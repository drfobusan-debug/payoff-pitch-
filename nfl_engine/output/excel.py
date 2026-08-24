"""The week as a workbook: the plays, every selection, and the record.

Four sheets, in the order they get read: **Plays** (what survived), **Selections**
(every row priced, vetoes included, because a rejected winner is only diagnosable
if it was written down), **Record** (the confusion-matrix splits) and **CLV**
(taken price against the close, which is the primary measurement while the graded
sample is small).

Written to bytes rather than to a path so the same object can be attached to an
email and saved, and so a test never needs a filesystem.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from nfl_engine.audit.ledger import ENGINE, LedgerEntry, Metrics
from nfl_engine.output.card import PAPER_NOTE, WeekCard

SELECTION_HEADER = (
    "Season",
    "Week",
    "Date",
    "Kickoff (UTC)",
    "Matchup",
    "Market",
    "Side",
    "Line",
    "Book",
    "Odds",
    "Opposite",
    "Tier",
    "Model %",
    "Fair %",
    "Model EV",
    "Exec EV",
    "Books paired",
    "Vetoes",
    "Result",
    "Units",
    "Close",
    "Close %",
    "CLV",
    "Close taken at",
    "Priced at",
    "Source",
    "Mode",
)

RECORD_HEADER = (
    "Split",
    "n",
    "Wins",
    "Losses",
    "Pushes",
    "Win %",
    "Break-even %",
    "PPV lift",
    "NPV lift",
    "ROI",
    "Units",
    "Mean CLV",
    "CLV beat %",
)

CLV_HEADER = (
    "Matchup",
    "Market",
    "Side",
    "Line",
    "Book",
    "Taken",
    "Close",
    "Close %",
    "CLV",
    "Close taken at",
    "Result",
)


def build_workbook(card: WeekCard, entries: list[LedgerEntry]) -> bytes:
    book = Workbook()
    _plays_sheet(book, card)
    _selections_sheet(book, entries, season=card.season, week=card.week)
    _record_sheet(book, card.record)
    _clv_sheet(book, entries)
    buffer = BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def _header(sheet, names: tuple[str, ...]) -> None:
    sheet.append(list(names))
    for index, name in enumerate(names, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")
        sheet.column_dimensions[get_column_letter(index)].width = max(9, min(len(name) + 4, 22))
    sheet.freeze_panes = "A2"


def _plays_sheet(book: Workbook, card: WeekCard) -> None:
    sheet = book.active
    sheet.title = "Plays"
    _header(
        sheet,
        (
            "Matchup",
            "Kickoff (UTC)",
            "Play",
            "Price",
            "Book",
            "Model %",
            "Fair %",
            "Exec EV",
            "Tier",
            "CLV",
            "Result",
        ),
    )
    for game in card.games:
        for play in game.plays:
            sheet.append(
                [
                    game.matchup,
                    game.kickoff,
                    play.label(),
                    play.odds,
                    play.book,
                    play.model_prob,
                    play.fair_prob,
                    play.ev_fair,
                    play.tier,
                    play.clv,
                    play.result,
                ]
            )
    sheet.append([])
    sheet.append([PAPER_NOTE])


def _selections_sheet(
    book: Workbook, entries: list[LedgerEntry], *, season: int, week: int
) -> None:
    sheet = book.create_sheet("Selections")
    _header(sheet, SELECTION_HEADER)
    scope = [e for e in entries if e.season == season and e.week == week]
    for entry in sorted(scope, key=lambda e: (e.matchup, e.market, -(e.ev_fair or 0.0))):
        sheet.append(
            [
                entry.season,
                entry.week,
                entry.date,
                entry.kickoff_utc,
                entry.matchup,
                entry.market,
                entry.side,
                entry.line,
                entry.book,
                entry.odds,
                entry.opposite_odds,
                entry.tier,
                entry.model_prob,
                entry.fair_prob,
                entry.ev_model,
                entry.ev_fair,
                entry.paired_books,
                entry.screens,
                entry.result,
                entry.pnl,
                entry.close_odds,
                entry.close_prob,
                entry.clv,
                entry.close_captured_at,
                entry.captured_at,
                entry.source,
                entry.mode,
            ]
        )


def _record_sheet(book: Workbook, record: list[Metrics]) -> None:
    sheet = book.create_sheet("Record")
    _header(sheet, RECORD_HEADER)
    for row in record:
        sheet.append(
            [
                row.label,
                row.n,
                row.wins,
                row.losses,
                row.pushes,
                row.win_pct,
                row.required_win_pct,
                row.ppv_lift,
                row.npv_lift,
                row.roi,
                row.units,
                row.mean_clv,
                row.clv_beat_pct,
            ]
        )


def _clv_sheet(book: Workbook, entries: list[LedgerEntry]) -> None:
    """Every engine row that has a closing price, worst move first.

    Sorted against us at the top on purpose: the rows the market moved away from
    are the ones worth reading, and a sheet sorted by our best result would bury
    them.
    """
    sheet = book.create_sheet("CLV")
    _header(sheet, CLV_HEADER)
    scored = [e for e in entries if e.clv is not None and e.source == ENGINE and not e.screens]
    for entry in sorted(scored, key=lambda e: e.clv or 0.0):
        sheet.append(
            [
                entry.matchup,
                entry.market,
                entry.side,
                entry.line,
                entry.book,
                entry.odds,
                entry.close_odds,
                entry.close_prob,
                entry.clv,
                entry.close_captured_at,
                entry.result,
            ]
        )
