"""Historical backtest harness: replay past slates and score model calibration.

The live engine turns model probabilities into EV/buy tiers using *market* prices.
A profitability (ROI) backtest therefore needs historical odds, which are a paid,
credit-heavy dependency. This module instead runs the free, odds-independent test
that answers the prior question -- *is the model predictive?* -- by replaying the
model over completed games and grading its own fixed-line probabilities against
the official box scores.

For each historical date we:
  * pull the actual slate + starting lineups (MLB Stats API),
  * rebuild every feature strictly as-of the day before (no look-ahead): Statcast
    rolling windows are sliced from a preloaded season frame, and the season-to-
    date Savant leaderboards are disabled,
  * emit the model's probability for each market at its own standard lines,
  * grade each pick win/loss/push against the final box score.

We then measure **calibration** (does "model says 62%" win ~62%?) via a reliability
table + Brier score, plus realized win% by market group. No wagering, no odds.
"""

from __future__ import annotations

import logging
from collections import defaultdict
from dataclasses import dataclass
from datetime import date as Date
from datetime import timedelta
from pathlib import Path

import pandas as pd

from mlb_engine.audit.grade import PUSH, WIN, grade
from mlb_engine.config import Config
from mlb_engine.data.mlb_statsapi import MLBStatsClient
from mlb_engine.data.parks import Park
from mlb_engine.data.results import GameResult, fetch_result
from mlb_engine.data.statcast import StatcastRepository
from mlb_engine.data.vsin import Split, VSINClient
from mlb_engine.filters.weather import WeatherEffect, WeatherProvider
from mlb_engine.market.ev import MarketQuote
from mlb_engine.pipeline import Pipeline, PipelineDeps
from mlb_engine.recommendations import Recommendation
from mlb_engine.schemas import Slate

log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Offline dependency stubs (no network, no look-ahead)
# ---------------------------------------------------------------------------
class SeasonStatcast(StatcastRepository):
    """Serve rolling windows by slicing one preloaded season frame in memory.

    Overriding ``load_range`` is enough: ``max_window``/``load_trailing`` call it
    with the trailing ``[start, end]`` window (ending the day before the slate),
    so every downstream feature stays correctly windowed as-of the game date.
    """

    # Numeric Statcast columns whose nullable (Int64/Float64) dtype must be
    # coerced to plain float so downstream ``float(series.mean())`` never sees a
    # pandas ``NAType`` (pybaseball's ``convert_dtypes`` otherwise leaks NA).
    _NUMERIC_COLS = (
        "release_speed", "inning", "strikes", "pfx_z", "launch_speed", "launch_angle",
        "launch_speed_angle", "hc_x", "hc_y", "zone", "estimated_woba_using_speedangle",
        "estimated_ba_using_speedangle", "woba_value", "woba_denom", "bat_speed",
        "swing_length", "release_pos_x", "release_pos_z", "release_extension",
        "release_spin_rate",
    )

    def __init__(self, frame: pd.DataFrame) -> None:
        f = frame.copy()
        for col in self._NUMERIC_COLS:
            if col in f.columns:
                f[col] = pd.to_numeric(f[col], errors="coerce").astype(float)
        self._frame = f

    def load_range(self, start: Date, end: Date, refresh: bool = False) -> pd.DataFrame:
        f = self._frame
        return f[(f["game_date"] >= start) & (f["game_date"] <= end)]


class NeutralWeather(WeatherProvider):
    """Weather-neutral stub (keeps the backtest offline and fast)."""

    def __init__(self) -> None:  # no session needed
        pass

    def fetch(self, park: Park, game_dt_utc: str | None) -> WeatherEffect:
        return WeatherEffect(None, 1.0, 1.0, note="backtest: weather neutral")


class NullVSIN(VSINClient):
    """VSIN stub returning no prices/splits (odds excluded from the backtest)."""

    def fetch(
        self, slate: Slate
    ) -> tuple[dict[tuple[str, str, str], list[MarketQuote]], dict[tuple[str, str, str], Split]]:
        return {}, {}


# ---------------------------------------------------------------------------
# Results
# ---------------------------------------------------------------------------
@dataclass
class CalibrationBin:
    lo: float
    hi: float
    n: int
    predicted: float  # mean model prob in the bin
    actual: float  # realized win rate in the bin (pushes excluded)


@dataclass
class GroupResult:
    group: str
    n: int
    wins: int
    losses: int
    pushes: int
    win_pct: float
    avg_model_prob: float
    brier: float
    bins: list[CalibrationBin]


def _group_of(rec: Recommendation) -> str:
    """Bucket a market into a human-facing group (reuses the display mapping)."""
    return rec.display_category


def _graded_rows(
    graded: list[tuple[Recommendation, str]],
) -> list[tuple[str, float, int]]:
    """Return ``(group, model_prob, won)`` per non-push pick (won in {0,1})."""
    rows: list[tuple[str, float, int]] = []
    for rec, result in graded:
        if result == PUSH:
            continue
        rows.append((_group_of(rec), rec.model_prob, 1 if result == WIN else 0))
    return rows


def _reliability(pairs: list[tuple[float, int]], n_bins: int = 10) -> list[CalibrationBin]:
    """Reliability table: bin by predicted prob, compare mean pred vs realized."""
    buckets: dict[int, list[tuple[float, int]]] = defaultdict(list)
    for prob, won in pairs:
        idx = min(int(prob * n_bins), n_bins - 1)
        buckets[idx].append((prob, won))
    out: list[CalibrationBin] = []
    for idx in range(n_bins):
        items = buckets.get(idx, [])
        if not items:
            continue
        preds = [p for p, _ in items]
        wins = [w for _, w in items]
        out.append(
            CalibrationBin(
                lo=idx / n_bins,
                hi=(idx + 1) / n_bins,
                n=len(items),
                predicted=sum(preds) / len(preds),
                actual=sum(wins) / len(wins),
            )
        )
    return out


def summarize(graded: list[tuple[Recommendation, str]]) -> list[GroupResult]:
    """Per-group + overall calibration/accuracy from graded picks."""
    rows = _graded_rows(graded)
    push_by_group: dict[str, int] = defaultdict(int)
    for rec, result in graded:
        if result == PUSH:
            push_by_group[_group_of(rec)] += 1

    by_group: dict[str, list[tuple[float, int]]] = defaultdict(list)
    for group, prob, won in rows:
        by_group[group].append((prob, won))

    def build(name: str, pairs: list[tuple[float, int]], pushes: int) -> GroupResult:
        n = len(pairs)
        wins = sum(w for _, w in pairs)
        brier = sum((p - w) ** 2 for p, w in pairs) / n if n else 0.0
        avg_p = sum(p for p, _ in pairs) / n if n else 0.0
        return GroupResult(
            group=name,
            n=n,
            wins=wins,
            losses=n - wins,
            pushes=pushes,
            win_pct=wins / n if n else 0.0,
            avg_model_prob=avg_p,
            brier=brier,
            bins=_reliability(pairs),
        )

    results = [build(g, by_group[g], push_by_group.get(g, 0)) for g in sorted(by_group)]
    results.append(build("ALL", rows_to_pairs(rows), sum(push_by_group.values())))
    return results


def rows_to_pairs(rows: list[tuple[str, float, int]]) -> list[tuple[float, int]]:
    return [(p, w) for _, p, w in rows]


# ---------------------------------------------------------------------------
# Confusion matrix / risk factors (odds-free, keyed on the model's own boundary)
# ---------------------------------------------------------------------------
@dataclass
class ConfusionResult:
    """Confusion matrix at a decision threshold on the model's probability.

    A *positive prediction* = the model favors the bet (``model_prob >= thr``);
    a *positive outcome* = the bet won. Pushes are excluded.
    """

    group: str
    threshold: float
    tp: int
    fp: int
    fn: int
    tn: int

    @property
    def n(self) -> int:
        return self.tp + self.fp + self.fn + self.tn

    @property
    def ppv(self) -> float:
        d = self.tp + self.fp
        return self.tp / d if d else 0.0

    @property
    def npv(self) -> float:
        d = self.tn + self.fn
        return self.tn / d if d else 0.0

    @property
    def sensitivity(self) -> float:
        d = self.tp + self.fn
        return self.tp / d if d else 0.0

    @property
    def specificity(self) -> float:
        d = self.tn + self.fp
        return self.tn / d if d else 0.0

    @property
    def fp_rate(self) -> float:
        """P(model favors it | it lost) -- false-positive rate."""
        d = self.fp + self.tn
        return self.fp / d if d else 0.0

    @property
    def fn_rate(self) -> float:
        """P(model fades it | it won) -- false-negative rate."""
        d = self.fn + self.tp
        return self.fn / d if d else 0.0


def confusion(
    graded: list[tuple[Recommendation, str]], threshold: float = 0.5
) -> list[ConfusionResult]:
    """Per-group + overall confusion matrix at ``threshold``."""
    counts: dict[str, list[int]] = defaultdict(lambda: [0, 0, 0, 0])  # tp, fp, fn, tn
    for rec, result in graded:
        if result == PUSH:
            continue
        won = result == WIN
        pred_pos = rec.model_prob >= threshold
        g = _group_of(rec)
        if pred_pos and won:
            counts[g][0] += 1
        elif pred_pos and not won:
            counts[g][1] += 1
        elif not pred_pos and won:
            counts[g][2] += 1
        else:
            counts[g][3] += 1

    def total() -> list[int]:
        acc = [0, 0, 0, 0]
        for c in counts.values():
            for i in range(4):
                acc[i] += c[i]
        return acc

    out = [
        ConfusionResult(g, threshold, c[0], c[1], c[2], c[3])
        for g, c in sorted(counts.items())
    ]
    t = total()
    out.append(ConfusionResult("ALL", threshold, t[0], t[1], t[2], t[3]))
    return out


def confidence_gap(graded: list[tuple[Recommendation, str]], threshold: float = 0.5) -> dict[str, float]:
    """Per-group over/under-confidence among *favored* picks (avg pred - win%).

    Positive => the model is over-confident on the bets it likes (a driver of
    false positives); negative => under-confident (a driver of false negatives).
    """
    fav: dict[str, list[tuple[float, int]]] = defaultdict(list)
    for rec, result in graded:
        if result == PUSH or rec.model_prob < threshold:
            continue
        fav[_group_of(rec)].append((rec.model_prob, 1 if result == WIN else 0))
    gaps: dict[str, float] = {}
    allp: list[tuple[float, int]] = []
    for g, pairs in fav.items():
        allp.extend(pairs)
        pred = sum(p for p, _ in pairs) / len(pairs)
        act = sum(w for _, w in pairs) / len(pairs)
        gaps[g] = pred - act
    if allp:
        gaps["ALL"] = (
            sum(p for p, _ in allp) / len(allp) - sum(w for _, w in allp) / len(allp)
        )
    return gaps


def _raw(rec: Recommendation) -> float:
    return rec.raw_prob if rec.raw_prob is not None else rec.model_prob


def false_positive_findings(
    graded: list[tuple[Recommendation, str]], min_n: int = 150, breakeven: float = 0.476
) -> list[str]:
    """Markets whose *favored* picks (prob>=0.5) lose more than breakeven.

    These are the commonalities driving the false-positive bias: the model likes
    them but they under-perform, so their EV/edge is phantom.
    """
    by_market: dict[str, list[int]] = defaultdict(list)
    over: list[int] = []
    under: list[int] = []
    for rec, result in graded:
        if result == PUSH or _raw(rec) < 0.5:
            continue
        lost = 0 if result == WIN else 1
        by_market[rec.market].append(lost)
        if rec.side == "over":
            over.append(lost)
        elif rec.side == "under":
            under.append(lost)
    out: list[str] = []
    for mk in sorted(by_market, key=lambda m: -sum(by_market[m]) / max(len(by_market[m]), 1)):
        o = by_market[mk]
        if len(o) < min_n:
            continue
        loss = sum(o) / len(o)
        if loss > 1 - breakeven:  # win rate below breakeven
            out.append(
                f"FALSE-POSITIVE RISK: {mk} favored picks lose {loss * 100:.1f}% "
                f"(win {100 - loss * 100:.1f}%, n={len(o)}) -- phantom edge, demoted by calibration"
            )
    if over and under:
        lo = sum(over) / len(over)
        lu = sum(under) / len(under)
        if lo - lu > 0.03:
            out.append(
                f"FALSE-POSITIVE RISK: over-side leans lose {lo * 100:.1f}% vs "
                f"under-side {lu * 100:.1f}% -- systematic OVER bias on props"
            )
    return out


def false_negative_findings(
    graded: list[tuple[Recommendation, str]], min_n: int = 150, breakeven: float = 0.524
) -> list[str]:
    """Faded bands (prob<0.5) that actually win above breakeven -- reclaimable."""
    band: dict[tuple[str, int], list[int]] = defaultdict(list)
    for rec, result in graded:
        if result == PUSH or _raw(rec) >= 0.5:
            continue
        band[(rec.market, int(_raw(rec) * 20))].append(1 if result == WIN else 0)
    out: list[str] = []
    for (mk, b), o in sorted(band.items()):
        if len(o) < min_n:
            continue
        wr = sum(o) / len(o)
        if wr > breakeven:
            out.append(
                f"RECLAIMABLE: {mk} faded band {b * 5}-{(b + 1) * 5}% actually wins "
                f"{wr * 100:.1f}% (n={len(o)}) -- calibration lifts it toward playable"
            )
    if not out:
        out.append("No sizable reclaimable false-negative pocket found -- fades are mostly correct.")
    return out


def risk_factors(
    groups: list[GroupResult],
    conf: list[ConfusionResult],
    gaps: dict[str, float],
) -> list[str]:
    """Plain-language risk findings on PPV/NPV and false-positive/negative bias."""
    out: list[str] = []
    by_conf = {c.group: c for c in conf}
    for gr in groups:
        c = by_conf.get(gr.group)
        if c is None or c.n < 30:
            continue
        gap = gaps.get(gr.group, 0.0)
        bias = (
            "false-positive bias (over-confident on bets it likes)"
            if c.fp_rate > c.fn_rate + 0.05
            else "false-negative bias (fades winners)"
            if c.fn_rate > c.fp_rate + 0.05
            else "balanced FP/FN"
        )
        cal = (
            f"over-confident by {gap * 100:.1f} pts"
            if gap > 0.03
            else f"under-confident by {abs(gap) * 100:.1f} pts"
            if gap < -0.03
            else "well-calibrated"
        )
        out.append(
            f"{gr.group}: PPV {c.ppv:.2f} / NPV {c.npv:.2f} | "
            f"FP-rate {c.fp_rate:.2f} vs FN-rate {c.fn_rate:.2f} -> {bias}; "
            f"favored picks {cal}; Brier {gr.brier:.3f}"
        )
    return out


# ---------------------------------------------------------------------------
# Runner
# ---------------------------------------------------------------------------
def sample_dates(start: Date, end: Date, step_days: int) -> list[Date]:
    """Every ``step_days``-th date in ``[start, end]`` (inclusive)."""
    out: list[Date] = []
    d = start
    while d <= end:
        out.append(d)
        d += timedelta(days=step_days)
    return out


def _make_pipeline(cfg: Config, frame: pd.DataFrame) -> Pipeline:
    deps = PipelineDeps(
        stats=MLBStatsClient(),
        statcast=SeasonStatcast(frame),
        weather=NeutralWeather(),
        vsin=NullVSIN(cfg.creds),
        oddsapi=None,
        rotowire=None,
        fangraphs=None,
    )
    return Pipeline(cfg, deps)


def run_backtest(
    cfg: Config,
    frame: pd.DataFrame,
    dates: list[Date],
    stats: MLBStatsClient | None = None,
) -> list[tuple[Recommendation, str]]:
    """Replay ``dates`` through the model and grade every pick vs final scores."""
    pipe = _make_pipeline(cfg, frame)
    client = stats or MLBStatsClient()
    graded: list[tuple[Recommendation, str]] = []
    for d in dates:
        try:
            recs = pipe.run(d, enrich_leaderboards=False)
        except Exception as exc:  # keep the backtest going across bad days
            log.warning("slate %s failed: %s", d, exc)
            continue
        if not recs:
            continue
        results: dict[int, GameResult] = {}
        for pk in {r.game_pk for r in recs}:
            try:
                results[pk] = fetch_result(pk, session=client.session)
            except Exception as exc:  # noqa: BLE001
                log.warning("result %s failed: %s", pk, exc)
        day_graded = 0
        for r in recs:
            res = results.get(r.game_pk)
            if res is None or not res.final:
                continue
            g = grade(r, res)
            if g is not None:
                graded.append((r, g))
                day_graded += 1
        log.info("Backtest %s: %d model picks, %d graded", d, len(recs), day_graded)
    return graded


def load_season_frame(cache_dir: Path, start: Date, end: Date) -> pd.DataFrame:
    """Load the cached season Statcast frame (downloading once if needed)."""
    return StatcastRepository(cache_dir).load_range(start, end)


# ---------------------------------------------------------------------------
# Workbook
# ---------------------------------------------------------------------------
def write_backtest_workbook(
    groups: list[GroupResult],
    conf: list[ConfusionResult],
    gaps: dict[str, float],
    findings: list[str],
    out_path: Path,
) -> Path:
    """Write a backtest report: Summary, Calibration, Confusion, Risk Factors."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")

    def style_header(ws, ncol: int) -> None:
        for c in range(1, ncol + 1):
            cell = ws.cell(1, c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    ws = wb.active
    ws.title = "Summary"
    cols = ["Market group", "N", "Wins", "Losses", "Pushes", "Win %", "Avg model %", "Brier", "Calib. gap (pts)"]
    ws.append(cols)
    for g in groups:
        ws.append([
            g.group, g.n, g.wins, g.losses, g.pushes, round(g.win_pct * 100, 1),
            round(g.avg_model_prob * 100, 1), round(g.brier, 4),
            round(gaps.get(g.group, 0.0) * 100, 1),
        ])
    style_header(ws, len(cols))

    wc = wb.create_sheet("Confusion (PPV-NPV-FP-FN)")
    cols = ["Market group", "N", "TP", "FP", "FN", "TN", "PPV", "NPV",
            "Sensitivity", "Specificity", "FP rate", "FN rate"]
    wc.append(cols)
    for c in conf:
        wc.append([
            c.group, c.n, c.tp, c.fp, c.fn, c.tn, round(c.ppv, 3), round(c.npv, 3),
            round(c.sensitivity, 3), round(c.specificity, 3),
            round(c.fp_rate, 3), round(c.fn_rate, 3),
        ])
    style_header(wc, len(cols))

    wcal = wb.create_sheet("Calibration")
    cols = ["Market group", "Prob band", "N", "Predicted %", "Actual win %", "Gap (pts)"]
    wcal.append(cols)
    for g in groups:
        for b in g.bins:
            wcal.append([
                g.group, f"{int(b.lo * 100)}-{int(b.hi * 100)}%", b.n,
                round(b.predicted * 100, 1), round(b.actual * 100, 1),
                round((b.predicted - b.actual) * 100, 1),
            ])
    style_header(wcal, len(cols))

    wr = wb.create_sheet("Risk Factors")
    wr.append(["Finding"])
    for line in findings:
        wr.append([line])
    style_header(wr, 1)
    wr.column_dimensions["A"].width = 120

    wb.save(out_path)
    return out_path
