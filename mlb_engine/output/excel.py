"""Write daily recommendations to a formatted Excel workbook.

Sheets:
  * Strong Buys   : strong-tier picks, categorized, neon-green -> fade spectrum.
  * Moderate Buys : moderate-tier picks, categorized, neon-yellow -> fade.
  * Fades         : the picks the model is against, neon-red -> fade.
  * Moneyline     : game family (ML / run line / totals), colored by tier.
  * First-5 (F5)  : first-five family, colored by tier.
  * Pitcher Props : pitcher family, colored by tier.
  * Batter Props  : batter family, colored by tier.
  * All           : every priced market, colored by tier.

Every sheet uses one color language: neon green fading to pale for strong buys,
neon yellow fading to pale for moderate buys, and neon red fading to pale for
fades. The most intense (neon) end marks the highest-conviction plays; a
``BEST`` flag highlights the standout buys.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date as Date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from mlb_engine.audit.analysis import (
    FALSE_NEGATIVE,
    FALSE_POSITIVE,
    TRUE_POSITIVE,
    PropInsight,
)
from mlb_engine.audit.clv import ClvSummary
from mlb_engine.audit.ledger import LedgerEntry, OverallMetrics
from mlb_engine.config import EVThresholds
from mlb_engine.market.tiers import Tier
from mlb_engine.recommendations import Recommendation

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)

COLUMNS = [
    "Date",
    "Matchup",
    "Category",
    "Market",
    "Selection",
    "Line",
    "Model %",
    "Market %",
    "Fair Odds",
    "Book",
    "Book Odds",
    "EV",
    "Edge",
    "Handle %",
    "Bets %",
    "Tier",
    "Signal",
    "Factor",
    "Score",
    "Profile",
    "Opta %",
    "AI",
    "VSiN",
    "VSiN Pick",
    "VSiN Edge",
    "Notes",
]

TIER_ORDER = {Tier.STRONG.value: 0, Tier.MODERATE.value: 1, Tier.PASS.value: 2}

# --- categorized sheet layout ----------------------------------------------

CATEGORY_ORDER = [
    "Moneyline",
    "Totals",
    "Run Lines",
    "First-5 (F5)",
    "Batter Props",
    "Pitcher Props",
    "Comeback (info)",
]
_GAME_CATEGORIES = {"Moneyline", "Totals", "Run Lines", "First-5 (F5)"}
_PROP_CATEGORIES = {"Batter Props", "Pitcher Props"}

# Per-game "family" tabs, keyed on Recommendation.category. These partition every
# priced market, so nothing is lost across the four family sheets.
FAMILY_TABS: list[tuple[str, str]] = [
    ("Moneyline", "game"),
    ("First-5 (F5)", "f5"),
    ("Pitcher Props", "pitcher"),
    ("Batter Props", "batter"),
]

# Compact grid shared by the tier tabs, the Fades tab and the family tabs.
GRID_COLUMNS = [
    "Best",
    "Tier",
    "Category",
    "EV",
    "Selection",
    "Matchup",
    "Date",
    "Book",
    "Odds",
    "Model %",
    "Market %",
    "AI",
    "Opta %",
    "VSiN",
    "VSiN Pick",
    "Edge",
    "Handle %",
    "Bets %",
    "Signal",
    "Factor",
    "Score",
    "Profile",
    "Notes",
]
GRID_WIDTHS = [
    7, 13, 15, 8, 30, 13, 12, 12, 8, 8, 8, 7, 8, 6, 20, 8, 9, 8, 8, 7, 7, 26, 40
]
GRID_CENTER = {
    "Best", "Tier", "EV", "Date", "Odds", "Model %", "Market %", "AI", "Opta %",
    "VSiN", "Edge", "Handle %", "Bets %",
}

# Scheme keys.
_STRONG = "strong"
_MODERATE = "moderate"
_FADE = "fade"
_SCHEME_RANK = {_STRONG: 0, _MODERATE: 1, _FADE: 2}


# neon (high-conviction) -> fade (pale) spectrums, one per class.
@dataclass(frozen=True)
class _Scheme:
    best: str  # BEST-row fill
    best_font: str  # BEST-row font color
    light: tuple[int, int, int]  # faded, low-conviction end of the gradient
    neon: tuple[int, int, int]  # neon, high-conviction end of the gradient
    header: str
    border: str


_SCHEMES: dict[str, _Scheme] = {
    _STRONG: _Scheme(
        best="00C853",
        best_font="00320F",
        light=(0xDE, 0xFF, 0xE7),
        neon=(0x00, 0xE6, 0x76),
        header="0B3D1E",
        border="9BE7B8",
    ),
    _MODERATE: _Scheme(
        best="FFD600",
        best_font="332B00",
        light=(0xFF, 0xFD, 0xD6),
        neon=(0xFF, 0xEA, 0x00),
        header="6B5600",
        border="FFEE7A",
    ),
    _FADE: _Scheme(
        best="FF1744",
        best_font="FFFFFF",
        light=(0xFF, 0xDA, 0xDA),
        neon=(0xFF, 0x17, 0x44),
        header="5A0A16",
        border="FFAEB8",
    ),
}


def _scheme_key(rec: Recommendation) -> str:
    """Green for strong buys, yellow for moderate, red for everything faded."""
    if rec.tier == Tier.STRONG:
        return _STRONG
    if rec.tier == Tier.MODERATE:
        return _MODERATE
    return _FADE


def _conviction(rec: Recommendation, key: str) -> float:
    """Sort/shade magnitude: distance from the no-vig price, in either direction.

    Both ends are probability points, so a shade means the same thing on a dog and
    on chalk; EV would make the gradient a function of price length.
    """
    if key == _FADE:
        if rec.fair_prob is not None:
            v = rec.fair_prob - rec.model_prob
        elif rec.ev is not None:
            v = -rec.ev
        else:
            v = 0.0
        return max(0.0, v)
    if rec.edge is not None:
        return rec.edge
    return rec.ev if rec.ev is not None else 0.0


def _is_best(rec: Recommendation, category: str) -> bool:
    """Flag the standout buys: the most edge over the price, not the most EV.

    Keyed off edge for the same reason the tiers are: ``EV = decimal_odds x edge``,
    so an EV bar picks out long prices rather than strong opinions. Props keep the
    price window, which is a separate judgement about which prop prices are worth
    taking at all.
    """
    edge = rec.edge
    odds = rec.market_american
    if edge is None or odds is None:
        return False
    thr = EVThresholds().for_market(rec.market)
    standout = thr.min_edge + 2 * thr.strong_edge_gap
    if category in _GAME_CATEGORIES:
        return edge >= standout
    if category in _PROP_CATEGORIES:
        return edge >= standout and -200 <= odds <= 120
    return False


def _american(odds: float | None) -> str:
    if odds is None:
        return "n/a"
    o = round(odds)
    return f"+{o}" if o > 0 else str(o)


def _interp(light: tuple[int, int, int], neon: tuple[int, int, int], t: float) -> str:
    rgb = tuple(int(light[i] + (neon[i] - light[i]) * t) for i in range(3))
    return f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"


def _grid_values(rec: Recommendation, cat: str, best: bool) -> dict[str, object]:
    return {
        "Best": "BEST" if best else "",
        "Tier": rec.tier.value,
        "Category": cat,
        "EV": round(rec.ev, 3) if rec.ev is not None else "",
        "Selection": rec.selection,
        "Matchup": rec.matchup,
        "Date": rec.game_date.isoformat(),
        "Book": rec.book or "",
        "Odds": _american(rec.market_american),
        "Model %": round(rec.model_prob * 100, 1),
        "Market %": round(rec.fair_prob * 100, 1) if rec.fair_prob is not None else "",
        "AI": rec.opta_mark,
        "Opta %": round(rec.opta_prob * 100, 1) if rec.opta_prob is not None else "",
        "VSiN": rec.vsin_mark,
        "VSiN Pick": rec.vsin_pick or "",
        "Edge": round(rec.edge, 3) if rec.edge is not None else "",
        "Handle %": rec.handle_pct if rec.handle_pct is not None else "",
        "Bets %": rec.bets_pct if rec.bets_pct is not None else "",
        "Signal": rec.signal or "",
        "Factor": round(rec.factor, 3) if rec.factor is not None else "",
        "Score": round(rec.score, 2) if rec.score is not None else "",
        "Profile": rec.profile or "",
        "Notes": "; ".join(rec.reasons),
    }


def _row_style(
    rec: Recommendation, key: str, best: bool, t: float
) -> tuple[PatternFill, Font | None, Border]:
    scheme = _SCHEMES[key]
    if best:
        fill = PatternFill("solid", fgColor=scheme.best)
        font: Font | None = Font(color=scheme.best_font, bold=True)
    else:
        fill = PatternFill(
            "solid", fgColor=_interp(scheme.light, scheme.neon, 0.15 + 0.7 * t)
        )
        font = None
    side = Side(style="thin", color=scheme.border)
    return fill, font, Border(left=side, right=side, top=side, bottom=side)


def _write_grid(
    ws: Worksheet, recs: list[Recommendation], header_key: str | None = None
) -> None:
    """Categorized grid; rows shaded neon->fade by class (green/yellow/red)."""
    header_fill = (
        PatternFill("solid", fgColor=_SCHEMES[header_key].header)
        if header_key
        else HEADER_FILL
    )
    center = Alignment(horizontal="center")

    cat_rank = {c: i for i, c in enumerate(CATEGORY_ORDER)}
    tagged = [(r, r.display_category, _scheme_key(r)) for r in recs]
    tagged.sort(
        key=lambda t: (
            _SCHEME_RANK[t[2]],
            cat_rank.get(t[1], len(CATEGORY_ORDER)),
            0 if _is_best(t[0], t[1]) else 1,
            -_conviction(t[0], t[2]),
        )
    )

    # conviction range per class, for the intra-class gradient.
    conv_range: dict[str, tuple[float, float]] = {}
    for rec, _cat, key in tagged:
        conv = _conviction(rec, key)
        lo, hi = conv_range.get(key, (conv, conv))
        conv_range[key] = (min(lo, conv), max(hi, conv))

    for c, name in enumerate(GRID_COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill = header_fill
        cell.font = HEADER_FONT
        cell.alignment = center

    for r, (rec, cat, key) in enumerate(tagged, start=2):
        best = key in (_STRONG, _MODERATE) and _is_best(rec, cat)
        lo, hi = conv_range[key]
        conv = _conviction(rec, key)
        t = 0.5 if hi == lo else (conv - lo) / (hi - lo)
        fill, font, border = _row_style(rec, key, best, t)
        values = _grid_values(rec, cat, best)
        for c, name in enumerate(GRID_COLUMNS, start=1):
            cell = ws.cell(row=r, column=c, value=values[name])
            cell.fill = fill
            cell.border = border
            if font is not None:
                cell.font = font
            if name in GRID_CENTER:
                cell.alignment = center

    for i, w in enumerate(GRID_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    if tagged:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(GRID_COLUMNS))}{len(tagged) + 1}"
    ws.freeze_panes = "A2"


def _write_all_sheet(ws: Worksheet, recs: list[Recommendation]) -> None:
    """Full-detail sheet, every row shaded by class (neon->fade)."""
    for c, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    tagged = [(r, _scheme_key(r)) for r in recs]
    conv_range: dict[str, tuple[float, float]] = {}
    for rec, key in tagged:
        conv = _conviction(rec, key)
        lo, hi = conv_range.get(key, (conv, conv))
        conv_range[key] = (min(lo, conv), max(hi, conv))

    for r, (rec, key) in enumerate(tagged, start=2):
        best = key in (_STRONG, _MODERATE) and _is_best(rec, rec.display_category)
        lo, hi = conv_range[key]
        conv = _conviction(rec, key)
        t = 0.5 if hi == lo else (conv - lo) / (hi - lo)
        fill, font, _border = _row_style(rec, key, best, t)
        row = rec.as_row()
        for c, name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=r, column=c, value=row[name])
            cell.fill = fill
            if font is not None:
                cell.font = font

    widths = [11, 12, 9, 14, 26, 6, 8, 9, 9, 11, 10, 8, 8, 9, 8, 12, 8, 7, 7, 30, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    if recs:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(recs) + 1}"
    ws.freeze_panes = "A2"


def write_workbook(recs: list[Recommendation], out_path: Path, slate_date: Date) -> Path:
    def sort_key(r: Recommendation) -> tuple[int, float]:
        return (TIER_ORDER.get(r.tier.value, 3), -(r.ev if r.ev is not None else -99))

    wb = Workbook()
    strong = [r for r in recs if r.tier == Tier.STRONG]
    moderate = [r for r in recs if r.tier == Tier.MODERATE]
    fades = [r for r in recs if r.tier == Tier.PASS]

    ws_strong = wb.active
    ws_strong.title = "Strong Buys"
    _write_grid(ws_strong, strong, header_key=_STRONG)

    _write_grid(wb.create_sheet("Moderate Buys"), moderate, header_key=_MODERATE)
    _write_grid(wb.create_sheet("Fades"), fades, header_key=_FADE)

    # Family tabs: buys and fades together, each row keeping its own class color.
    for title, cat in FAMILY_TABS:
        family = [r for r in recs if r.category == cat]
        _write_grid(wb.create_sheet(title), family)

    _write_all_sheet(wb.create_sheet("All"), sorted(recs, key=sort_key))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


_METRIC_HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
_RESULT_FILL = {
    "win": PatternFill("solid", fgColor="C6EFCE"),
    "loss": PatternFill("solid", fgColor="FFC7CE"),
    "push": PatternFill("solid", fgColor="E7E6E6"),
}
_OVERALL_COLUMNS = [
    "Tier",
    "N",
    "Wins",
    "Losses",
    "Pushes",
    "Win %",
    "PPV",
    "NPV",
    "Sensitivity",
    "Specificity",
    "Needs %",
    "ROI",
    "Units",
]
_DAILY_COLUMNS = [
    "Date",
    "Buy N",
    "Wins",
    "Losses",
    "Win %",
    "ROI",
    "Units",
]
_BET_COLUMNS = [
    "Date",
    "Category",
    "Selection",
    "Matchup",
    "Book",
    "Odds",
    "Tier",
    "Model %",
    "Market %",
    "EV",
    "Close",
    "CLV",
    "CLV EV",
    "Result",
    "P/L",
]


def _pct(v: float) -> str:
    return f"{v * 100:.1f}%"


def _metric_row(m: OverallMetrics) -> list[object]:
    return [
        m.tier,
        m.n,
        m.wins,
        m.losses,
        m.pushes,
        _pct(m.win_pct),
        m.ppv,
        m.npv,
        m.sensitivity,
        m.specificity,
        _pct(m.required_win_pct),
        _pct(m.roi),
        m.units,
    ]


_INSIGHT_COLUMNS = ["Type", "Market", "N", "Rate", "Recommendation"]
_INSIGHT_FILL = {
    FALSE_POSITIVE: PatternFill("solid", fgColor="FFC7CE"),
    FALSE_NEGATIVE: PatternFill("solid", fgColor="FFEB9C"),
    TRUE_POSITIVE: PatternFill("solid", fgColor="C6EFCE"),
}
_INSIGHT_LABEL = {
    FALSE_POSITIVE: "False positive",
    FALSE_NEGATIVE: "False negative",
    TRUE_POSITIVE: "True positive",
}


def _write_metric_header(ws: Worksheet, columns: list[str]) -> None:
    for c, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill = _METRIC_HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _write_metric_sheet(
    ws: Worksheet, rows: list[OverallMetrics], label_header: str = "Tier"
) -> None:
    cols = [label_header, *_OVERALL_COLUMNS[1:]]
    _write_metric_header(ws, cols)
    for r, m in enumerate(rows, start=2):
        for c, v in enumerate(_metric_row(m), start=1):
            ws.cell(row=r, column=c, value=v)
    for i, w in enumerate([14, 6, 6, 7, 7, 8, 7, 7, 12, 12, 9, 8, 8], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows) + 1}"
    ws.freeze_panes = "A2"


def _write_insight_sheet(ws: Worksheet, insights: list[PropInsight]) -> None:
    _write_metric_header(ws, _INSIGHT_COLUMNS)
    order = {FALSE_POSITIVE: 0, FALSE_NEGATIVE: 1, TRUE_POSITIVE: 2}
    ordered = sorted(insights, key=lambda i: (order.get(i.kind, 9), i.market))
    for r, ins in enumerate(ordered, start=2):
        vals = [
            _INSIGHT_LABEL.get(ins.kind, ins.kind),
            ins.market,
            ins.n,
            _pct(ins.rate),
            ins.finding,
        ]
        for c, v in enumerate(vals, start=1):
            ws.cell(row=r, column=c, value=v)
        fill = _INSIGHT_FILL.get(ins.kind)
        if fill:
            ws.cell(row=r, column=1).fill = fill
    for i, w in enumerate([15, 14, 6, 8, 100], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    if ordered:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(_INSIGHT_COLUMNS))}{len(ordered) + 1}"
    ws.freeze_panes = "A2"


_CLV_COLUMNS = ["Market", "N", "Mean CLV", "Beat close %", "Mean CLV EV"]
_CLV_FILL = {
    True: PatternFill("solid", fgColor="C6EFCE"),
    False: PatternFill("solid", fgColor="FFC7CE"),
}


def _write_clv_sheet(ws: Worksheet, rows: list[ClvSummary]) -> None:
    _write_metric_header(ws, _CLV_COLUMNS)
    for r, m in enumerate(rows, start=2):
        vals: list[object] = [
            m.label,
            m.n,
            f"{m.mean_clv * 100:+.2f}",
            _pct(m.beat_close_pct),
            f"{m.mean_clv_ev:+.4f}",
        ]
        for c, v in enumerate(vals, start=1):
            ws.cell(row=r, column=c, value=v)
        ws.cell(row=r, column=5).fill = _CLV_FILL[m.positive]
    for i, w in enumerate([16, 7, 11, 13, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def write_ledger_workbook(
    entries: list[LedgerEntry],
    overall: list[OverallMetrics],
    daily: list[OverallMetrics],
    out_path: Path,
    daily_engine: list[OverallMetrics] | None = None,
    prop_rows: list[OverallMetrics] | None = None,
    insights: list[PropInsight] | None = None,
    runline_rows: list[OverallMetrics] | None = None,
    clv_rows: list[ClvSummary] | None = None,
) -> Path:
    wb = Workbook()

    ws_overall = wb.active
    ws_overall.title = "Overall"
    _write_metric_sheet(ws_overall, overall)

    if daily_engine:
        _write_metric_sheet(wb.create_sheet("Daily PPV-NPV"), daily_engine, "Date")

    if prop_rows:
        _write_metric_sheet(wb.create_sheet("Prop PPV-NPV"), prop_rows, "Prop market")

    if runline_rows:
        _write_metric_sheet(wb.create_sheet("Run Line NPV"), runline_rows, "Run line / gate")

    if clv_rows:
        _write_clv_sheet(wb.create_sheet("Closing Line Value"), clv_rows)

    if insights:
        _write_insight_sheet(wb.create_sheet("Prop Insights"), insights)

    ws_daily = wb.create_sheet("Daily")
    _write_metric_header(ws_daily, _DAILY_COLUMNS)
    for r, m in enumerate(daily, start=2):
        vals = [m.tier, m.n, m.wins, m.losses, _pct(m.win_pct), _pct(m.roi), m.units]
        for c, v in enumerate(vals, start=1):
            ws_daily.cell(row=r, column=c, value=v)
    for i, w in enumerate([12, 8, 7, 8, 8, 8, 8], start=1):
        ws_daily.column_dimensions[get_column_letter(i)].width = w
    if daily:
        ws_daily.auto_filter.ref = f"A1:{get_column_letter(len(_DAILY_COLUMNS))}{len(daily) + 1}"
    ws_daily.freeze_panes = "A2"

    ws_bets = wb.create_sheet("Bets")
    _write_metric_header(ws_bets, _BET_COLUMNS)
    for r, e in enumerate(entries, start=2):
        vals = [
            e.date,
            e.category,
            e.selection,
            e.matchup,
            e.book,
            _american(e.odds),
            e.tier,
            round(e.model_prob * 100, 1),
            round(e.fair_prob * 100, 1) if e.fair_prob is not None else "",
            e.ev if e.ev is not None else "",
            _american(e.close_odds),
            round(e.clv * 100, 2) if e.clv is not None else "",
            round(e.clv_ev, 3) if e.clv_ev is not None else "",
            e.result,
            round(e.pnl, 3),
        ]
        for c, v in enumerate(vals, start=1):
            ws_bets.cell(row=r, column=c, value=v)
        fill = _RESULT_FILL.get(e.result)
        if fill:
            ws_bets.cell(row=r, column=_BET_COLUMNS.index("Result") + 1).fill = fill
    for i, w in enumerate([12, 15, 30, 13, 13, 8, 10, 8, 9, 8, 8, 8, 8, 8, 8], start=1):
        ws_bets.column_dimensions[get_column_letter(i)].width = w
    if entries:
        ws_bets.auto_filter.ref = f"A1:{get_column_letter(len(_BET_COLUMNS))}{len(entries) + 1}"
    ws_bets.freeze_panes = "A2"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
