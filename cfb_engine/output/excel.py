"""Write daily recommendations to a formatted Excel workbook.

Sheets:
  * Strong Buys / Moderate Buys : the buys, brightest = highest EV.
  * Fades                       : the sides the model is against.
  * Moneyline / ATS / Totals    : one tab per market family, every priced side.
  * All                         : every priced market.

One color language across sheets: green for Strong, yellow for Moderate, red
for Fades, each fading from bright (high conviction) to pale.
"""

from __future__ import annotations

from datetime import date as Date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from cfb_engine.audit.clv import ClvSummary
from cfb_engine.audit.ledger import LedgerEntry, OverallMetrics
from cfb_engine.audit.priced import PricedStat
from cfb_engine.audit.probation import Probation
from cfb_engine.market.tiers import Tier
from cfb_engine.recommendations import Recommendation

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CENTER = Alignment(horizontal="center")

COLUMNS = [
    "Date", "Matchup", "Market", "Selection", "Line",
    "Model %", "Market %", "Fair Odds", "Book", "Book Odds",
    "EV", "Edge", "Tier", "Notes",
]
WIDTHS = [11, 22, 13, 22, 7, 8, 8, 9, 12, 10, 8, 8, 12, 46]
CENTER_COLS = {"Line", "Model %", "Market %", "Fair Odds", "Book Odds", "EV", "Edge", "Tier"}

TIER_ORDER = {Tier.STRONG.value: 0, Tier.MODERATE.value: 1, Tier.PASS.value: 2}
_MARKET_TABS = [("Moneyline", "game_ml"), ("ATS", "game_ats"), ("Totals", "game_total")]

# bright (high conviction) -> pale, one gradient per tier.
_SCHEME = {
    Tier.STRONG: ((0xDE, 0xFF, 0xE7), (0x00, 0xE6, 0x76), "0B3D1E"),
    Tier.MODERATE: ((0xFF, 0xFD, 0xD6), (0xFF, 0xEA, 0x00), "6B5600"),
    Tier.PASS: ((0xFF, 0xDA, 0xDA), (0xFF, 0x17, 0x44), "5A0A16"),
}


def _interp(light: tuple[int, int, int], neon: tuple[int, int, int], t: float) -> str:
    rgb = tuple(int(light[i] + (neon[i] - light[i]) * t) for i in range(3))
    return f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"


def _conviction(rec: Recommendation) -> float:
    if rec.tier == Tier.PASS:
        if rec.fair_prob is not None:
            return max(0.0, rec.fair_prob - rec.model_prob)
        return max(0.0, -(rec.ev or 0.0))
    return rec.ev if rec.ev is not None else 0.0


def _write_sheet(ws: Worksheet, recs: list[Recommendation], header: str | None = None) -> None:
    header_fill = PatternFill("solid", fgColor=header) if header else HEADER_FILL
    for c, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill = header_fill
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    ordered = sorted(
        recs, key=lambda r: (TIER_ORDER.get(r.tier.value, 3), -_conviction(r))
    )
    # per-tier conviction range for the intra-tier gradient.
    ranges: dict[Tier, tuple[float, float]] = {}
    for r in ordered:
        conv = _conviction(r)
        lo, hi = ranges.get(r.tier, (conv, conv))
        ranges[r.tier] = (min(lo, conv), max(hi, conv))

    for row_idx, rec in enumerate(ordered, start=2):
        light, neon, _ = _SCHEME[rec.tier]
        lo, hi = ranges[rec.tier]
        t = 0.5 if hi == lo else (_conviction(rec) - lo) / (hi - lo)
        fill = PatternFill("solid", fgColor=_interp(light, neon, 0.15 + 0.7 * t))
        row = rec.as_row()
        for c, name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=c, value=row[name])
            cell.fill = fill
            if name in CENTER_COLS:
                cell.alignment = CENTER

    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    if ordered:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(ordered) + 1}"
    ws.freeze_panes = "A2"


def write_workbook(recs: list[Recommendation], out_path: Path, slate_date: Date) -> Path:
    wb = Workbook()
    strong = [r for r in recs if r.tier == Tier.STRONG]
    moderate = [r for r in recs if r.tier == Tier.MODERATE]
    fades = [r for r in recs if r.tier == Tier.PASS]

    ws = wb.active
    ws.title = "Strong Buys"
    _write_sheet(ws, strong, header=_SCHEME[Tier.STRONG][2])
    _write_sheet(wb.create_sheet("Moderate Buys"), moderate, header=_SCHEME[Tier.MODERATE][2])
    _write_sheet(wb.create_sheet("Fades"), fades, header=_SCHEME[Tier.PASS][2])
    for title, market in _MARKET_TABS:
        _write_sheet(wb.create_sheet(title), [r for r in recs if r.market == market])
    _write_sheet(wb.create_sheet("All"), recs)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


# --- audit ledger workbook -------------------------------------------------
_METRIC_HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
_RESULT_FILL = {
    "win": PatternFill("solid", fgColor="C6EFCE"),
    "loss": PatternFill("solid", fgColor="FFC7CE"),
    "push": PatternFill("solid", fgColor="E7E6E6"),
}
_OVERALL_COLUMNS = [
    "Tier", "N", "Wins", "Losses", "Pushes", "Win %", "PPV", "NPV",
    "Sensitivity", "Specificity", "Needs %", "ROI", "Units",
]
_DAILY_COLUMNS = ["Date", "Buy N", "Wins", "Losses", "Win %", "ROI", "Units"]
_BET_COLUMNS = [
    "Date", "Category", "Selection", "Matchup", "Book", "Odds", "Tier",
    "Model %", "Market %", "EV", "Close", "CLV", "CLV EV", "Result", "P/L",
]
_CLV_COLUMNS = ["Market", "N", "Mean CLV", "Beat close %", "Mean CLV EV"]
_MONEY_COLUMNS = [
    "Market", "N", "Won", "Win%", "Needs", "Gap", "ROI", "Units",
    "Two-sided", "One-way", "One-way units", "CLV", "Beat close", "Beat number",
]
_PROBATION_COLUMNS = [
    "Verdict", "What", "Kind", "N", "ROI", "SE", "1st half", "2nd half", "Finding",
]
_CLV_FILL = {
    True: PatternFill("solid", fgColor="C6EFCE"),
    False: PatternFill("solid", fgColor="FFC7CE"),
}
_VERDICT_FILL = {
    "SHUT": PatternFill("solid", fgColor="FFC7CE"),
    "LIFT": PatternFill("solid", fgColor="FFC7CE"),
    "SHIP": PatternFill("solid", fgColor="FFEB9C"),
}


def _pct(v: float) -> str:
    return f"{v * 100:.1f}%"


def _nan_pct(v: float) -> str:
    """Percent, or a dash where the sample is empty rather than zero."""
    return "-" if v != v else f"{v * 100:.1f}%"


def _american(odds: float | None) -> str:
    if odds is None:
        return "n/a"
    o = round(odds)
    return f"+{o}" if o > 0 else str(o)


def _metric_header(ws: Worksheet, columns: list[str]) -> None:
    for c, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill = _METRIC_HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER


def _metric_row(m: OverallMetrics) -> list[object]:
    return [
        m.tier, m.n, m.wins, m.losses, m.pushes, _pct(m.win_pct), m.ppv, m.npv,
        m.sensitivity, m.specificity, _pct(m.required_win_pct), _pct(m.roi), m.units,
    ]


def _write_metric_sheet(
    ws: Worksheet, rows: list[OverallMetrics], label_header: str = "Tier"
) -> None:
    cols = [label_header, *_OVERALL_COLUMNS[1:]]
    _metric_header(ws, cols)
    for r, m in enumerate(rows, start=2):
        for c, v in enumerate(_metric_row(m), start=1):
            ws.cell(row=r, column=c, value=v)
    for i, w in enumerate([16, 6, 6, 7, 7, 8, 7, 7, 12, 12, 9, 8, 8], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows) + 1}"
    ws.freeze_panes = "A2"


def _write_clv_sheet(ws: Worksheet, rows: list[ClvSummary]) -> None:
    _metric_header(ws, _CLV_COLUMNS)
    for r, m in enumerate(rows, start=2):
        vals: list[object] = [
            m.label, m.n, f"{m.mean_clv * 100:+.2f}", _pct(m.beat_close_pct),
            f"{m.mean_clv_ev:+.4f}",
        ]
        for c, v in enumerate(vals, start=1):
            ws.cell(row=r, column=c, value=v)
        ws.cell(row=r, column=5).fill = _CLV_FILL[m.positive]
    for i, w in enumerate([18, 7, 11, 13, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _write_money_sheet(ws: Worksheet, rows: list[PricedStat]) -> None:
    """The realized money record: only bets that were priced and bought.

    Sits beside the PPV sheets because a win rate is not a return: ``Needs`` is
    the rate the prices charged and ``Gap`` is what decides the sign of ROI.
    """
    _metric_header(ws, _MONEY_COLUMNS)
    for r, s in enumerate(rows, start=2):
        vals: list[object] = [
            s.label, s.n, s.wins, _nan_pct(s.win_rate), _nan_pct(s.breakeven),
            _nan_pct(s.shortfall), _nan_pct(s.roi), s.units,
            s.two_sided, s.n_one_way, s.units_one_way,
            "-" if s.clv != s.clv else f"{s.clv * 100:+.2f}",
            _nan_pct(s.clv_rate), _nan_pct(s.number_rate),
        ]
        for c, v in enumerate(vals, start=1):
            ws.cell(row=r, column=c, value=v)
        if s.n:
            ws.cell(row=r, column=7).fill = _CLV_FILL[s.roi >= 0]
    for i, w in enumerate([16, 6, 6, 8, 8, 8, 8, 8, 10, 9, 13, 8, 11, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _write_probation_sheet(ws: Worksheet, rows: list[Probation]) -> None:
    """Every verdict, including the WATCHING ones: an empty table would read as
    a clean bill of health when it actually means too few graded bets."""
    _metric_header(ws, _PROBATION_COLUMNS)
    for r, p in enumerate(rows, start=2):
        vals: list[object] = [
            p.status, p.name, p.kind, p.n, _nan_pct(p.roi), f"{p.se * 100:.1f}",
            _nan_pct(p.first_half), _nan_pct(p.second_half), p.finding,
        ]
        for c, v in enumerate(vals, start=1):
            ws.cell(row=r, column=c, value=v)
        fill = _VERDICT_FILL.get(p.status)
        if fill:
            ws.cell(row=r, column=1).fill = fill
    for i, w in enumerate([10, 30, 11, 6, 8, 7, 9, 9, 90], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def write_ledger_workbook(
    entries: list[LedgerEntry],
    overall: list[OverallMetrics],
    daily: list[OverallMetrics],
    out_path: Path,
    market_rows: list[OverallMetrics] | None = None,
    clv_rows: list[ClvSummary] | None = None,
    price_rows: list[OverallMetrics] | None = None,
    money_rows: list[PricedStat] | None = None,
    probation_rows: list[Probation] | None = None,
) -> Path:
    wb = Workbook()
    ws_overall = wb.active
    ws_overall.title = "Overall"
    _write_metric_sheet(ws_overall, overall)

    if money_rows:
        _write_money_sheet(wb.create_sheet("Money (priced buys)"), money_rows)
    if probation_rows:
        _write_probation_sheet(wb.create_sheet("Probation"), probation_rows)
    if market_rows:
        _write_metric_sheet(wb.create_sheet("By Market"), market_rows, "Market")
    if price_rows:
        _write_metric_sheet(wb.create_sheet("By Price"), price_rows, "Price band")
    if clv_rows:
        _write_clv_sheet(wb.create_sheet("Closing Line Value"), clv_rows)

    ws_daily = wb.create_sheet("Daily")
    _metric_header(ws_daily, _DAILY_COLUMNS)
    for r, m in enumerate(daily, start=2):
        vals = [m.tier, m.n, m.wins, m.losses, _pct(m.win_pct), _pct(m.roi), m.units]
        for c, v in enumerate(vals, start=1):
            ws_daily.cell(row=r, column=c, value=v)
    for i, w in enumerate([12, 8, 7, 8, 8, 8, 8], start=1):
        ws_daily.column_dimensions[get_column_letter(i)].width = w
    ws_daily.freeze_panes = "A2"

    ws_bets = wb.create_sheet("Bets")
    _metric_header(ws_bets, _BET_COLUMNS)
    for r, e in enumerate(entries, start=2):
        vals = [
            e.date, e.category, e.selection, e.matchup, e.book, _american(e.odds), e.tier,
            round(e.model_prob * 100, 1),
            round(e.fair_prob * 100, 1) if e.fair_prob is not None else "",
            e.ev if e.ev is not None else "",
            _american(e.close_odds),
            round(e.clv * 100, 2) if e.clv is not None else "",
            round(e.clv_ev, 3) if e.clv_ev is not None else "",
            e.result, round(e.pnl, 3),
        ]
        for c, v in enumerate(vals, start=1):
            ws_bets.cell(row=r, column=c, value=v)
        fill = _RESULT_FILL.get(e.result)
        if fill:
            ws_bets.cell(row=r, column=_BET_COLUMNS.index("Result") + 1).fill = fill
    for i, w in enumerate([12, 14, 24, 22, 12, 8, 12, 8, 9, 8, 8, 8, 8, 8, 8], start=1):
        ws_bets.column_dimensions[get_column_letter(i)].width = w
    ws_bets.freeze_panes = "A2"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
