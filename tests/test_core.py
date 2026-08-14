"""Unit tests for the model, market, and audit layers (no network)."""

from __future__ import annotations

from datetime import date

import numpy as np

from mlb_engine.audit.grade import LOSS, PUSH, WIN, grade
from mlb_engine.audit.scorecard import FIELDS, append_scorecard, build_scorecard
from mlb_engine.config import EVThresholds
from mlb_engine.data.results import GameResult, PlayerLine, _ip_to_outs
from mlb_engine.features.rolling import LEAGUE_RATES, OutcomeRates, rates_from_events
from mlb_engine.market.ev import MarketQuote, ev_per_dollar, evaluate
from mlb_engine.market.odds import (
    american_to_decimal,
    american_to_prob,
    no_vig_two_way,
    prob_to_american,
)
from mlb_engine.market.tiers import Tier, classify
from mlb_engine.models.markov_f5 import (
    f5_from_lineups,
    f5_from_rates,
    team_f5_distribution,
)
from mlb_engine.models.matchup import apply_multipliers, combine
from mlb_engine.models.montecarlo import MonteCarlo, TeamSimConfig
from mlb_engine.models.rbi_rule import evaluate_lineup, rbi_multiplier
from mlb_engine.recommendations import Recommendation


def _league_rates() -> OutcomeRates:
    r = LEAGUE_RATES
    return OutcomeRates(500, r["1B"], r["2B"], r["3B"], r["HR"], r["BB"], r["K"], r["OUT"])


# ---- odds math ----
def test_odds_roundtrip():
    for a in (-200, -110, 100, 150, 250):
        p = american_to_prob(a)
        assert 0 < p < 1
    assert abs(american_to_decimal(100) - 2.0) < 1e-9
    assert abs(american_to_decimal(-200) - 1.5) < 1e-9


def test_prob_to_american_inverse():
    for p in (0.4, 0.55, 0.7):
        a = prob_to_american(p)
        assert abs(american_to_prob(a) - p) < 1e-6


def test_no_vig_sums_to_one():
    a, b = no_vig_two_way(-110, -110)
    assert abs(a + b - 1.0) < 1e-9
    assert abs(a - 0.5) < 1e-9


# ---- matchup ----
def test_combine_normalizes():
    lg = _league_rates()
    out = combine(lg, lg)
    assert abs(sum(out.values()) - 1.0) < 1e-9
    # combining league vs league should return ~league rates
    assert abs(out["HR"] - LEAGUE_RATES["HR"]) < 1e-3


def test_apply_multipliers_renormalizes():
    lg = _league_rates()
    base = combine(lg, lg)
    out = apply_multipliers(base, {"HR": 2.0})
    assert abs(sum(out.values()) - 1.0) < 1e-9
    assert out["HR"] > base["HR"]


# ---- markov F5 ----
def test_f5_reasonable():
    lg = LEAGUE_RATES
    r = f5_from_rates(dict(lg), dict(lg))
    mean = sum(i * p for i, p in enumerate(r.home_dist))
    assert 1.5 < mean < 3.5  # ~2.4 runs/team through 5
    assert abs(r.p_home_ml + r.p_away_ml + r.p_tie - 1.0) < 1e-6
    assert abs(sum(r.total_dist) - 1.0) < 1e-6


def test_f5_lineup_matches_convolution_without_tto():
    lg = dict(LEAGUE_RATES)
    conv = f5_from_rates(lg, lg)
    dp = team_f5_distribution([lg] * 9, tto_factors=(1.0, 1.0, 1.0, 1.0))
    m_conv = sum(i * p for i, p in enumerate(conv.home_dist))
    m_dp = sum(i * p for i, p in enumerate(dp))
    assert abs(m_conv - m_dp) < 1e-4


def test_f5_tto_raises_scoring():
    lg = dict(LEAGUE_RATES)
    base = sum(i * p for i, p in enumerate(team_f5_distribution([lg] * 9, (1.0, 1.0, 1.0, 1.0))))
    tto = sum(i * p for i, p in enumerate(team_f5_distribution([lg] * 9)))
    assert tto > base
    r = f5_from_lineups([lg] * 9, [lg] * 9)
    assert abs(r.p_home_ml + r.p_away_ml + r.p_tie - 1.0) < 1e-6


# ---- weather (WAM park-config filter) ----
def test_weather_park_config_gates_wind():
    from mlb_engine.data.parks import get_park
    from mlb_engine.filters.weather import WeatherConditions, _effect

    out = WeatherConditions(85, 50, 15, 0, 15)  # 15 mph straight out to CF
    wrigley = _effect(out, get_park(17))  # open bowl, wind-receptive
    oracle = _effect(out, get_park(2395))  # shielded
    assert wrigley > 1.15  # wind reaches the field -> big HR boost
    assert oracle < wrigley  # architecture suppresses the same wind

    blow_in = WeatherConditions(85, 50, 15, 180, -15)
    wrigley_in = _effect(blow_in, get_park(17))
    assert wrigley_in < 1.0  # in-from-CF suppresses power


# ---- monte carlo ----
def test_montecarlo_runs():
    lg = LEAGUE_RATES
    bat = [dict(lg) for _ in range(9)]
    cfg = TeamSimConfig(bat_vs_starter=bat, bat_vs_pen=bat)
    res = MonteCarlo(400, seed=1).simulate(cfg, cfg)
    assert res.home_runs_full.shape == (400,)
    assert 3.0 < res.home_runs_full.mean() < 6.5
    # home never trails-and-bats: full >= f5
    assert (res.home_runs_full >= res.home_runs_f5).all()


# ---- rbi rule ----
def _profile(rates: OutcomeRates):
    from mlb_engine.features.rolling import BatterProfile

    return BatterProfile(
        mlbam_id=1, home=rates, away=rates, vs_rhp=rates, vs_lhp=rates, overall=rates
    )


def test_rbi_rule_flags_high_obp():
    hi = OutcomeRates(200, 0.18, 0.06, 0.005, 0.05, 0.12, 0.18, 0.40)  # obp high
    lo = OutcomeRates(200, 0.10, 0.02, 0.0, 0.01, 0.04, 0.30, 0.53)  # obp low
    flags = evaluate_lineup([_profile(hi)] * 9)
    assert all(f.flagged for f in flags)
    assert rbi_multiplier(flags[0]) > 1.0
    flags2 = evaluate_lineup([_profile(lo)] * 9)
    assert not any(f.flagged for f in flags2)
    assert rbi_multiplier(flags2[0]) == 1.0


def _reg(xslg: float, zone_contact: float):
    from mlb_engine.features.regression import BatterRegression

    return BatterRegression(
        bbe=40, barrel_rate=0.08, hard_hit=0.40, sweet_spot=0.33, bat_speed=72.0,
        max_ev=108.0, whiff=0.24, zone_contact=zone_contact, xba=0.25, xslg=xslg,
        babip=0.29, woba=0.32, xwoba=0.32,
    )


def test_rbi_ppv_npv_tiers():
    hi = OutcomeRates(200, 0.18, 0.06, 0.005, 0.05, 0.12, 0.18, 0.40)
    profs = [_profile(hi)] * 9
    # PPV: elite xSLG with runners on boosts RBI above the volume-only boost.
    elite = evaluate_lineup(profs, regressions=[_reg(0.520, 0.82)] * 9)
    vol_only = evaluate_lineup(profs)
    assert rbi_multiplier(elite[0]) > rbi_multiplier(vol_only[0])
    # NPV: in-zone contact collapse caps RBI despite big opportunity.
    collapse = evaluate_lineup(profs, regressions=[_reg(0.400, 0.60)] * 9)
    assert rbi_multiplier(collapse[0]) < rbi_multiplier(vol_only[0])


# ---- bullpen vs batter ----
def test_bullpen_profile_excludes_starter_and_uses_late_innings():
    import pandas as pd

    from mlb_engine.features.rolling import build_bullpen_profile

    gd = date(2024, 7, 10)
    rows = [
        # starter (NYY, pitching in Top as home team): 1st + carries into 6th
        (gd, 100, 1, "Top", "single"),
        (gd, 100, 6, "Top", "single"),
        # relievers in late innings -> heavy strikeouts
        (gd, 200, 7, "Top", "strikeout"),
        (gd, 200, 8, "Top", "strikeout"),
        (gd, 201, 9, "Top", "strikeout"),
        # noise: another team's pitcher in NYY's batting half (Bot, home=NYY) - excluded
        (gd, 300, 8, "Bot", "single"),
    ]
    df = pd.DataFrame(
        rows, columns=["game_date", "pitcher", "inning", "inning_topbot", "events"]
    )
    df["batter"] = 1
    df["home_team"] = "NYY"
    df["away_team"] = "BOS"

    pen = build_bullpen_profile(df, "NYY", date(2024, 7, 19), 21, min_inning=6)
    # relievers were all strikeouts -> pen K rate well above league (0.225),
    # and the starter's late single must not leak in.
    assert pen.allowed.p_k > LEAGUE_RATES["K"]


def _relief_frame(rows):
    import pandas as pd

    df = pd.DataFrame(
        rows, columns=["game_date", "pitcher", "inning", "inning_topbot", "events"]
    )
    df["batter"] = 1
    df["home_team"] = "NYY"
    df["away_team"] = "BOS"
    return df


def test_bullpen_leverage_split_isolates_late_arms():
    from mlb_engine.features.rolling import build_bullpen_profile

    gd = date(2024, 7, 10)
    rows = (
        # 7th-inning middle relief gives up singles (dilutes the aggregate)
        [(gd, 201, 7, "Top", "single") for _ in range(30)]
        # 8th/9th high-leverage arms strike everyone out
        + [(gd, 200, 8, "Top", "strikeout") for _ in range(20)]
        + [(gd, 202, 9, "Top", "strikeout") for _ in range(10)]
    )
    pen = build_bullpen_profile(_relief_frame(rows), "NYY", date(2024, 7, 19), 21, min_inning=6)
    # 8th+ arms (all Ks) grade far better than the single-diluted 6th+ aggregate.
    assert pen.allowed_leverage.p_k > pen.allowed.p_k


def test_bullpen_leverage_falls_back_when_thin():
    from mlb_engine.features.rolling import build_bullpen_profile

    gd = date(2024, 7, 10)
    rows = (
        [(gd, 201, 7, "Top", "single") for _ in range(30)]
        # only 5 8th-inning PAs: below MIN_LEVERAGE_PA -> no separate profile
        + [(gd, 200, 8, "Top", "strikeout") for _ in range(5)]
    )
    pen = build_bullpen_profile(_relief_frame(rows), "NYY", date(2024, 7, 19), 21, min_inning=6)
    assert pen.allowed_leverage.p_k == pen.allowed.p_k


def test_leverage_pen_suppresses_close_game_scoring():
    lg = LEAGUE_RATES
    bat = [dict(lg) for _ in range(9)]
    hot = {"1B": 0.5, "2B": 0.0, "3B": 0.0, "HR": 0.0, "BB": 0.0, "K": 0.0, "OUT": 0.5}
    cold = {"1B": 0.0, "2B": 0.0, "3B": 0.0, "HR": 0.0, "BB": 0.0, "K": 0.5, "OUT": 0.5}
    hot_pen = [dict(hot) for _ in range(9)]
    cold_pen = [dict(cold) for _ in range(9)]
    # Early hook so the pen pitches most of the game.
    base = dict(bat_vs_starter=bat, starter_bf_cap=9, starter_pitch_cap=200)
    no_split = TeamSimConfig(bat_vs_pen=hot_pen, **base)
    with_split = TeamSimConfig(bat_vs_pen=hot_pen, bat_vs_pen_close=cold_pen, **base)
    res_no = MonteCarlo(600, seed=7).simulate(no_split, no_split)
    res_yes = MonteCarlo(600, seed=7).simulate(with_split, with_split)
    # Shutdown leverage arms in the close innings pull scoring down sharply.
    assert res_yes.home_runs_full.mean() < res_no.home_runs_full.mean()


def test_bullpen_npv_walk_trap_and_fatigue():
    import pandas as pd

    from mlb_engine.features.rolling import BullpenProfile

    lg = rates_from_events(pd.Series(dtype=object))
    empty = pd.DataFrame()
    # Walk trap: zone% below .40 -> BB boosted.
    trap = BullpenProfile(lg, lg, empty, zone_pct=0.34, recent_load=1.0)
    assert trap.npv_multipliers().get("BB", 1.0) > 1.0
    # Fatigue: heavy recent workload -> HR/hits boosted.
    tired = BullpenProfile(lg, lg, empty, zone_pct=0.50, recent_load=1.4)
    assert tired.npv_multipliers().get("HR", 1.0) > 1.0
    # Rotowire availability override (rested) suppresses the fatigue penalty.
    assert "HR" not in tired.npv_multipliers(availability=1.0)


# ---- fielding defense ----
def test_defense_hit_multiplier_bounds_and_direction():
    from mlb_engine.filters.defense import defense_hit_multiplier

    # Elite defense (+FRV) suppresses BIP hits (<1); poor defense inflates (>1).
    assert defense_hit_multiplier(20.0) < 1.0
    assert defense_hit_multiplier(-20.0) > 1.0
    assert defense_hit_multiplier(0.0) == 1.0
    # Bounded to +/-6%.
    assert defense_hit_multiplier(9999.0) >= 0.94
    assert defense_hit_multiplier(-9999.0) <= 1.06


def test_team_defense_positional_and_npv():
    from mlb_engine.filters.defense import TeamDefense

    # Team FRV fallback suppresses both singles and XBH.
    good = TeamDefense(frv=40.0).bip_multipliers()
    assert good["1B"] < 1.0 and good["2B"] < 1.0 and good["3B"] < 1.0
    assert set(good) == {"1B", "2B", "3B"}  # never touches K/BB/HR

    # Per-position OAA: elite OF range hits XBH harder than singles.
    of = TeamDefense(outfield_oaa=15.0, infield_oaa=0.0).bip_multipliers()
    assert of["2B"] < of["1B"]

    # NPV: broken middle infield (SS+2B OAA < -8) leaks grounder singles.
    broken = TeamDefense(middle_if_oaa=-15.0).bip_multipliers()
    assert broken["1B"] > 1.0

    # Neutral profile -> no-op multipliers.
    neutral = TeamDefense().bip_multipliers()
    assert neutral["1B"] == 1.0 and neutral["2B"] == 1.0


# ---- human element + umpire ----
def test_human_divisional_and_umpire_zone():
    from mlb_engine.filters.human import HumanFactors
    from mlb_engine.filters.umpire import zone_runs_for_name

    # Divisional familiarity -> fewer Ks, a touch more contact.
    div = HumanFactors(divisional=True).offense_multipliers()
    assert div["K"] < 1.0
    assert div.get("1B", 1.0) > 1.0

    # Tight-zone "Over" ump (Moscoso) -> fewer Ks, more walks.
    over = HumanFactors(umpire_zone_runs=zone_runs_for_name("Edwin Moscoso")).offense_multipliers()
    assert over["K"] < 1.0
    assert over["BB"] > 1.0

    # Wide-zone "Under" ump (Barrett) -> more Ks, fewer walks.
    under = HumanFactors(umpire_zone_runs=zone_runs_for_name("Lance Barrett")).offense_multipliers()
    assert under["K"] > 1.0
    assert under["BB"] < 1.0

    # Unknown umpire -> neutral.
    assert zone_runs_for_name("Nobody McUnknown") is None


# ---- manager tendencies ----
def test_no_manager_sets_the_starter_hook() -> None:
    """The hook is measured per start, so no manager may carry a cap.

    The hand-entered caps ranged 19-29 batters faced across five managers; the
    measured p75 over 3,299 starts ranges 23-26 across all thirty teams, and
    correlated with the entered value at r = +0.22. Los Angeles had the longest
    leash in baseball and was entered as the third quickest.
    """
    from mlb_engine.data.managers import MANAGERS, ManagerProfile

    fields = set(ManagerProfile.__dataclass_fields__)
    assert "starter_bf_cap" not in fields
    assert "starter_pitch_cap" not in fields
    for tid in (139, 141, 119, 113, 137):  # TB, TOR, LAD, CIN, SF
        assert tid not in MANAGERS


def test_manager_platoon_and_speed():
    from mlb_engine.data.managers import get_manager

    # Unknown team -> neutral default, no tilts.
    neutral = get_manager(999999)
    assert neutral.offense_multipliers() == {}
    assert neutral.pen_multipliers() == {}
    # Platoon maximizer (Baldelli, MIN=142) tilts only the bullpen matchup.
    assert get_manager(142).pen_multipliers().get("K", 1.0) < 1.0
    # Speed engine (Vogt, CLE=114) boosts advancement + lowers K.
    speed = get_manager(114).offense_multipliers()
    assert speed.get("2B", 1.0) > 1.0 and speed.get("K", 1.0) < 1.0


def test_the_third_time_through_window_reads_measured_depth() -> None:
    """The comeback signal must stay reachable now no manager reports 26+ BF."""
    from mlb_engine.models.comeback import TTTO_LONG_LEASH, ComebackSignal
    from mlb_engine.models.comeback import evaluate as evaluate_comeback

    deep = evaluate_comeback(ComebackSignal(opp_starter_bf_cap=TTTO_LONG_LEASH))
    short = evaluate_comeback(ComebackSignal(opp_starter_bf_cap=TTTO_LONG_LEASH - 1))
    assert deep.score > short.score
    assert any("long-leash" in r for r in deep.reasons)
    # Reachable from a real starter: the measured p90 start is 27 batters faced.
    assert TTTO_LONG_LEASH <= 27


# ---- VSIN public splits -> moneyline quotes ----
def test_vsin_fetch_quotes_maps_to_slate():
    import datetime

    from mlb_engine.config import Credentials
    from mlb_engine.data.vsin import VSINClient
    from mlb_engine.schemas import Game, Slate, TeamGameInfo, Venue

    def team(tid, name, ab, home):
        return TeamGameInfo(team_id=tid, name=name, abbrev=ab, is_home=home)

    game = Game(
        game_pk=1, game_date=datetime.date(2026, 7, 20), status="Preview",
        venue=Venue(venue_id=1, name="x"),
        home=team(114, "Cleveland Guardians", "CLE", True),
        away=team(142, "Minnesota Twins", "MIN", False),
    )
    slate = Slate(slate_date=datetime.date(2026, 7, 20), games=[game])

    client = VSINClient(Credentials())

    from mlb_engine.data.vsin import _RawRow

    def fake_fetch_book(src):
        # away row (Twins) carries the Over; home row (Guardians) the Under.
        return [
            _RawRow("Minnesota Twins", -1.5, 96.0, 42.0, 7.5, 13.0, 61.0, -131.0, 84.0, 62.0),
            _RawRow("Cleveland Guardians", 1.5, 4.0, 58.0, 7.5, 87.0, 39.0, 109.0, 16.0, 38.0),
            _RawRow("Unlisted Team", -1.5, 50.0, 50.0, 8.5, 50.0, 50.0, -120.0, 50.0, 50.0),
        ]

    client._fetch_book = fake_fetch_book  # type: ignore[method-assign]
    quotes, splits = client.fetch(slate)

    # two books -> two priced ML quotes per matched selection; unlisted team dropped.
    assert set(quotes) == {
        ("MIN @ CLE", "game_ml", "MIN ML"),
        ("MIN @ CLE", "game_ml", "CLE ML"),
    }
    min_q = quotes[("MIN @ CLE", "game_ml", "MIN ML")]
    assert len(min_q) == 2  # DK + circa
    assert min_q[0].american == -131.0
    assert min_q[0].handle_pct == 84.0 and min_q[0].bets_pct == 62.0

    # run-line + total handle/bets surface as splits (no price).
    assert splits[("MIN @ CLE", "game_rl", "MIN -1.5")].handle_pct == 96.0
    assert splits[("MIN @ CLE", "game_rl", "CLE +1.5")].bets_pct == 58.0
    assert splits[("MIN @ CLE", "game_total", "Over 7.5")].handle_pct == 13.0
    assert splits[("MIN @ CLE", "game_total", "Under 7.5")].handle_pct == 87.0
    assert client.fetch_quotes(slate).keys() == quotes.keys()

    # An alternate line is the same public money: VSIN posts the split against
    # its own line, but the engine routinely picks 8.5 or 9.5 on the same game,
    # and keying on the full selection dropped the split on every one of them.
    from mlb_engine.data.vsin import lookup_split

    assert lookup_split(splits, "MIN @ CLE", "game_total", "Over 9.5").handle_pct == 13.0
    assert lookup_split(splits, "MIN @ CLE", "game_total", "Under 8.5").handle_pct == 87.0
    assert lookup_split(splits, "MIN @ CLE", "game_rl", "MIN -2.5").handle_pct == 96.0
    # The exact line still wins when it is present, and an uncovered game or a
    # market VSIN never reports stays absent rather than borrowing a side.
    assert lookup_split(splits, "MIN @ CLE", "game_total", "Over 7.5").handle_pct == 13.0
    assert lookup_split(splits, "TB @ ATH", "game_total", "Over 8.5") is None
    assert lookup_split(splits, "MIN @ CLE", "batter_h", "Jose Ramirez H o0.5") is None


def test_split_side_ignores_the_line():
    from mlb_engine.data.vsin import split_side

    assert split_side("game_total", "Over 7.5") == "Over"
    assert split_side("game_total", "Under 9.5") == "Under"
    assert split_side("game_rl", "BAL -1.5") == "BAL"
    assert split_side("game_ml", "LAD ML") == "LAD"
    # Props have no public split anywhere, so there is no side to fall back to.
    assert split_side("batter_h", "Aaron Judge H o1.5") is None
    assert split_side("game_total", "") is None


def test_circa_weighted_consensus_and_divergence():
    from mlb_engine.market.ev import MarketQuote, evaluate

    # DK soft (+120), Circa sharp (-110); Circa's higher implied prob should
    # dominate the weighted no-vig fair estimate.
    dk = MarketQuote(book="draftkings", american=120.0, handle_pct=40.0, bets_pct=55.0)
    circa = MarketQuote(book="circa", american=-110.0, handle_pct=70.0, bets_pct=48.0)
    res = evaluate(0.5, [dk, circa])
    # best price for EV payout is the +120 DK line.
    assert res.best_quote.book == "draftkings"
    # weighted fair sits closer to Circa's implied prob than a 50/50 blend.
    from mlb_engine.market.odds import american_to_prob
    even = (american_to_prob(120.0) + american_to_prob(-110.0)) / 2
    assert res.fair_prob > even
    # divergence weights Circa's +22 heavier than DK's -15 -> net positive.
    assert res.sharp_divergence is not None and res.sharp_divergence > 0


def test_runline_sharp_money_bump():
    from mlb_engine.market.runline import RunLineSignal, runline_adjustment

    sig = RunLineSignal(sharp_money_side="home")
    steps, reasons = runline_adjustment("home", -1.5, sig)
    assert steps >= 1 and any("sharp money" in r for r in reasons)
    steps2, _ = runline_adjustment("away", -1.5, sig)
    assert steps2 == 0


def test_oddsapi_maps_game_f5_and_props():
    import datetime

    from mlb_engine.data.oddsapi import OddsAPIClient
    from mlb_engine.schemas import Game, Slate, TeamGameInfo, Venue

    def team(tid, name, ab, home):
        return TeamGameInfo(team_id=tid, name=name, abbrev=ab, is_home=home)

    game = Game(
        game_pk=1, game_date=datetime.date(2026, 7, 20), status="Preview",
        venue=Venue(venue_id=1, name="x"),
        home=team(114, "Cleveland Guardians", "CLE", True),
        away=team(142, "Minnesota Twins", "MIN", False),
    )
    slate = Slate(slate_date=datetime.date(2026, 7, 20), games=[game])

    bulk = [{
        "id": "evt1", "home_team": "Cleveland Guardians", "away_team": "Minnesota Twins",
        "bookmakers": [{"key": "draftkings", "markets": [
            {"key": "h2h", "outcomes": [
                {"name": "Minnesota Twins", "price": -130},
                {"name": "Cleveland Guardians", "price": 110}]},
            {"key": "spreads", "outcomes": [
                {"name": "Minnesota Twins", "price": 105, "point": -1.5},
                {"name": "Cleveland Guardians", "price": -125, "point": 1.5}]},
            {"key": "totals", "outcomes": [
                {"name": "Over", "price": -110, "point": 8.5},
                {"name": "Under", "price": -110, "point": 8.5}]},
        ]}],
    }]
    event = {
        "id": "evt1", "home_team": "Cleveland Guardians", "away_team": "Minnesota Twins",
        "bookmakers": [{"key": "fanduel", "markets": [
            {"key": "h2h_1st_5_innings", "outcomes": [
                {"name": "Minnesota Twins", "price": -120}]},
            {"key": "totals_1st_5_innings", "outcomes": [
                {"name": "Over", "price": 100, "point": 4.5}]},
            {"key": "batter_hits", "outcomes": [
                {"name": "Over", "description": "Byron Buxton", "price": -150, "point": 0.5},
                {"name": "Under", "description": "Byron Buxton", "price": 120, "point": 0.5}]},
            {"key": "pitcher_strikeouts", "outcomes": [
                {"name": "Over", "description": "Pablo Lopez", "price": -115, "point": 5.5}]},
        ]}],
    }

    client = OddsAPIClient("k")

    def fake_get(url, **params):
        return event if "/events/" in url else bulk

    client._get_json = fake_get  # type: ignore[method-assign]
    q = client.fetch(slate)

    assert ("MIN @ CLE", "game_ml", "MIN ML") in q
    assert ("MIN @ CLE", "game_rl", "MIN -1.5") in q
    assert ("MIN @ CLE", "game_total", "Over 8.5") in q
    assert ("MIN @ CLE", "f5_ml", "MIN F5 ML") in q
    assert ("MIN @ CLE", "f5_total", "F5 Over 4.5") in q
    assert ("MIN @ CLE", "batter_h", "Byron Buxton H o0.5") in q
    assert ("MIN @ CLE", "pitcher_k", "Pablo Lopez Ks o5.5") in q
    assert q[("MIN @ CLE", "game_ml", "MIN ML")][0].american == -130.0
    # Both sides of a prop are priced, so a prop can be passed on its merits
    # rather than for want of a price on the side the model prefers.
    assert ("MIN @ CLE", "batter_h", "Byron Buxton H u0.5") in q
    assert q[("MIN @ CLE", "batter_h", "Byron Buxton H u0.5")][0].american == 120.0


def test_merge_quotes_dedupes_by_book():
    from mlb_engine.data.oddsapi import Quotes  # noqa: F401
    from mlb_engine.market.ev import MarketQuote
    from mlb_engine.pipeline import _merge_quotes

    key = ("MIN @ CLE", "game_ml", "MIN ML")
    odds = {key: [MarketQuote(book="draftkings", american=-130.0)]}
    vsin = {key: [
        MarketQuote(book="draftkings", american=-131.0, handle_pct=80.0, bets_pct=60.0),
        MarketQuote(book="circa", american=-126.0, handle_pct=70.0, bets_pct=50.0),
    ]}
    merged = _merge_quotes(odds, vsin)
    books = sorted(q.book for q in merged[key])
    assert books == ["circa", "draftkings"]  # DK kept once (from Odds API)
    dk = next(q for q in merged[key] if q.book == "draftkings")
    assert dk.american == -130.0  # primary (Odds API) wins


_ROTO_HTML = """
<div class="lineup__box">
  <div class="lineup__teams">
    <div class="lineup__team is-visit"><div class="lineup__abbr">MIN</div></div>
    <div class="lineup__team is-home"><div class="lineup__abbr">CLE</div></div>
  </div>
  <a class="lineup__matchup">
    <div class="lineup__mteam is-visit">Twins <span class="lineup__wl">(1-2)</span></div>
    <div class="lineup__mteam is-home">Guardians <span class="lineup__wl">(2-1)</span></div>
  </a>
  <ul class="lineup__list is-visit">
    <li class="lineup__player-highlight">
      <div class="lineup__player-highlight-name">
        <a href="/baseball/player/joe-ryan-16036">Joe Ryan</a>
        <span class="lineup__throws">R</span>
      </div>
    </li>
    <li class="lineup__status is-expected">Expected Lineup</li>
    <li class="lineup__player"><div class="lineup__pos">LF</div>
      <a title="Trevor Larnach" href="/x">T. Larnach</a>
      <span class="lineup__bats">L</span></li>
    <li class="lineup__player"><div class="lineup__pos">DH</div>
      <a title="Byron Buxton" href="/x">Byron Buxton</a>
      <span class="lineup__bats">R</span></li>
  </ul>
  <ul class="lineup__list is-home">
    <li class="lineup__player-highlight">
      <div class="lineup__player-highlight-name">
        <a href="/baseball/player/tanner-bibee-1">Tanner Bibee</a>
        <span class="lineup__throws">R</span>
      </div>
    </li>
    <li class="lineup__status is-confirmed">Confirmed Lineup</li>
    <li class="lineup__player"><div class="lineup__pos">2B</div>
      <a title="Andres Gimenez" href="/x">A. Gimenez</a>
      <span class="lineup__bats">L</span></li>
  </ul>
</div>
"""


def test_rotowire_parse_daily_lineups():
    from mlb_engine.data.rotowire import norm_person, parse_daily_lineups

    games = parse_daily_lineups(_ROTO_HTML)
    assert len(games) == 1
    g = games[0]
    assert g.away.abbrev == "MIN" and g.away.nickname == "Twins"
    assert g.home.abbrev == "CLE" and g.home.nickname == "Guardians"
    assert g.away.pitcher == "Joe Ryan" and g.away.pitcher_throws == "R"
    assert g.away.confirmed is False and g.home.confirmed is True
    assert g.away.batters[0].name == "Trevor Larnach"
    assert g.away.batters[0].position == "LF" and g.away.batters[0].bats == "L"
    assert norm_person("José Ramírez Jr.") == "jose ramirez"


def test_match_roto_game_by_nickname():
    import datetime

    from mlb_engine.data.rotowire import RotoGame, RotoLineup
    from mlb_engine.pipeline import _match_roto_game
    from mlb_engine.schemas import Game, TeamGameInfo, Venue

    game = Game(
        game_pk=1, game_date=datetime.date(2026, 7, 20), status="Preview",
        venue=Venue(venue_id=1, name="x"),
        home=TeamGameInfo(team_id=114, name="Cleveland Guardians", abbrev="CLE", is_home=True),
        away=TeamGameInfo(team_id=142, name="Minnesota Twins", abbrev="MIN", is_home=False),
    )
    rg = RotoGame(
        away=RotoLineup("MIN", "Twins", False, "Joe Ryan", "R", []),
        home=RotoLineup("CLE", "Guardians", True, "Tanner Bibee", "R", []),
    )
    other = RotoGame(
        away=RotoLineup("NYY", "Yankees", False, None, None, []),
        home=RotoLineup("BOS", "Red Sox", False, None, None, []),
    )
    assert _match_roto_game(game, [other, rg]) is rg
    assert _match_roto_game(game, [other]) is None


def test_tail_xslg_fold_in():
    import numpy as np
    import pandas as pd

    from mlb_engine.features.tails import TailAdjuster

    rng = np.random.default_rng(0)
    # Population of ordinary batted balls for many batters (xwOBA/hard-hit/barrel).
    rows = []
    for bid in range(1, 41):
        for _ in range(20):
            rows.append({
                "batter": bid, "pitcher": 999,
                "launch_speed": float(rng.normal(88, 4)),
                "launch_speed_angle": 3,
                "estimated_woba_using_speedangle": float(rng.normal(0.32, 0.02)),
                "description": "hit_into_play", "events": "single",
            })
    df = pd.DataFrame(rows)

    # xSLG: batter 1 is a massive outlier (>2 SD), the rest cluster tightly.
    xslg = {bid: 0.40 for bid in range(1, 41)}
    xslg[1] = 0.95
    adj = TailAdjuster.build(df, xslg)
    assert adj.batter_z[1]["xslg"] >= 2.0
    # Outlier xSLG lifts batter 1's power multiplier above neutral.
    assert adj.batter_multiplier(1).get("HR", 1.0) > 1.0
    # Without xSLG the same batter is neutral (no other tail metric triggers).
    assert TailAdjuster.build(df).batter_multiplier(1) == {}


def test_fangraphs_tail_csv_name_matching(tmp_path):
    import pandas as pd

    from mlb_engine.data.fangraphs import load_fangraphs_tail_csv
    from mlb_engine.pipeline import _zscores

    # Custom-report style export with no MLBAMID -> name matching.
    csv = tmp_path / "fg.csv"
    pd.DataFrame(
        {
            "Name": ["José Ramírez", "Aaron Judge", "Joe Ryan"],
            "wRC+": [180, 200, float("nan")],
            "xSLG": [0.55, 0.62, float("nan")],
            "SIERA": [float("nan"), float("nan"), 2.50],
            "Stf+": [float("nan"), float("nan"), 120],
        }
    ).to_csv(csv, index=False)

    fg = load_fangraphs_tail_csv(csv)
    assert not fg.is_empty()
    assert fg.wrc_plus.by_name["jose ramirez"] == 180.0
    assert fg.xslg.by_name["aaron judge"] == 0.62
    assert fg.siera.by_name["joe ryan"] == 2.50
    assert fg.stuff_plus.by_name["joe ryan"] == 120.0
    # NaN cells are skipped, not coerced to 0; no MLBAMID column -> by_id empty.
    assert "joe ryan" not in fg.wrc_plus.by_name
    assert fg.wrc_plus.by_id == {}

    # SIERA is inverted downstream: lower SIERA -> higher (better) directional z.
    z = _zscores({"a": 2.5, "b": 3.5, "c": 4.5})
    assert z["a"] < 0 < z["c"]  # raw: low SIERA has negative z, so it gets negated


def test_fangraphs_tail_csv_prefers_mlbamid(tmp_path):
    import pandas as pd

    from mlb_engine.data.fangraphs import load_fangraphs_tail_csv
    from mlb_engine.pipeline import _metric_to_id_z

    # Real FanGraphs pitcher export shape: has an MLBAMID column.
    csv = tmp_path / "fg_sp.csv"
    pd.DataFrame(
        {
            "Name": ["Tarik Skubal", "Chris Sale", "Joe Ryan"],
            "SIERA": [2.08, 3.03, 3.60],
            "Stuff+": [118.7, 119.9, 105.0],
            "MLBAMID": [669373, 519242, 657746],
        }
    ).to_csv(csv, index=False)

    fg = load_fangraphs_tail_csv(csv)
    assert fg.siera.by_id[669373] == 2.08
    assert fg.stuff_plus.by_id[519242] == 119.9
    assert fg.siera.by_name == {}  # id present -> name path unused

    # SIERA inverted: Skubal (lowest SIERA) should get the highest (best) z.
    allowed = {669373, 519242, 657746}
    z = _metric_to_id_z(fg.siera, {}, allowed, invert=True)
    assert z[669373] == max(z.values())


def test_vsin_parse_helpers():
    from mlb_engine.data.vsin import _american, _norm_name, _pct

    assert _norm_name("St. Louis Cardinals") == _norm_name("ST Louis Cardinals")
    assert _pct("84% \u25b2") == 84.0
    assert _pct("nan") is None
    assert _american("-131") == -131.0
    assert _american("109") == 109.0


# ---- bullpen fatigue feed (StatsAPI parsing) ----
def test_bullpen_fatigue_from_boxscores():
    import datetime

    from mlb_engine.data.mlb_statsapi import MLBStatsClient

    tid = 100
    sched = {"dates": [
        {"date": "2026-07-17", "games": [
            {"gamePk": 1, "status": {"abstractGameState": "Final"}}]},
        {"date": "2026-07-18", "games": [
            {"gamePk": 2, "status": {"abstractGameState": "Final"}}]},
    ]}

    def _box(relievers):
        players = {}
        pitchers = []
        for pid, gs, npitch in relievers:
            pitchers.append(pid)
            players[f"ID{pid}"] = {"stats": {"pitching": {
                "gamesStarted": gs, "numberOfPitches": npitch}}}
        return {"teams": {
            "home": {"team": {"id": tid}, "pitchers": pitchers, "players": players},
            "away": {"team": {"id": 999}, "pitchers": [], "players": {}}}}

    # Day 1: relievers 10,11 throw; Day 2: reliever 10 again (back-to-back) + 12.
    boxes = {
        1: _box([(5, 1, 95), (10, 0, 20), (11, 0, 18)]),  # 5 is the starter
        2: _box([(6, 1, 90), (10, 0, 15), (12, 0, 35)]),
    }

    client = MLBStatsClient()

    def fake_get(path, **params):
        if path == "schedule":
            return sched
        pk = int(path.split("/")[1])
        return boxes[pk]

    client._get = fake_get  # type: ignore[method-assign]
    score = client.bullpen_fatigue(tid, datetime.date(2026, 7, 19))
    # reliever 10 = back-to-back, reliever 12 = 35 pitches yesterday -> 2 gassed.
    assert score == 40.0


# ---- comeback resilience ----
def test_comeback_flag_from_signals():
    from mlb_engine.models.comeback import ComebackSignal, evaluate

    # Strong contact edge + high OBP + long-leash opp starter -> resilient.
    strong = evaluate(ComebackSignal(
        xwoba_diff=0.045, team_obp=0.345, opp_starter_bf_cap=29,
    ))
    assert strong.resilient and strong.score >= 0.60 and strong.reasons

    # Outclassed, low-OBP team facing a quick hook -> not resilient.
    weak = evaluate(ComebackSignal(
        xwoba_diff=-0.050, team_obp=0.290, opp_starter_bf_cap=19,
    ))
    assert not weak.resilient and weak.score < strong.score

    # Bullpen-fatigue hook only fires when supplied.
    base = evaluate(ComebackSignal(xwoba_diff=0.0, team_obp=0.320, opp_starter_bf_cap=24))
    fatigued = evaluate(ComebackSignal(
        xwoba_diff=0.0, team_obp=0.320, opp_starter_bf_cap=24, opp_bullpen_fatigue=90,
    ))
    assert fatigued.score > base.score
    assert any("fatigued" in r for r in fatigued.reasons)


# ---- run-line PPV layer ----
def test_runline_xwoba_confirms_and_contradicts():
    from mlb_engine.market.runline import RunLineSignal, runline_adjustment

    # Home owns a strong xwOBA edge.
    sig = RunLineSignal(xwoba_diff=0.040)
    # Home -1.5 confirmed; away -1.5 (against the edge) flagged; home +1.5 outclass-safe.
    assert runline_adjustment("home", -1.5, sig)[0] == 1
    assert runline_adjustment("away", -1.5, sig)[0] == -1
    # Away is the outclassed dog -> +1.5 risk (downgrade).
    assert runline_adjustment("away", 1.5, sig)[0] == -1

    # Near-even by xwOBA -> both +1.5 dogs gain value.
    even = RunLineSignal(xwoba_diff=0.005)
    assert runline_adjustment("home", 1.5, even)[0] == 1
    assert runline_adjustment("away", 1.5, even)[0] == 1

    # Hooks: cold-but-sound + depleted favorite bullpen stack onto opponent +1.5.
    hook = RunLineSignal(xwoba_diff=None, cold_sound_side="away", fav_pen_depleted_side="home")
    steps, reasons = runline_adjustment("away", 1.5, hook)
    assert steps == 2 and len(reasons) == 2

    # Non-run-line selection -> untouched.
    assert runline_adjustment("home", None, sig) == (0, [])


def test_runline_luck_gap_nudges():
    from mlb_engine.market.runline import RunLineSignal, runline_adjustment

    # Home overperforms its contact quality (lucky), away lags (unlucky/buy-low).
    sig = RunLineSignal(luck_gap_home=1.5, luck_gap_away=-1.5)
    # Backing the lucky favorite at -1.5 -> fade (regression risk).
    steps, reasons = runline_adjustment("home", -1.5, sig)
    assert steps == -1 and any("regression" in r for r in reasons)
    # Backing the dog +1.5 when the favorite (home) is overperforming -> support.
    up, upr = runline_adjustment("away", 1.5, sig)
    assert up == 1 and any("back dog" in r for r in upr)
    # Backing the underrated team at -1.5 -> support.
    assert runline_adjustment("away", -1.5, sig)[0] == 1
    # Missing either side's gap -> neutral (backward compatible).
    assert runline_adjustment("home", -1.5, RunLineSignal(luck_gap_home=1.5)) == (0, [])


def test_team_form_build_and_luck_gaps():
    import pandas as pd

    from mlb_engine.features.team_form import build_team_forms, compute_luck_gaps

    # AAA bats poorly (proxy low) but has a high actual RD -> lucky/fade.
    # BBB bats well (proxy high) but a low actual RD -> unlucky/buy-low.
    rows = []
    for _ in range(250):
        rows.append({"home_team": "BBB", "away_team": "AAA", "inning_topbot": "Top",
                     "estimated_woba_using_speedangle": 0.300})  # AAA batting (low)
        rows.append({"home_team": "BBB", "away_team": "AAA", "inning_topbot": "Bot",
                     "estimated_woba_using_speedangle": 0.400})  # BBB batting (high)
    df = pd.DataFrame(rows)
    run_diffs = {"AAA": (0.8, 100), "BBB": (-0.8, 100)}

    forms = build_team_forms(df, run_diffs)
    assert set(forms) == {"AAA", "BBB"}
    assert abs(forms["AAA"].xwoba_for - 0.300) < 1e-9
    assert abs(forms["AAA"].xwoba_against - 0.400) < 1e-9
    assert forms["AAA"].xrd_proxy < 0 < forms["BBB"].xrd_proxy

    gaps = compute_luck_gaps(forms)
    assert gaps["AAA"] > 0 > gaps["BBB"]  # lucky vs unlucky

    # Thin batted-ball sample is dropped.
    thin = build_team_forms(df.head(10), run_diffs)
    assert thin == {}


def test_team_form_round_trip(tmp_path):
    from mlb_engine.features.team_form import TeamForm, load_team_forms, save_team_forms

    forms = {"AAA": TeamForm("AAA", 0.33, 0.31, 0.5, 100)}
    path = tmp_path / "team_form.json"
    save_team_forms(forms, path)
    assert load_team_forms(path) == forms
    # Missing cache -> empty (off/neutral).
    assert load_team_forms(tmp_path / "nope.json") == {}


# ---- distribution tails ----
def test_tail_bonus_and_penalty_symmetric():
    import numpy as np
    import pandas as pd

    from mlb_engine.features.tails import TailAdjuster

    rng = np.random.default_rng(3)
    rows = []
    # A population of average batters (ids 1000..1039) with ~15 BBE each.
    for pid in range(1000, 1040):
        for _ in range(16):
            rows.append(
                {
                    "batter": pid,
                    "pitcher": 1,
                    "launch_speed": float(rng.normal(88, 3)),
                    "launch_speed_angle": 1,
                    "estimated_woba_using_speedangle": float(rng.normal(0.32, 0.02)),
                    "description": "hit_into_play",
                    "events": "single",
                }
            )
    # An elite outlier (id 9001) and a bottom-tail hitter (id 9002).
    for _ in range(16):
        rows.append({"batter": 9001, "pitcher": 1, "launch_speed": 106.0,
                     "launch_speed_angle": 6, "estimated_woba_using_speedangle": 0.55,
                     "description": "hit_into_play", "events": "home_run"})
        rows.append({"batter": 9002, "pitcher": 1, "launch_speed": 72.0,
                     "launch_speed_angle": 1, "estimated_woba_using_speedangle": 0.18,
                     "description": "hit_into_play", "events": "field_out"})
    df = pd.DataFrame(rows)

    tails = TailAdjuster.build(df)
    elite = tails.batter_multiplier(9001)
    poor = tails.batter_multiplier(9002)
    assert elite and elite["HR"] > 1.0  # >2 SD above -> bonus
    assert poor and poor["HR"] < 1.0    # >2 SD below -> penalty
    assert tails.batter_multiplier(1000) == {}  # average -> neutral


# ---- arsenal matching (pitch mix) ----
def test_arsenal_matchup_high_whiff_vs_slider():
    import pandas as pd

    from mlb_engine.features.pitch_mix import (
        arsenal_matchup_multiplier,
        build_arsenal,
        build_batter_pitch_profile,
    )

    # Slider-heavy pitcher with lots of swinging strikes.
    prows = pd.DataFrame(
        {
            "pitch_type": ["SL"] * 60 + ["FF"] * 40,
            "description": (["swinging_strike"] * 30 + ["foul"] * 30) + ["hit_into_play"] * 40,
            "estimated_woba_using_speedangle": [None] * 100,
        }
    )
    arsenal = build_arsenal(prows)
    assert arsenal.usage["BRK"] > arsenal.usage["FB"]
    assert arsenal.swstr["BRK"] > 0.0

    # Batter who whiffs badly on breaking balls.
    brows = pd.DataFrame(
        {
            "pitch_type": ["SL"] * 40,
            "description": ["swinging_strike"] * 28 + ["hit_into_play"] * 12,
            "estimated_woba_using_speedangle": [0.2] * 40,
        }
    )
    bpp = build_batter_pitch_profile(brows)
    mult = arsenal_matchup_multiplier(arsenal, bpp)
    assert mult["K"] > 1.0  # high-whiff hitter vs. whiff-heavy slider -> more Ks

    # Empty inputs -> neutral (no fabricated edge).
    assert arsenal_matchup_multiplier(build_arsenal(pd.DataFrame()), bpp) == {}


# ---- schedule pacing (DGANG) ----
def test_dgang_tax_only_night_then_day():
    from mlb_engine.filters.schedule import dgang_multipliers, is_day, is_night

    # A ~7pm local night game -> next-day ~1pm day game (rest=1) taxes offense.
    prev_night = 19.0
    today_day = 13.0
    assert is_night(prev_night) and is_day(today_day)
    tax = dgang_multipliers(prev_night, today_day, rest_days=1)
    assert tax and tax["1B"] < 1.0 and tax["K"] > 1.0

    # Day-after-day, or two days rest, or missing times -> no tax.
    assert dgang_multipliers(13.0, 13.0, 1) == {}
    assert dgang_multipliers(prev_night, today_day, 2) == {}
    assert dgang_multipliers(None, today_day, 1) == {}


# ---- travel / circadian ----
def test_travel_eastward_worse_than_west_and_boosts_opponent_hr():
    from datetime import timedelta

    from mlb_engine.data.parks import Park
    from mlb_engine.filters.travel_rest import PrevGame, compute

    east_park = Park(1, "East", 40.8, -73.9, 0.0, "open", 100.0)  # NYC-ish
    west_lon = -122.4  # SF-ish
    today = date(2024, 7, 19)
    yday = today - timedelta(days=1)

    # Team flew EAST (from west coast to NYC) with 1 day rest.
    east_trip = compute(PrevGame(yday, 37.8, west_lon), east_park, today)
    # Team flew WEST the same distance/rest.
    west_park = Park(2, "West", 37.8, west_lon, 0.0, "open", 100.0)
    west_trip = compute(PrevGame(yday, 40.8, -73.9), west_park, today)

    assert east_trip.east and not west_trip.east
    assert east_trip.offense_mult < west_trip.offense_mult  # eastward hurts more
    assert east_trip.hr_allowed_mult > 1.0  # tired staff -> opponent HR spike

    # Mid-trip (same city as prior game) -> fully adapted, no penalty.
    mid = compute(PrevGame(yday, 40.8, -73.9), east_park, today)
    assert mid.offense_mult == 1.0
    assert mid.hr_allowed_mult == 1.0


# ---- ev + tiers ----
def test_ev_positive_when_underpriced():
    # model 60%, priced at +100 (fair 50%) -> positive EV
    ev = ev_per_dollar(0.60, 100)
    assert ev > 0.15


def test_classify_tiers():
    thr = EVThresholds()
    q = [MarketQuote("draftkings", 120, opposite_american=-140, handle_pct=70, bets_pct=45)]
    res = evaluate(0.50, q)
    tier, reasons = classify(res, thr)
    assert tier in (Tier.STRONG, Tier.MODERATE)
    # negative edge -> pass
    res2 = evaluate(0.30, q)
    tier2, _ = classify(res2, thr)
    assert tier2 == Tier.PASS


# ---- grading ----
def _rec(**kw) -> Recommendation:
    base = dict(
        game_date=date(2024, 7, 19),
        game_pk=1,
        matchup="AAA @ BBB",
        category="game",
        market="game_ml",
        selection="x",
        model_prob=0.5,
    )
    base.update(kw)
    return Recommendation(**base)


def test_grade_ml_and_total():
    res = GameResult(1, True, 5, 3, 3, 1)
    assert grade(_rec(market="game_ml", team_side="home", side="win"), res) == WIN
    assert grade(_rec(market="game_ml", team_side="away", side="win"), res) == LOSS
    over = _rec(category="game", market="game_total", line=7.5, side="over")
    assert grade(over, res) == WIN  # total 8 > 7.5
    under = _rec(category="game", market="game_total", line=8.5, side="under")
    assert grade(under, res) == WIN  # total 8 < 8.5


def test_grade_batter_prop():
    res = GameResult(
        1, True, 5, 3, 3, 1, players={99: PlayerLine(batting={"H": 2, "HR": 1, "R": 1, "RBI": 2})}
    )
    r = _rec(category="batter", market="batter_h", player_id=99, stat="H", line=1.5, side="over")
    assert grade(r, res) == WIN
    r2 = _rec(category="batter", market="batter_hr", player_id=99, stat="HR", line=1.5, side="over")
    assert grade(r2, res) == LOSS


def test_grade_batter_total_bases():
    # 1B + 2*2B + 3*3B + 4*HR = 1 + 2*1 + 3*0 + 4*1 = 7 total bases
    res = GameResult(
        1, True, 5, 3, 3, 1,
        players={99: PlayerLine(batting={"H": 3, "1B": 1, "2B": 1, "3B": 0, "HR": 1})},
    )
    win = _rec(category="batter", market="batter_tb", player_id=99, stat="TB", line=3.5, side="over")
    assert grade(win, res) == WIN  # 7 > 3.5
    loss = _rec(category="batter", market="batter_tb", player_id=99, stat="TB", line=8.5, side="over")
    assert grade(loss, res) == LOSS  # 7 < 8.5
    push = _rec(category="batter", market="batter_tb", player_id=99, stat="TB", line=7, side="over")
    assert grade(push, res) == PUSH  # 7 == 7


def test_grade_voids_props_on_players_who_never_appeared():
    # 99 played; 77 was a late scratch and so is absent from the box score.
    res = GameResult(
        1, True, 5, 3, 3, 1,
        players={99: PlayerLine(batting={"PA": 4, "H": 2, "R": 1, "RBI": 2})},
    )
    played = _rec(category="batter", market="batter_h", player_id=99, stat="H", line=1.5, side="over")
    assert grade(played, res) == WIN

    # Reading a missing player's stats back gives zero for everything, which
    # would sink this over. A book voids it instead, and so do we.
    scratched = _rec(
        category="batter", market="batter_h", player_id=77, stat="H", line=0.5, side="over"
    )
    assert grade(scratched, res) is None

    # A pinch runner appears in the box score but never bats.
    res.players[88] = PlayerLine(batting={"PA": 0, "H": 0, "R": 1, "RBI": 0})
    runner = _rec(
        category="batter", market="batter_h", player_id=88, stat="H", line=0.5, side="over"
    )
    assert grade(runner, res) is None

    # Same for a starter who was scratched before first pitch.
    sp = _rec(
        category="pitcher", market="pitcher_k", player_id=77, stat="K", line=5.5, side="over"
    )
    assert grade(sp, res) is None
    res.players[66] = PlayerLine(pitching={"BF": 24, "K": 7, "outs": 18})
    threw = _rec(
        category="pitcher", market="pitcher_k", player_id=66, stat="K", line=5.5, side="over"
    )
    assert grade(threw, res) == WIN


def test_grade_under_on_a_scratch_is_not_a_free_win():
    # The mirror image, and the reason "absent -> zero" is not a harmless default:
    # it would hand every under an automatic win on players who never played.
    res = GameResult(1, True, 5, 3, 3, 1, players={})
    under = _rec(
        category="batter", market="batter_h", player_id=77, stat="H", line=0.5, side="under"
    )
    assert grade(under, res) is None


def test_props_total_bases_market():
    import numpy as np

    from mlb_engine.models.montecarlo import GameSimResult
    from mlb_engine.models.props import batter_markets

    z = np.zeros((4, 9), dtype=np.int16)
    bat = {s: z.copy() for s in ("H", "1B", "2B", "3B", "HR", "R", "RBI")}
    # slot 0 across 4 sims: TB = 1B + 2*2B + 3*3B + 4*HR
    bat["1B"][:, 0] = [1, 0, 0, 2]
    bat["2B"][:, 0] = [0, 1, 0, 0]
    bat["HR"][:, 0] = [0, 0, 1, 0]  # TB per sim = [1, 2, 4, 2]
    res = GameSimResult(4, np.zeros(4), np.zeros(4), np.zeros(4), np.zeros(4), {"home": bat}, {})

    tb = [m for m in batter_markets(res, "home", 0, "Player X") if m.market == "batter_tb"]
    assert {m.line for m in tb} == {1.5, 2.5, 3.5}
    by_line = {m.line: m.prob for m in tb}
    assert abs(by_line[1.5] - 0.75) < 1e-9  # 3 of 4 sims (2,4,2) exceed 1.5
    assert abs(by_line[2.5] - 0.25) < 1e-9  # only the 4-TB sim exceeds 2.5


def test_grade_push():
    res = GameResult(1, True, 4, 4, 2, 2)
    r = _rec(market="game_ml", team_side="home", side="win")
    assert grade(r, res) == PUSH


def test_ip_to_outs():
    assert _ip_to_outs("5.2") == 17
    assert _ip_to_outs("6.0") == 18


# ---- scorecard ----
def test_scorecard_metrics():
    graded = [
        (_rec(tier=Tier.STRONG), WIN),
        (_rec(tier=Tier.STRONG), WIN),
        (_rec(tier=Tier.STRONG), LOSS),
        (_rec(tier=Tier.PASS), LOSS),
        (_rec(tier=Tier.PASS), LOSS),
    ]
    rows = build_scorecard(graded, date(2024, 7, 19))
    strong = next(r for r in rows if r.tier == Tier.STRONG.value)
    assert strong.n == 3
    assert strong.wins == 2
    assert abs(strong.ppv - 2 / 3) < 1e-3
    # 2 of 5 graded rows won, so a 2/3 hit rate is 26.7 points of real skill.
    assert abs(strong.base_win - 0.4) < 1e-3
    assert abs(strong.ppv_lift - (2 / 3 - 0.4)) < 1e-3


def test_scorecard_npv_is_free_when_a_tier_passes_on_everything():
    # Pass never bets, so all three of its "negatives" that lost score a perfect
    # NPV -- against a base loss rate that is itself 0.6. The lift is what shows
    # the abstention was worth nothing.
    graded = [
        (_rec(tier=Tier.PASS), LOSS),
        (_rec(tier=Tier.PASS), LOSS),
        (_rec(tier=Tier.PASS), LOSS),
        (_rec(tier=Tier.STRONG), WIN),
        (_rec(tier=Tier.STRONG), WIN),
    ]
    buy = next(r for r in build_scorecard(graded, date(2024, 7, 19)) if r.tier == "Buy (S+M)")
    assert buy.npv == 1.0
    assert abs(buy.base_loss - 0.6) < 1e-3
    assert abs(buy.npv_lift - 0.4) < 1e-3


def test_append_scorecard_widens_a_file_written_under_an_older_header(tmp_path):
    # Appending new fields to an old header would shift every value in the new
    # rows one column left of where a reader expects it.
    path = tmp_path / "scorecard.csv"
    path.write_text("date,tier,n,wins,losses,ppv,npv,sensitivity,specificity,roi\n2024-07-18,Strong,2,1,1,0.5,0.5,0.5,0.5,-0.05\n")
    append_scorecard(build_scorecard([(_rec(tier=Tier.STRONG), WIN)], date(2024, 7, 19)), path)

    import csv as _csv

    with path.open(newline="") as f:
        rows = list(_csv.DictReader(f))
    assert list(rows[0]) == FIELDS
    assert rows[0]["date"] == "2024-07-18"
    assert rows[0]["ppv"] == "0.5"
    assert rows[0]["npv_lift"] == ""  # cannot be recomputed; never guessed
    assert rows[-1]["date"] == "2024-07-19"
    assert rows[-1]["ppv"] == "1.0"


def test_rates_from_events_sums_to_one():
    import pandas as pd

    ev = pd.Series(["single", "home_run", "strikeout", "walk", "field_out"] * 20)
    r = rates_from_events(ev)
    total = sum(r.as_dict().values())
    assert abs(total - 1.0) < 1e-9
    assert 0 < r.obp < 1


def test_the_league_rates_are_a_distribution():
    """Exactly, not nearly: several callers assert their own normalisation to 1e-9.

    Four decimal places on seven independently rounded numbers misses a distribution
    by 1e-4 about half the time, which is 1e5 outside what those callers allow. The
    refit script emits OUT as the residual for this reason.
    """
    assert abs(sum(LEAGUE_RATES.values()) - 1.0) < 1e-9
    assert all(v > 0 for v in LEAGUE_RATES.values())


def test_the_league_rates_are_the_measured_league():
    """A tripwire on the constants, because no invariant inside the engine can be one.

    ``LEAGUE_RATES`` is the log5 denominator in ``combine``, so a value that is too
    small inflates that outcome in every matchup on the slate. It sat at BB 0.085
    against a measured 0.1012 -- walks 18.6% high everywhere, doubles 8.5% light.

    Note what cannot catch this. ``test_combine_normalizes`` asserts that a
    league-average batter facing a league-average pitcher returns the league, which
    reads like the right invariant and is vacuous: the batter is *built from*
    ``LEAGUE_RATES``, so ``b * p / lg`` returns ``lg`` identically for any values at
    all, right or wrong. ``combine`` was always self-consistent; the constant was
    wrong, and only the data can say so. Hence a pinned measurement, refittable with
    ``python -m scripts.league_rates``.
    """
    measured = {  # 115,504 classified PA, 2026-03-25..07-22, the engine's own bucketer
        "1B": 0.1417,
        "2B": 0.0416,
        "3B": 0.0036,
        "HR": 0.0311,
        "BB": 0.1020,
        "K": 0.2228,
        "OUT": 0.4572,
    }
    assert LEAGUE_RATES == measured


def test_a_free_pass_is_not_an_out():
    """The intentional walk and catcher's interference put the batter on first.

    Both fell through the bucketer's ``else`` and were counted as outs, and the
    intentional walk is not spread evenly: it is aimed at the best hitter in the
    lineup, so the error concentrated on exactly the bats worth pricing. Yordan
    Alvarez's walk rate over the measured window was 10.8% against a true 15.2%.
    """
    import pandas as pd

    from mlb_engine.features.rolling import _bucket_counts

    counts = _bucket_counts(pd.Series(["intent_walk", "catcher_interf", "walk", "hit_by_pitch"]))
    assert counts["BB"] == 4
    assert counts["OUT"] == 0


def test_an_intentional_walk_is_charged_to_the_pitcher_and_interference_is_not():
    """``WALK_EVENTS`` doubles as a pitcher's walks allowed, where the two differ."""
    from mlb_engine.features.rolling import FREE_PASS_EVENTS, WALK_EVENTS

    assert "intent_walk" in WALK_EVENTS
    assert "catcher_interf" not in WALK_EVENTS
    assert WALK_EVENTS < FREE_PASS_EVENTS
    assert "catcher_interf" in FREE_PASS_EVENTS


def test_np_import_available():
    assert np.array([1, 2, 3]).sum() == 6


# ---- strikeout-model upgrades ----
def _pitch_rows(n_pitches, csw_frac, swstr_frac, k_frac, stand="R", pa_frac=0.25):
    """Synthetic pitch-level rows with controllable CSW%/SwStr% and K%."""
    import pandas as pd

    n_whiff = int(n_pitches * swstr_frac)
    n_called = int(n_pitches * (csw_frac - swstr_frac))
    desc = (
        ["swinging_strike"] * n_whiff
        + ["called_strike"] * max(n_called, 0)
        + ["hit_into_play"] * (n_pitches - n_whiff - max(n_called, 0))
    )
    n_pa = int(n_pitches * pa_frac)
    n_k = int(n_pa * k_frac)
    events = ["strikeout"] * n_k + ["field_out"] * (n_pa - n_k)
    events = events + [None] * (n_pitches - n_pa)
    return pd.DataFrame(
        {
            "description": desc[:n_pitches],
            "events": events[:n_pitches],
            "stand": [stand] * n_pitches,
            "strikes": [0] * n_pitches,
            "launch_speed": [None] * n_pitches,
            "pfx_z": [None] * n_pitches,
            "release_extension": [None] * n_pitches,
            "release_pos_x": [None] * n_pitches,
            "release_pos_z": [None] * n_pitches,
            "release_spin_rate": [None] * n_pitches,
        }
    )


def test_expected_k_pct_tracks_stuff():
    from mlb_engine.features.regression import build_pitcher_regression

    hi = build_pitcher_regression(_pitch_rows(1000, csw_frac=0.36, swstr_frac=0.16, k_frac=0.30))
    lo = build_pitcher_regression(_pitch_rows(1000, csw_frac=0.24, swstr_frac=0.07, k_frac=0.15))
    assert hi.expected_k_pct() > lo.expected_k_pct()
    # A whiffy arm's xK% should clear the ~.22 league mean; a soft-tosser's shouldn't.
    assert hi.expected_k_pct() > 0.24
    assert lo.expected_k_pct() < 0.20


def _pitch_rows_disc(n_pitches, zone_frac, chase_frac, fstrike_frac):
    """Pitch rows with controllable Zone%, chase (O-Swing%), and F-strike%."""
    import pandas as pd

    n_zone = int(n_pitches * zone_frac)
    n_out = n_pitches - n_zone
    n_chase = int(n_out * chase_frac)
    # In-zone: called strikes; out-of-zone: swings (chase) then balls.
    desc = (
        ["called_strike"] * n_zone
        + ["swinging_strike"] * n_chase
        + ["ball"] * (n_out - n_chase)
    )
    zone = [1] * n_zone + [13] * n_out
    n_first = int(n_pitches * 0.4)
    n_fs = int(n_first * fstrike_frac)
    balls = [0] * n_first + [1] * (n_pitches - n_first)
    strikes = [0] * n_first + [0] * (n_pitches - n_first)
    ptype = ["S"] * n_fs + ["B"] * (n_first - n_fs) + ["B"] * (n_pitches - n_first)
    return pd.DataFrame(
        {
            "description": desc[:n_pitches],
            "zone": zone[:n_pitches],
            "balls": balls[:n_pitches],
            "strikes": strikes[:n_pitches],
            "type": ptype[:n_pitches],
            "events": [None] * n_pitches,
            "launch_speed": [None] * n_pitches,
            "pfx_z": [None] * n_pitches,
            "release_extension": [None] * n_pitches,
            "release_pos_x": [None] * n_pitches,
            "release_pos_z": [None] * n_pitches,
            "release_spin_rate": [None] * n_pitches,
        }
    )


def test_expected_bb_pct_tracks_command():
    from mlb_engine.features.regression import build_pitcher_regression

    # Wild arm: low zone / low chase / low F-strike -> high xBB.
    wild = build_pitcher_regression(_pitch_rows_disc(1000, 0.40, 0.22, 0.50))
    # Command arm: high zone / high chase / high F-strike -> low xBB (NPV screen).
    sharp = build_pitcher_regression(_pitch_rows_disc(1000, 0.55, 0.38, 0.70))
    assert wild.expected_bb_pct() > sharp.expected_bb_pct()
    assert wild.expected_bb_pct() > 0.085
    assert sharp.expected_bb_pct() < 0.085


def test_blend_bb_rate_small_sample_leans_on_prior():
    import pandas as pd

    from mlb_engine.features.rolling import blend_bb_rate, rates_from_events

    thin = rates_from_events(pd.Series(["walk"] * 1 + ["field_out"] * 9))
    blended = blend_bb_rate(thin, bb_prior=0.16, prior_weight=150.0)
    assert blended.p_bb > thin.p_bb
    assert abs(sum(blended.as_dict().values()) - 1.0) < 1e-9


def test_blend_k_rate_small_sample_leans_on_prior():
    import pandas as pd

    from mlb_engine.features.rolling import blend_k_rate, rates_from_events

    # Thin sample (few PA) with a low observed K -> pulled up toward a high xK prior.
    thin = rates_from_events(pd.Series(["strikeout"] * 2 + ["field_out"] * 8))
    blended = blend_k_rate(thin, k_prior=0.32, prior_weight=150.0)
    assert blended.p_k > thin.p_k
    assert abs(sum(blended.as_dict().values()) - 1.0) < 1e-9


def test_platoon_k_multiplier():
    import pandas as pd

    from mlb_engine.features.regression import build_pitcher_regression

    # Big reverse split: dominates RHB (high K), struggles vs LHB (low K).
    rows = pd.concat(
        [
            _pitch_rows(600, 0.32, 0.13, k_frac=0.34, stand="R", pa_frac=0.30),
            _pitch_rows(600, 0.26, 0.09, k_frac=0.16, stand="L", pa_frac=0.30),
        ],
        ignore_index=True,
    )
    reg = build_pitcher_regression(rows)
    assert reg.platoon_k_multiplier("R") > 1.0
    assert reg.platoon_k_multiplier("L") < 1.0
    assert reg.platoon_k_multiplier(None) == 1.0


def test_expected_bf_cap_workload_and_opener():
    from datetime import date as _date

    import pandas as pd

    from mlb_engine.features.workload import expected_bf_cap

    as_of = _date(2024, 7, 20)
    # Five ~22-BF starts -> cap tracks workload (avg+buffer), under the manager cap.
    deep = pd.DataFrame(
        {
            "game_date": sum(([_date(2024, 7, d)] * 22 for d in (2, 5, 8, 11, 14)), []),
            "events": ["field_out"] * 110,
            "pitcher": [1] * 110,
        }
    )
    cap = expected_bf_cap(deep, as_of, form_days=28, manager_cap=29)
    assert 23 <= cap <= 26

    # Opener: repeated ~6-BF outings collapses the cap.
    opener = pd.DataFrame(
        {
            "game_date": sum(([_date(2024, 7, d)] * 6 for d in (4, 7, 10, 13)), []),
            "events": ["field_out"] * 24,
            "pitcher": [2] * 24,
        }
    )
    assert expected_bf_cap(opener, as_of, form_days=28, manager_cap=24) <= 12


def test_framing_lookup_and_human_factor():
    from mlb_engine.data.catcher_framing import framing_runs_for_name
    from mlb_engine.filters.human import HumanFactors

    assert framing_runs_for_name("Patrick Bailey") and framing_runs_for_name("Patrick Bailey") > 0
    assert framing_runs_for_name("Unknown Guy") is None
    # Elite framer -> steals strikes -> K multiplier > 1 for the pitching side.
    m = HumanFactors(catcher_framing_runs=15.0).offense_multipliers()
    assert m.get("K", 1.0) > 1.0
    assert m.get("BB", 1.0) < 1.0


def test_load_framing_parses_savant_columns(monkeypatch):
    import mlb_engine.data.catcher_framing as cf

    class _Resp:
        text = "id,name,pitches,rv_tot\n111,A B,5000,8.4\n222,C D,4800,-6.1\n"

        def raise_for_status(self):
            return None

    monkeypatch.setattr(cf.http, "get", lambda *a, **k: _Resp())
    out = cf.load_framing(2024)
    assert out == {111: 8.4, 222: -6.1}


def test_load_framing_neutral_on_failure(monkeypatch):
    import mlb_engine.data.catcher_framing as cf

    def _boom(*a, **k):
        raise RuntimeError("network down")

    monkeypatch.setattr(cf.http, "get", _boom)
    assert cf.load_framing(2024) == {}


def test_strong_only_and_min_edge_selection():
    from mlb_engine.config import EVThresholds
    from mlb_engine.market.ev import EVResult, MarketQuote
    from mlb_engine.market.tiers import Tier, classify

    q = MarketQuote(book="bk", american=-110)

    def _res(ev, edge):
        return EVResult(
            model_prob=0.5, best_quote=q, decimal=1.91, ev=ev,
            fair_prob=0.5 - edge, edge=edge, sharp_divergence=None,
        )

    moderate = _res(ev=0.05, edge=0.03)
    assert classify(moderate, EVThresholds())[0] == Tier.MODERATE
    # strong_only downgrades Moderate to Pass.
    assert classify(moderate, EVThresholds(strong_only=True))[0] == Tier.PASS
    # A raised min_edge also rejects a thin edge.
    assert classify(_res(ev=0.05, edge=0.01), EVThresholds(min_edge=0.03))[0] == Tier.PASS


def test_tier_does_not_reward_the_longer_price():
    """Same edge, two prices: EV differs, the tier must not."""
    from mlb_engine.config import EVThresholds
    from mlb_engine.market.ev import MarketQuote, evaluate
    from mlb_engine.market.tiers import classify

    thr = EVThresholds()
    # A 5-point edge over the devigged price, quoted as a dog and as a favourite.
    dog = evaluate(1 / 3 + 0.05, [MarketQuote("dk", 200, opposite_american=-200)])
    fave = evaluate(5 / 7 + 0.05, [MarketQuote("dk", -250, opposite_american=250)])
    assert abs(dog.edge - 0.05) < 1e-9 and abs(fave.edge - 0.05) < 1e-9
    assert dog.ev > fave.ev  # EV = decimal odds x edge, so the dog looks bigger
    assert classify(dog, thr)[0] is classify(fave, thr)[0] is Tier.STRONG


def test_implausible_edge_is_a_pass():
    """Past max_edge the disagreement reads as a model error, not a bet."""
    from mlb_engine.config import EVThresholds
    from mlb_engine.market.ev import EVResult, MarketQuote
    from mlb_engine.market.tiers import classify

    q = MarketQuote(book="bk", american=-110)
    huge = EVResult(
        model_prob=0.70, best_quote=q, decimal=1.91, ev=0.337,
        fair_prob=0.50, edge=0.20, sharp_divergence=None,
    )
    assert classify(huge, EVThresholds())[0] is Tier.PASS
    assert any("> 0.08" in r for r in classify(huge, EVThresholds())[1])
    # The cap is what rejects it, not the EV or the thin-edge guard.
    assert classify(huge, EVThresholds(max_edge=1.0))[0] is Tier.STRONG


def test_zero_ev_price_is_a_pass():
    """An edge over the consensus is not a bet if the best price does not pay."""
    from mlb_engine.config import EVThresholds
    from mlb_engine.market.ev import EVResult, MarketQuote
    from mlb_engine.market.tiers import classify

    q = MarketQuote(book="bk", american=-140)
    vigged = EVResult(
        model_prob=0.58, best_quote=q, decimal=1.714, ev=-0.0058,
        fair_prob=0.53, edge=0.05, sharp_divergence=None,
    )
    assert classify(vigged, EVThresholds())[0] is Tier.PASS


# ---- ledger ----
def test_ledger_entries_and_pnl():
    from mlb_engine.audit.ledger import entries_from_graded

    graded = [
        (_rec(tier=Tier.STRONG, market="game_ml", market_american=100, ev=0.2), WIN),
        (_rec(tier=Tier.STRONG, market="game_ml", market_american=-110), LOSS),
        (_rec(tier=Tier.MODERATE, market="game_ml", market_american=100), PUSH),
    ]
    entries = entries_from_graded(graded, date(2024, 7, 19))
    assert [e.pnl for e in entries] == [1.0, -1.0, 0.0]
    assert entries[0].category == "Moneyline"


def test_ledger_overall_and_dedup(tmp_path):
    from mlb_engine.audit.ledger import (
        entries_from_graded,
        load_ledger,
        overall_metrics,
        update_ledger,
    )

    path = tmp_path / "ledger.csv"
    day1 = entries_from_graded(
        [
            (_rec(tier=Tier.STRONG, market_american=100), WIN),
            (_rec(tier=Tier.STRONG, market_american=-110), LOSS),
        ],
        date(2024, 7, 19),
    )
    update_ledger(path, day1, date(2024, 7, 19))
    day2 = entries_from_graded(
        [(_rec(tier=Tier.STRONG, market_american=100), WIN)], date(2024, 7, 20)
    )
    all_entries = update_ledger(path, day2, date(2024, 7, 20))
    assert len(all_entries) == 3
    assert len(load_ledger(path)) == 3

    # re-auditing day 1 replaces (does not duplicate) that date's rows
    all_entries = update_ledger(path, day1, date(2024, 7, 19))
    assert len(all_entries) == 3

    strong = next(m for m in overall_metrics(all_entries) if m.tier == Tier.STRONG.value)
    assert strong.n == 3 and strong.wins == 2 and abs(strong.ppv - 2 / 3) < 1e-3
    assert abs(strong.units - 1.0) < 1e-9  # +1 -1 +1
    # Two +100s and one -110: the prices demanded (.5 + .5 + .5238) / 3 to break
    # even, so a 66.7% win rate is genuinely ahead rather than assumed to be.
    assert abs(strong.required_win_pct - 0.5079) < 1e-3
    assert strong.win_pct > strong.required_win_pct


# ---- backtest analytics ----
def test_backtest_summarize_and_confusion():
    from mlb_engine.backtest import confidence_gap, confusion, sample_dates, summarize

    # favored (0.7) over that won, and faded (0.3) over that lost -> perfect
    graded = [
        (_rec(market="batter_h", model_prob=0.7), WIN),
        (_rec(market="batter_h", model_prob=0.3), LOSS),
        (_rec(market="game_total", model_prob=0.6, line=8.5, side="over"), LOSS),
        (_rec(market="game_ml", model_prob=0.55), PUSH),
    ]
    summ = {g.group: g for g in summarize(graded)}
    assert summ["Batter Props"].n == 2 and summ["Batter Props"].wins == 1
    assert summ["ALL"].pushes == 1  # the ML push counted, excluded from n
    assert summ["ALL"].n == 3

    conf = {c.group: c for c in confusion(graded)}
    bp = conf["Batter Props"]
    assert (bp.tp, bp.fp, bp.fn, bp.tn) == (1, 0, 0, 1)
    assert bp.ppv == 1.0 and bp.npv == 1.0

    gaps = confidence_gap(graded)
    # favored batter pick predicted 0.70, won (1.0) -> under-confident (negative)
    assert gaps["Batter Props"] < 0

    assert len(sample_dates(date(2024, 6, 1), date(2024, 6, 9), 2)) == 5


def test_isotonic_calibration_pulls_overconfident_down():
    from mlb_engine.calibration import Calibrator, IsotonicMap

    # raw 0.9 only wins ~50% -> calibrated must drop below 0.9 (monotone)
    pairs = [(0.9, i % 2) for i in range(200)] + [(0.1, 0) for _ in range(200)]
    m = IsotonicMap.fit(pairs)
    assert m.apply(0.9) < 0.9
    assert m.apply(0.1) <= m.apply(0.9)  # monotone (ranking preserved)

    cal = Calibrator.fit([("pitcher_k", 0.9, i % 2) for i in range(600)])
    assert cal.apply("pitcher_k", 0.9) < 0.9
    assert cal.apply("unknown_market", 0.9) < 0.9  # falls back to pooled default


def test_false_positive_and_negative_findings():
    from mlb_engine.backtest import false_negative_findings, false_positive_findings

    # 200 favored pitcher_er picks that mostly lose -> flagged as FP risk
    fp_graded = [
        (_rec(market="pitcher_er", model_prob=0.6, side="over"), LOSS if i % 4 else WIN)
        for i in range(200)
    ]
    fps = false_positive_findings(fp_graded, min_n=50)
    assert any("pitcher_er" in f for f in fps)

    fns = false_negative_findings(fp_graded, min_n=50)
    assert any("mostly correct" in f for f in fns)  # no faded picks here


# ---- whole-engine / daily / per-prop ledger metrics ----
def test_engine_metrics_directional_confusion():
    from mlb_engine.audit.ledger import engine_metrics, entries_from_graded

    # favored (>=.5) win -> TP; favored lose -> FP; faded (<.5) lose -> TN; faded win -> FN
    graded = [
        (_rec(model_prob=0.60), WIN),   # TP
        (_rec(model_prob=0.55), LOSS),  # FP
        (_rec(model_prob=0.40), LOSS),  # TN
        (_rec(model_prob=0.45), WIN),   # FN
        (_rec(model_prob=0.70), PUSH),  # excluded (favored push)
    ]
    m = engine_metrics(entries_from_graded(graded, date(2024, 7, 19)))
    assert m.tier == "ENGINE (p>=.5)"
    assert m.n == 2  # favored decided picks (TP + FP)
    assert abs(m.ppv - 0.5) < 1e-9  # 1 TP / (1 TP + 1 FP)
    assert abs(m.npv - 0.5) < 1e-9  # 1 TN / (1 TN + 1 FN)
    assert abs(m.sensitivity - 0.5) < 1e-9
    assert abs(m.specificity - 0.5) < 1e-9
    assert m.pushes == 1


def test_daily_engine_and_prop_metrics():
    from mlb_engine.audit.ledger import (
        daily_engine_metrics,
        entries_from_graded,
        prop_metrics,
    )

    entries = entries_from_graded(
        [(_rec(market="batter_hr", model_prob=0.6), WIN)], date(2024, 7, 19)
    ) + entries_from_graded(
        [
            (_rec(market="pitcher_k", model_prob=0.6), LOSS),
            (_rec(market="game_ml", model_prob=0.6), WIN),  # not a prop
        ],
        date(2024, 7, 20),
    )

    daily = daily_engine_metrics(entries)
    assert [d.tier for d in daily] == ["2024-07-19", "2024-07-20"]
    # day 1: the lone favored pick (batter_hr) won -> PPV 1.0.
    # day 2 (whole engine): favored pitcher_k lost + favored game_ml won -> PPV 0.5.
    assert daily[0].ppv == 1.0 and daily[1].ppv == 0.5

    props = {p.tier: p for p in prop_metrics(entries)}
    assert set(props) == {"batter_hr", "pitcher_k", "ALL PROPS"}  # game_ml excluded
    assert props["batter_hr"].ppv == 1.0 and props["pitcher_k"].ppv == 0.0
    assert props["ALL PROPS"].n == 2  # both favored prop picks


# ---- per-prop FP / FN / TP insights ----
def test_prop_insights_fp_fn_tp():
    from mlb_engine.audit.analysis import (
        FALSE_NEGATIVE,
        FALSE_POSITIVE,
        TRUE_POSITIVE,
        false_negative_insights,
        false_positive_insights,
        true_positive_insights,
    )
    from mlb_engine.audit.ledger import entries_from_graded

    # favored HR picks that mostly LOSE -> false-positive risk (raise PPV)
    fp = [(_rec(market="batter_hr", model_prob=0.62), LOSS if i % 4 else WIN) for i in range(40)]
    # favored K picks that mostly WIN -> true positive (reinforce)
    tp = [(_rec(market="pitcher_k", model_prob=0.62), WIN if i % 4 else LOSS) for i in range(40)]
    # faded 1B picks that mostly WIN -> false negative (reclaim, raise NPV)
    fn = [(_rec(market="batter_1b", model_prob=0.40), WIN if i % 4 else LOSS) for i in range(40)]
    entries = entries_from_graded(fp + tp + fn, date(2024, 7, 19))

    fps = false_positive_insights(entries, min_n=20)
    assert any(i.market == "batter_hr" and i.kind == FALSE_POSITIVE for i in fps)

    tps = true_positive_insights(entries, min_n=20)
    assert any(i.market == "pitcher_k" and i.kind == TRUE_POSITIVE for i in tps)

    fns = false_negative_insights(entries, min_n=20)
    assert any(i.market == "batter_1b" and i.kind == FALSE_NEGATIVE for i in fns)


def test_picked_margin_run_line():
    from mlb_engine.audit.grade import picked_margin

    res = GameResult(1, True, 5, 3, 3, 1)  # home wins 5-3 (full), 3-1 (F5)
    assert picked_margin(_rec(market="game_rl", team_side="home", line=-1.5), res) == 2.0
    assert picked_margin(_rec(market="game_rl", team_side="away", line=1.5), res) == -2.0
    assert picked_margin(_rec(market="f5_rl", team_side="home", line=-1.5), res) == 2.0
    # non run-line markets carry no margin
    assert picked_margin(_rec(market="game_ml", team_side="home"), res) is None
    # missing team_side -> None
    assert picked_margin(_rec(market="game_rl", line=-1.5), res) is None


def test_run_line_miss_matrix_and_findings():
    from mlb_engine.audit.analysis import run_line_miss_findings, run_line_miss_matrix
    from mlb_engine.audit.ledger import entries_from_graded

    results = {
        1: GameResult(1, True, 4, 3, 0, 0),  # home +1  -> -1.5 fav loses by 1
        2: GameResult(2, True, 6, 3, 0, 0),  # home +3  -> -1.5 fav covers
        3: GameResult(3, True, 2, 5, 0, 0),  # home -3  -> -1.5 fav loses outright
        4: GameResult(4, True, 7, 1, 0, 0),  # home +6  -> away +1.5 dog blown out
        5: GameResult(5, True, 5, 3, 0, 0),  # home +2  -> away +1.5 dog loses close
    }
    graded = [
        (_rec(game_pk=1, market="game_rl", team_side="home", line=-1.5, model_prob=0.6), LOSS),
        (_rec(game_pk=2, market="game_rl", team_side="home", line=-1.5, model_prob=0.6), WIN),
        (_rec(game_pk=3, market="game_rl", team_side="home", line=-1.5, model_prob=0.6), LOSS),
        (_rec(game_pk=4, market="game_rl", team_side="away", line=1.5, model_prob=0.6), LOSS),
        (_rec(game_pk=5, market="game_rl", team_side="away", line=1.5, model_prob=0.6), LOSS),
    ]
    entries = entries_from_graded(graded, date(2024, 7, 19), results)
    assert entries[0].margin == 1.0  # one-run win recorded

    m = run_line_miss_matrix(entries)
    assert (m.fav_n, m.fav_cover, m.fav_one_run, m.fav_outright) == (3, 1, 1, 1)
    assert (m.dog_n, m.dog_cover, m.dog_moderate, m.dog_blowout) == (2, 0, 1, 1)

    finds = run_line_miss_findings(entries)
    assert any("one-run" in f for f in finds)
    assert any("blowout" in f for f in finds)


def test_run_line_miss_matrix_persists_across_ledger_io(tmp_path):
    from mlb_engine.audit.analysis import run_line_miss_matrix
    from mlb_engine.audit.ledger import entries_from_graded, load_ledger, update_ledger

    results = {1: GameResult(1, True, 4, 3, 0, 0)}
    graded = [
        (_rec(game_pk=1, market="game_rl", team_side="home", line=-1.5, model_prob=0.6), LOSS)
    ]
    entries = entries_from_graded(graded, date(2024, 7, 19), results)
    path = tmp_path / "ledger.csv"
    update_ledger(path, entries, date(2024, 7, 19))
    reloaded = load_ledger(path)
    assert reloaded[0].margin == 1.0  # margin survives the CSV round-trip
    assert run_line_miss_matrix(reloaded).fav_one_run == 1


def test_report_renders_run_line_miss_matrix():
    from mlb_engine.audit.ledger import entries_from_graded
    from mlb_engine.output.report import build_report_data, render_markdown_report

    results = {i: GameResult(i, True, 4, 3, 0, 0) for i in range(1, 6)}
    graded = [
        (_rec(game_pk=i, market="game_rl", team_side="home", line=-1.5, model_prob=0.6), LOSS)
        for i in range(1, 6)
    ]
    entries = entries_from_graded(graded, date(2026, 7, 23), results)
    data = build_report_data(entries, period_label="Daily", subtitle="x")
    assert data.rl_matrix.has_data
    md = render_markdown_report(data)
    assert "Run-line miss matrix" in md


def test_ledger_workbook_with_analysis(tmp_path):
    from mlb_engine.audit.analysis import prop_insights
    from mlb_engine.audit.ledger import (
        daily_engine_metrics,
        daily_rollup,
        engine_metrics,
        entries_from_graded,
        overall_metrics,
        prop_metrics,
    )
    from mlb_engine.output.excel import write_ledger_workbook

    entries = entries_from_graded(
        [(_rec(market="batter_hr", model_prob=0.6, market_american=100), WIN)],
        date(2024, 7, 19),
    )
    out = tmp_path / "ledger.xlsx"
    write_ledger_workbook(
        entries,
        [engine_metrics(entries), *overall_metrics(entries)],
        daily_rollup(entries),
        out,
        daily_engine=daily_engine_metrics(entries),
        prop_rows=prop_metrics(entries),
        insights=prop_insights(entries, min_n=1),
    )
    assert out.exists()

    from openpyxl import load_workbook

    wb = load_workbook(out)
    for sheet in ("Overall", "Daily PPV-NPV", "Prop PPV-NPV", "Prop Insights", "Daily", "Bets"):
        assert sheet in wb.sheetnames


def test_best_flag_is_reachable_under_the_edge_cap():
    """A BEST bar in EV points is unreachable once edge is capped."""
    from mlb_engine.config import EVThresholds
    from mlb_engine.output.excel import _is_best

    cap = EVThresholds().max_edge
    # A prop at the cap: the most a capped buy can ever disagree with the price.
    prop = _rec(
        category="pitcher", market="pitcher_k", selection="P Ks o5.5",
        model_prob=0.58, fair_prob=0.58 - cap, edge=cap, ev=0.17,
        market_american=110, tier=Tier.STRONG,
    )
    assert _is_best(prop, "Pitcher Props")
    # The price window still excludes prop longshots.
    assert not _is_best(
        _rec(
            category="pitcher", market="pitcher_k", selection="P Ks o8.5",
            model_prob=0.30, fair_prob=0.30 - cap, edge=cap, ev=0.60,
            market_american=400, tier=Tier.STRONG,
        ),
        "Pitcher Props",
    )
    # A thin-edge buy at a long price is not a standout, however big its EV.
    assert not _is_best(
        _rec(
            category="game", market="game_ml", selection="AAA ML",
            model_prob=0.36, fair_prob=0.335, edge=0.025, ev=0.08,
            market_american=200, tier=Tier.MODERATE,
        ),
        "Moneyline",
    )


def test_recommendation_workbook_tabs_and_colors(tmp_path):
    from openpyxl import load_workbook

    from mlb_engine.output.excel import write_workbook

    recs = [
        _rec(category="game", market="game_ml", selection="BBB ML",
             model_prob=0.62, fair_prob=0.50, ev=0.18, edge=0.12,
             market_american=110, tier=Tier.STRONG),
        _rec(category="game", market="game_total", selection="Over 8.5",
             model_prob=0.55, fair_prob=0.51, ev=0.05, edge=0.04,
             market_american=-105, tier=Tier.MODERATE),
        _rec(category="f5", market="f5_ml", selection="AAA F5",
             model_prob=0.40, fair_prob=0.52, ev=-0.10, edge=-0.12,
             market_american=-120, tier=Tier.PASS),
        _rec(category="pitcher", market="pitcher_k", selection="P Ks o5.5",
             model_prob=0.58, fair_prob=0.50, ev=0.22, edge=0.09,
             market_american=100, tier=Tier.STRONG, line=5.5),
        _rec(category="batter", market="batter_h", selection="B Hits o0.5",
             model_prob=0.45, fair_prob=0.55, ev=-0.06, edge=-0.10,
             market_american=-140, tier=Tier.PASS),
    ]
    out = tmp_path / "recs.xlsx"
    write_workbook(recs, out, date(2024, 7, 19))
    assert out.exists()

    wb = load_workbook(out)
    for sheet in (
        "Strong Buys", "Moderate Buys", "Fades",
        "Moneyline", "First-5 (F5)", "Pitcher Props", "Batter Props", "All",
    ):
        assert sheet in wb.sheetnames

    # Fades tab holds only the PASS picks the model is against.
    fades = wb["Fades"]
    assert fades.max_row == 3  # header + 2 fades
    # Family tabs partition by category: the F5 fade lands on the F5 tab.
    assert wb["First-5 (F5)"].max_row == 2
    assert wb["Batter Props"].max_row == 2

    # Every non-header cell on the All tab is shaded (no blank fill).
    all_ws = wb["All"]
    filled = [
        all_ws.cell(row=r, column=1).fill.fgColor.rgb
        for r in range(2, all_ws.max_row + 1)
    ]
    assert all(c not in (None, "00000000") for c in filled)


# ---- pitcher-outs: efficiency + pitch-count exit model ----
def _pitcher_statcast(
    *,
    dates,
    pitches_per_start,
    pa_per_start,
    f_strike_frac,
    bb_frac,
    gb_frac,
    hit_frac=0.0,
):
    """Build a minimal pitch-level Statcast slice for one pitcher.

    Each start gets ``pitches_per_start`` rows; ``pa_per_start`` of them end a PA
    (``events`` set). The first pitch of every PA is a 0-0 count, a fraction of
    which are strikes; a fraction of PAs are walks, a fraction are hits, and a
    fraction of batted balls are ground balls.
    """
    import pandas as pd

    rows = []
    for d in dates:
        pa_end_idx = set(range(pa_per_start))  # first N rows are PA-ending pitches
        n_walk = int(pa_per_start * bb_frac)
        n_hit = int(pa_per_start * hit_frac)
        for i in range(pitches_per_start):
            is_pa = i in pa_end_idx
            first_pitch = is_pa  # treat each PA-ender as its own 0-0 first pitch
            is_strike = first_pitch and (i < int(pa_per_start * f_strike_frac))
            is_walk = is_pa and (i < n_walk)
            is_hit = is_pa and (n_walk <= i < n_walk + n_hit)
            is_gb = is_pa and (not is_walk and not is_hit) and (
                i < int(pa_per_start * (bb_frac + hit_frac + gb_frac))
            )
            if not is_pa:
                event = None
            elif is_walk:
                event = "walk"
            elif is_hit:
                event = "single"
            else:
                event = "field_out"
            rows.append(
                {
                    "game_date": d,
                    "pitcher": 100,
                    "events": event,
                    "type": "S" if is_strike else "B",
                    "balls": 0 if first_pitch else 1,
                    "strikes": 0 if first_pitch else 1,
                    "bb_type": "ground_ball" if is_gb else None,
                }
            )
    return pd.DataFrame(rows)


def test_pitcher_efficiency_metrics_and_cap():
    from mlb_engine.features.efficiency import build_pitcher_efficiency

    df = _pitcher_statcast(
        dates=[date(2024, 4, 20), date(2024, 4, 25)],
        pitches_per_start=80,
        pa_per_start=20,
        f_strike_frac=0.65,
        bb_frac=0.10,
        gb_frac=0.50,
    )
    eff = build_pitcher_efficiency(df, date(2024, 5, 1), 28, manager_pitch_cap=110)
    assert eff.pa == 40  # 20 PA x 2 starts
    assert eff.pitches == 160
    assert abs(eff.pitches_per_pa - 4.0) < 1e-6
    assert abs(eff.f_strike_pct - 0.65) < 0.02
    assert abs(eff.bb_pct - 0.10) < 0.02
    # recent avg 80 pitches/start + buffer, under the 110 manager cap
    assert eff.pitch_cap == 88


def test_efficiency_scaler_tracks_command():
    from mlb_engine.features.efficiency import build_pitcher_efficiency

    common = dict(
        dates=[date(2024, 4, 20), date(2024, 4, 25)],
        pitches_per_start=90,
        pa_per_start=24,
        bb_frac=0.08,
        gb_frac=0.45,
    )
    efficient = build_pitcher_efficiency(
        _pitcher_statcast(f_strike_frac=0.72, **common), date(2024, 5, 1), 28, 100
    )
    wild = build_pitcher_efficiency(
        _pitcher_statcast(f_strike_frac=0.52, **common), date(2024, 5, 1), 28, 100
    )
    # High F-Strike% -> shorter counts -> lower expected P/PA prior.
    assert efficient.expected_pitches_per_pa() < wild.expected_pitches_per_pa()


def test_pitch_cap_and_efficiency_reduce_outs():
    lg = LEAGUE_RATES
    bat = [dict(lg) for _ in range(9)]
    # Same batters-faced ceiling; only the pitch budget/efficiency differ.
    deep = TeamSimConfig(
        bat_vs_starter=bat, bat_vs_pen=bat,
        starter_bf_cap=40, starter_pitch_cap=110, pitch_eff=0.9,
    )
    quick = TeamSimConfig(
        bat_vs_starter=bat, bat_vs_pen=bat,
        starter_bf_cap=40, starter_pitch_cap=80, pitch_eff=1.2,
    )
    res_deep = MonteCarlo(600, seed=3).simulate(deep, deep)
    res_quick = MonteCarlo(600, seed=3).simulate(quick, quick)
    # A tighter pitch cap + inefficiency pulls the starter earlier -> fewer outs.
    assert res_quick.pit["home"]["outs"].mean() < res_deep.pit["home"]["outs"].mean()


def test_gb_double_play_lifts_outs_per_start():
    lg = LEAGUE_RATES
    bat = [dict(lg) for _ in range(9)]
    no_dp = TeamSimConfig(
        bat_vs_starter=bat, bat_vs_pen=bat,
        starter_bf_cap=30, starter_pitch_cap=200, gb_dp_rate=0.0,
    )
    dp = TeamSimConfig(
        bat_vs_starter=bat, bat_vs_pen=bat,
        starter_bf_cap=30, starter_pitch_cap=200, gb_dp_rate=0.20,
    )
    res_no = MonteCarlo(600, seed=5).simulate(no_dp, no_dp)
    res_dp = MonteCarlo(600, seed=5).simulate(dp, dp)
    # Double plays record two outs on one PA -> more outs over a fixed BF window.
    assert res_dp.pit["home"]["outs"].mean() > res_no.pit["home"]["outs"].mean()


def test_ev_thresholds_per_market_override(monkeypatch):
    base = EVThresholds(min_edge=0.02, max_edge=0.08)
    # No override -> unchanged.
    assert base.for_market("batter_hr").min_edge == 0.02
    monkeypatch.setenv("MLBE_MAX_EDGE_PITCHER_OUTS", "0.06")
    monkeypatch.setenv("MLBE_MIN_EDGE_PITCHER_OUTS", "0.05")
    tuned = base.for_market("pitcher_outs")
    assert tuned.max_edge == 0.06
    assert tuned.min_edge == 0.05
    # Other markets still use the global cutoff.
    assert base.for_market("batter_hr").min_edge == 0.02


def test_calibration_min_samples_env(monkeypatch):
    from mlb_engine.calibration import Calibrator

    graded = [("pitcher_outs", 0.6, 1) for _ in range(30)]
    graded += [("pitcher_outs", 0.4, 0) for _ in range(30)]
    # Default 500 -> pitcher_outs too thin, no dedicated map.
    assert "pitcher_outs" not in Calibrator.fit(graded).maps
    monkeypatch.setenv("MLBE_CALIB_MIN_SAMPLES", "20")
    assert "pitcher_outs" in Calibrator.fit(graded).maps


def test_control_cap_factor_from_whip_bb9():
    from mlb_engine.features.efficiency import PitcherEfficiency

    def _eff(whip, bb9):
        return PitcherEfficiency(
            pa=100, pitches=390, pitches_per_pa=3.9, f_strike_pct=0.60,
            bb_pct=0.08, gb_pct=0.43, whip=whip, bb9=bb9, pitch_cap=95,
        )

    # Clean control -> no trim; high traffic -> ceiling trimmed below 1.0.
    assert _eff(1.0, 2.0).control_cap_factor() == 1.0
    assert _eff(1.7, 5.0).control_cap_factor() < 1.0


def test_high_whip_shrinks_pitch_cap():
    from mlb_engine.features.efficiency import build_pitcher_efficiency

    common = dict(
        dates=[date(2024, 4, 20), date(2024, 4, 25)],
        pitches_per_start=110,  # manager cap (95) binds the base ceiling
        pa_per_start=24,
        f_strike_frac=0.62,
        gb_frac=0.30,
    )
    clean = build_pitcher_efficiency(
        _pitcher_statcast(bb_frac=0.05, hit_frac=0.18, **common),
        date(2024, 5, 1), 28, 95,
    )
    wild = build_pitcher_efficiency(
        _pitcher_statcast(bb_frac=0.15, hit_frac=0.30, **common),
        date(2024, 5, 1), 28, 95,
    )
    assert clean.pitch_cap == 95
    assert wild.pitch_cap < clean.pitch_cap
    assert wild.whip > clean.whip and wild.bb9 > clean.bb9


def _lineup_statcast(ids, *, pitches, pa, day=date(2024, 4, 20)):
    import pandas as pd

    rows = []
    for bid in ids:
        for i in range(pitches):
            rows.append(
                {
                    "game_date": day,
                    "batter": bid,
                    "events": "field_out" if i < pa else None,
                }
            )
    return pd.DataFrame(rows)


def test_opponent_discipline_factor():
    from mlb_engine.features.efficiency import opponent_discipline_factor

    ids = [1, 2, 3]
    # Patient: 5.0 pitches/PA seen -> burns the budget faster (factor > 1).
    patient = _lineup_statcast(ids, pitches=100, pa=20)  # 60 PA, 5.0 P/PA
    f_patient = opponent_discipline_factor(patient, ids, date(2024, 5, 1), 28)
    assert f_patient > 1.0
    # Aggressive: 2.5 pitches/PA -> lets the starter work deep (factor < 1).
    aggressive = _lineup_statcast(ids, pitches=50, pa=20)  # 60 PA, 2.5 P/PA
    f_aggr = opponent_discipline_factor(aggressive, ids, date(2024, 5, 1), 28)
    assert f_aggr < 1.0
    # No batters / no data -> neutral.
    assert opponent_discipline_factor(patient, [], date(2024, 5, 1), 28) == 1.0


# ---- daily card ----
def _card_recs() -> list[Recommendation]:
    d = date(2026, 7, 24)

    def rec(market, selection, mp, am, ev, edge, tier, line=None):
        return Recommendation(
            game_date=d,
            game_pk=1,
            matchup="KC @ DET",
            category="game",
            market=market,
            selection=selection,
            model_prob=mp,
            line=line,
            market_american=am,
            ev=ev,
            edge=edge,
            tier=tier,
            park_name="Comerica Park",
            park_factor=98.0,
            carry_factor=0.85,
            roof="open",
            wx_summary="78F 40% wind 6mph (2 in to CF)",
            wx_hr_mult=0.97,
            wx_note="",
            xrd=0.8,
            xrd_sd=3.9,
        )

    return [
        rec("pitcher_k", "Tarik Skubal Ks o4.5", 0.80, -150, 0.10, 0.10, Tier.MODERATE, 4.5),
        rec("pitcher_k", "Beck Way Ks o4.5", 0.25, 200, -0.10, -0.05, Tier.PASS, 4.5),
        rec("pitcher_er", "Tarik Skubal ER o2.5", 0.16, 120, 0.05, 0.04, Tier.PASS, 2.5),
        rec("game_ml", "DET ML", 0.687, -285, -0.07, -0.02, Tier.PASS),
        rec("game_ml", "KC ML", 0.354, 254, 0.25, 0.07, Tier.STRONG),
        rec("game_total", "Under 7.5", 0.607, -108, 0.17, 0.077, Tier.STRONG, 7.5),
        rec("game_rl", "KC +1.5", 0.605, 116, 0.31, 0.13, Tier.STRONG, 1.5),
    ]


def test_card_build_picks_positive_ev_and_starters():
    from mlb_engine.output.card import build_cards

    cards = build_cards(_card_recs())
    assert len(cards) == 1
    c = cards[0]
    assert c.matchup == "KC @ DET"
    # Only positive-EV buys become plays; the DET ML fade and PASS props drop out.
    sels = {p.selection for p in c.plays}
    assert "DET ML" not in sels
    assert "KC +1.5" in sels and "Under 7.5" in sels
    # Implied probability is derived from the book price.
    kc_rl = next(p for p in c.plays if p.selection == "KC +1.5")
    assert 0.45 < (kc_rl.implied_prob or 0) < 0.47
    # Starters recovered with Skubal owning the strikeout edge.
    names = [s.name for s in c.starters]
    assert "Tarik Skubal" in names and "Beck Way" in names
    # Narrative is a list of paragraphs covering starters, park, weather, market.
    blob = " ".join(c.narrative)
    assert "Tarik Skubal" in blob
    assert "Comerica Park" in blob and "park factor" in blob
    assert "wind" in blob.lower()
    # DET ML is a negative-EV favorite -> flagged as too rich, not a play.
    assert "too rich" in blob
    # Expected run differential (xRD/G) is surfaced, home (DET) perspective.
    assert "Expected run differential: DET +0.8" in blob


def test_card_renderers_emit_md_and_html():
    from mlb_engine.output.card import build_cards, render_html, render_markdown

    cards = build_cards(_card_recs())
    md = render_markdown(cards, date(2026, 7, 24))
    html_body = render_html(cards, date(2026, 7, 24))
    assert "KC @ DET" in md and "KC +1.5" in md and "EV" in md
    assert html_body.startswith("<!DOCTYPE html>")
    assert "KC @ DET" in html_body and "<li>" in html_body


def test_email_raises_without_credentials():
    import pytest

    from mlb_engine.config import Config, Credentials
    from mlb_engine.output.email import EmailNotConfigured, send_card_email

    cfg = Config(creds=Credentials(gmail_user="x@y.com", gmail_app_password=None))
    with pytest.raises(EmailNotConfigured):
        send_card_email(cfg, subject="s", html_body="<p>", text_body="t", to="a@b.com")


# ---- audit report ----
def _ledger_entry(market, model_prob, result, *, odds=-110, ev=0.0, tier=Tier.PASS,
                  selection="Over 7.5", date_str="2026-07-23", pnl=None):
    from mlb_engine.audit.ledger import LedgerEntry

    if pnl is None:
        won = result == WIN
        pnl = (0.91 if odds == -110 else (odds / 100.0)) if won else (
            0.0 if result == PUSH else -1.0
        )
    return LedgerEntry(
        date=date_str, matchup="AAA @ BBB", category="game", market=market,
        selection=selection, line=7.5, book="dk", odds=odds, tier=tier.value,
        model_prob=model_prob, ev=ev, result=result, pnl=pnl,
    )


def _report_ledger():
    entries = []
    # A clean, profitable play market (pitcher_k): favored picks mostly win.
    entries += [_ledger_entry("pitcher_k", 0.7, WIN) for _ in range(8)]
    entries += [_ledger_entry("pitcher_k", 0.7, LOSS) for _ in range(2)]
    # A losing pocket (f5_total): favored picks mostly lose big.
    entries += [_ledger_entry("f5_total", 0.6, LOSS, odds=100) for _ in range(7)]
    entries += [_ledger_entry("f5_total", 0.6, WIN, odds=100) for _ in range(3)]
    # A market the model always fades -> abstain row.
    entries += [_ledger_entry("batter_hr", 0.2, LOSS, selection="Over 0.5") for _ in range(6)]
    return entries


def test_market_metrics_sorted_and_abstain_last():
    from mlb_engine.audit.ledger import market_metrics

    rows = market_metrics(_report_ledger())
    # highest ROI first, model-never-favored (n==0) row sinks to the bottom
    assert rows[0].tier == "pitcher_k"
    assert rows[-1].tier == "batter_hr" and rows[-1].n == 0


def test_report_classifies_and_renders():
    from mlb_engine.output.report import (
        FADE,
        NEUTRAL,
        PLAY,
        build_report_data,
        render_html_report,
        render_markdown_report,
    )

    data = build_report_data(
        _report_ledger(), period_label="Daily", subtitle="slate graded 2026-07-23"
    )
    verdicts = {r.market: r.verdict for r in data.rows}
    assert verdicts["pitcher_k"] == PLAY
    assert verdicts["f5_total"] == FADE
    assert verdicts["batter_hr"] == NEUTRAL  # abstained
    assert any(r.abstained for r in data.rows if r.market == "batter_hr")

    md = render_markdown_report(data)
    assert "## Executive summary" in md
    assert "## Market scorecard" in md
    assert "Min p to Play" in md
    assert "Recommendations" in md
    assert "Pitcher strikeouts" in md

    html_body = render_html_report(data)
    assert html_body.startswith("<!DOCTYPE html>")
    assert "Market scorecard" in html_body and "<table>" in html_body


def test_weekly_window_filters_to_seven_days():
    from mlb_engine.output.report import weekly_entries

    old = _ledger_entry("game_ml", 0.6, WIN, date_str="2026-07-01")
    recent = _ledger_entry("game_ml", 0.6, WIN, date_str="2026-07-23")
    kept = weekly_entries([old, recent], date(2026, 7, 23))
    assert recent in kept and old not in kept


# ---- contact-quality floor on batter props ----
def _floor_reg(xslg: float, k_pct: float, bbe: int = 40):
    from mlb_engine.features.regression import BatterRegression

    return BatterRegression(
        bbe=bbe, barrel_rate=0.08, hard_hit=0.40, sweet_spot=0.33, bat_speed=72.0,
        max_ev=108.0, whiff=0.24, zone_contact=0.82, xba=0.25, xslg=xslg,
        babip=0.29, woba=0.32, xwoba=0.32, k_pct=k_pct, bb_pct=0.08,
    )


def test_power_floor_reason_gates():
    from mlb_engine.models.selectors import power_floor_reason

    kw = {"xslg_floor": 0.400, "k_ceiling": 0.25}
    # Power markets: below the xSLG floor -> excluded; above -> kept.
    assert power_floor_reason(_floor_reg(0.330, 0.20), "HR", **kw)
    assert power_floor_reason(_floor_reg(0.330, 0.20), "TB", **kw)
    assert power_floor_reason(_floor_reg(0.480, 0.20), "HR", **kw) is None
    # A high-K slugger is NOT gated out of the power markets.
    assert power_floor_reason(_floor_reg(0.520, 0.32), "HR", **kw) is None
    # Contact markets: above the K% ceiling -> excluded; below -> kept.
    assert power_floor_reason(_floor_reg(0.480, 0.30), "H", **kw)
    assert power_floor_reason(_floor_reg(0.480, 0.30), "HRR", **kw)
    assert power_floor_reason(_floor_reg(0.480, 0.18), "1B", **kw) is None
    # Never gate on a thin sample or a missing (NaN) feature.
    assert power_floor_reason(_floor_reg(0.330, 0.20, bbe=5), "HR", **kw) is None
    assert power_floor_reason(_floor_reg(0.480, float("nan")), "H", **kw) is None


def test_batter_regression_k_bb_pct():
    import pandas as pd

    from mlb_engine.features.regression import build_batter_regression

    events = ["strikeout"] * 3 + ["walk"] + ["single"] * 4 + ["field_out"] * 2
    n = len(events)
    df = pd.DataFrame(
        {
            "events": events,
            "batter": [1] * n,
            "launch_speed": [float("nan")] * 4 + [95.0] * 4 + [80.0] * 2,
            "launch_angle": [float("nan")] * n,
            "launch_speed_angle": [float("nan")] * n,
            "bat_speed": [float("nan")] * n,
            "description": (
                ["swinging_strike"] * 3 + ["ball"] + ["hit_into_play"] * 6
            ),
            "estimated_ba_using_speedangle": [float("nan")] * n,
            "estimated_woba_using_speedangle": [float("nan")] * n,
            "woba_value": [float("nan")] * n,
            "zone": [5] * n,
        }
    )
    breg = build_batter_regression(df)
    assert breg.k_pct == 0.3  # 3 / 10
    assert breg.bb_pct == 0.1  # 1 / 10


# ---- singles "Under" NPV screen ----
def test_singles_under_score_flags_tto_and_flyball():
    from mlb_engine.features.singles_under import (
        SINGLES_UNDER_STRONG,
        SinglesUnderProfile,
        singles_under_score,
    )

    # Both fitted flags: he misses the ball, and lifts it when he doesn't.
    slugger = SinglesUnderProfile(
        pa=95, bip=60, k_pct=0.31, bb_pct=0.14, z_swing=0.55, avg_la=22.0,
        barrel=0.18, hard_hit=0.52, pull_rate=0.40,
    )
    score, reasons = singles_under_score(slugger)
    assert score >= SINGLES_UNDER_STRONG
    assert any("K%" in r for r in reasons)
    assert any("fly-ball" in r for r in reasons)

    # The dropped flags no longer score: a passive, barrel-heavy, pull-happy
    # bat who makes contact on the ground trips nothing (none of those measures
    # predicted a no-single game out of time).
    unfitted = SinglesUnderProfile(
        pa=95, bip=60, k_pct=0.18, bb_pct=0.15, z_swing=0.52, avg_la=3.0,
        barrel=0.19, hard_hit=0.55, pull_rate=0.50,
    )
    assert singles_under_score(unfitted) == (0.0, [])

    # A contact, line-drive hitter trips no flags.
    contact = SinglesUnderProfile(
        pa=95, bip=70, k_pct=0.14, bb_pct=0.07, z_swing=0.72, avg_la=10.0,
        barrel=0.05, hard_hit=0.35, pull_rate=0.38,
    )
    cscore, creasons = singles_under_score(contact)
    assert cscore == 0.0 and creasons == []


def test_singles_under_thin_sample_is_neutral():
    from mlb_engine.features.singles_under import (
        SinglesUnderProfile,
        singles_under_score,
    )

    thin = SinglesUnderProfile(
        pa=10, bip=6, k_pct=0.40, bb_pct=0.20, z_swing=0.50, avg_la=25.0,
        barrel=0.20, hard_hit=0.55, pull_rate=0.50,
    )
    assert not thin.has_data
    assert singles_under_score(thin) == (0.0, [])


def test_build_singles_under_from_statcast():
    import numpy as np
    import pandas as pd

    from mlb_engine.features.singles_under import build_singles_under

    # 30 K, 15 BB, 55 batted balls (all pulled fly balls) over 100 PA.
    events = ["strikeout"] * 30 + ["walk"] * 15 + ["field_out"] * 55
    n = len(events)
    ls = [float("nan")] * 45 + [100.0] * 55  # only batted balls have exit velo
    la = [float("nan")] * 45 + [25.0] * 55  # steep fly-ball angle
    zone = [5] * 60 + [12] * 40  # 60 in-zone, 40 out
    # Batted balls pulled to LF for a RHB (hc_x < origin, deep).
    hc_x = [float("nan")] * 45 + [80.0] * 55
    hc_y = [float("nan")] * 45 + [100.0] * 55
    df = pd.DataFrame(
        {
            "events": events,
            "description": ["swinging_strike"] * 45 + ["hit_into_play"] * 55,
            "launch_speed": ls,
            "launch_angle": la,
            "launch_speed_angle": [float("nan")] * n,
            "zone": zone,
            "hc_x": hc_x,
            "hc_y": hc_y,
        }
    )
    p = build_singles_under(df, "R")
    assert p.pa == 100
    assert np.isclose(p.k_pct, 0.30)
    assert np.isclose(p.bb_pct, 0.15)
    assert p.avg_la == 25.0
    assert p.bip == 55


# ---- SIERA from Statcast ----
def test_pitcher_siera_ace_vs_scrub_and_empty():
    import pandas as pd

    from mlb_engine.features.siera import pitcher_siera

    def _mk(n_pa, k, bb, gb, fb, pu):
        # n_pa PA: k strikeouts, bb walks, rest batted balls split gb/fb/pu.
        events = (
            ["strikeout"] * k
            + ["walk"] * bb
            + ["field_out"] * (n_pa - k - bb)
        )
        bip = n_pa - k - bb
        bt = (
            [None] * (k + bb)
            + ["ground_ball"] * gb
            + ["fly_ball"] * fb
            + ["popup"] * pu
            + [None] * (bip - gb - fb - pu)
        )
        return pd.DataFrame({"events": events, "bb_type": bt})

    # High-K, low-BB, grounder-leaning -> low (elite) SIERA.
    ace = pitcher_siera(_mk(200, 76, 10, 60, 30, 5))
    # Low-K, high-BB, fly-ball-leaning -> high SIERA.
    scrub = pitcher_siera(_mk(200, 24, 22, 30, 55, 12))
    assert ace.has_data and scrub.has_data
    assert ace.siera < 3.4 < scrub.siera
    assert 1.0 < ace.siera < 3.4
    assert 4.0 < scrub.siera < 6.5

    # No plate appearances -> neutral, not trusted.
    empty = pitcher_siera(pd.DataFrame({"events": [None, None], "bb_type": [None, None]}))
    assert empty.pa == 0 and not empty.has_data
    assert empty.siera != empty.siera  # nan


def test_siera_matchup_gate_helpers():
    from mlb_engine.features.siera import Siera, faces_ace, faces_scrub

    ace = Siera(pa=180, so_rate=0.35, bb_rate=0.05, net_gb_rate=0.1, siera=2.1)
    mid = Siera(pa=150, so_rate=0.22, bb_rate=0.08, net_gb_rate=0.05, siera=3.9)
    scrub = Siera(pa=140, so_rate=0.13, bb_rate=0.10, net_gb_rate=0.06, siera=4.9)
    thin = Siera(pa=20, so_rate=0.35, bb_rate=0.05, net_gb_rate=0.1, siera=2.0)

    # Ace triggers the over-exclude; mid/scrub do not.
    assert faces_ace(ace, 3.4) and not faces_ace(mid, 3.4) and not faces_ace(scrub, 3.4)
    # Scrub triggers the under-veto; ace/mid do not.
    assert faces_scrub(scrub, 4.4) and not faces_scrub(mid, 4.4) and not faces_scrub(ace, 4.4)
    # Thin sample and missing data stay neutral on both sides.
    assert not faces_ace(thin, 3.4) and not faces_scrub(thin, 4.4)
    assert not faces_ace(None, 3.4) and not faces_scrub(None, 4.4)


def test_card_render_pdf_produces_pdf_bytes():
    from mlb_engine.output.card import build_cards, render_html, render_pdf

    cards = build_cards(_card_recs())
    pdf = render_pdf(render_html(cards, date(2026, 7, 24)))
    assert pdf[:5] == b"%PDF-"
    assert len(pdf) > 1000


def test_email_attachments_infer_mime_type(monkeypatch):
    import smtplib

    from mlb_engine.config import Config, Credentials
    from mlb_engine.output import email as email_mod

    captured = {}

    class _FakeSMTP:
        def __init__(self, *a, **k):
            pass

        def __enter__(self):
            return self

        def __exit__(self, *a):
            return False

        def login(self, *a):
            pass

        def send_message(self, msg):
            captured["msg"] = msg

    monkeypatch.setattr(smtplib, "SMTP_SSL", _FakeSMTP)
    cfg = Config(creds=Credentials(gmail_user="x@y.com", gmail_app_password="pw"))
    email_mod.send_card_email(
        cfg,
        subject="s",
        html_body="<p>hi</p>",
        text_body="t",
        to="a@b.com",
        attachments=[("card.pdf", b"%PDF-1.4 x"), ("bets.xlsx", b"PK\x03\x04")],
    )
    types = {
        part.get_filename(): part.get_content_type()
        for part in captured["msg"].iter_attachments()
    }
    assert types["card.pdf"] == "application/pdf"
    assert types["bets.xlsx"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def test_reaching_on_an_error_is_neither_an_out_nor_a_plate_appearance_outcome():
    """The bucketer's ``else`` called it an out. The batter is standing on first.

    It cannot go anywhere in the seven -- not a hit, not a walk, not an out -- so the
    plate appearance leaves the denominator entirely and the run models add the
    league's rate back. The alternative bucketings each corrupt a market that gets
    bet: ``1B`` inflates batter_h and batter_tb, ``BB`` inflates batter_bb, ``OUT``
    is the status quo and shortens innings that in fact continued.
    """
    import pandas as pd

    from mlb_engine.features.rolling import _bucket_counts

    counts = _bucket_counts(pd.Series(["field_error", "single", "field_out"]))
    assert counts["OUT"] == 1
    assert counts["1B"] == 1
    assert counts["BB"] == 0
    assert sum(counts.values()) == 2, "the error is not in the denominator"


def test_a_plate_appearance_the_third_out_cut_short_is_not_an_out():
    """``truncated_pa`` is a PA that never finished: caught stealing, or the game ended.

    These rows sit at 0-0, 1-1, 2-2 and end on a ball or a called strike, so the
    batter did nothing -- he was not retired, and he does not have a plate appearance.
    """
    import pandas as pd

    from mlb_engine.features.rolling import _bucket_counts

    counts = _bucket_counts(pd.Series(["truncated_pa", "truncated_pa", "strikeout"]))
    assert counts["K"] == 1
    assert counts["OUT"] == 0
    assert sum(counts.values()) == 1
