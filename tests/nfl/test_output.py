"""The weekly package: what it says, and what it still ships when a part fails.

The contracts worth holding, both learned on the MLB side:

* a vetoed price appears on the card as a veto, never as a silent absence;
* one optional artifact failing (no WeasyPrint libraries, no SMTP credentials)
  costs the caller that artifact and nothing else -- the workbook is written
  either way.
"""

from __future__ import annotations

import argparse
from io import BytesIO

import pytest
from openpyxl import load_workbook

from nfl_engine import cli
from nfl_engine.audit.ledger import LedgerEntry, save_ledger
from nfl_engine.output.card import build_card, render_html, render_markdown
from nfl_engine.output.email import EmailNotConfigured
from nfl_engine.output.excel import build_workbook

SEASON, WEEK = 2026, 1


def row(
    matchup: str,
    market: str,
    side: str,
    *,
    ev_fair: float,
    line: float | None = None,
    screens: str = "",
    source: str = "engine",
    clv: float | None = None,
    result: str = "",
    kickoff: str = "2026-09-13T17:00:00Z",
) -> LedgerEntry:
    return LedgerEntry(
        season=SEASON,
        week=WEEK,
        date="2026-09-13",
        matchup=matchup,
        market=market,
        side=side,
        line=line,
        book="dk",
        odds=-110.0,
        opposite_odds=-110.0,
        tier="Strong buy" if ev_fair > 0.03 else "Moderate buy",
        model_prob=0.55,
        fair_prob=0.52,
        ev_model=0.05,
        ev_fair=ev_fair,
        paired_books=3,
        screens=screens,
        result=result,
        clv=clv,
        source=source,
        captured_at="2026-09-09T18:00:00Z",
        kickoff_utc=kickoff,
    )


@pytest.fixture
def entries() -> list[LedgerEntry]:
    return [
        row("BUF @ KC", "moneyline", "KC", ev_fair=0.02, clv=-0.01),
        row("BUF @ KC", "spread", "KC", ev_fair=0.05, line=-2.5, clv=0.03),
        row("BUF @ KC", "total", "over", ev_fair=0.01, line=47.5, screens="thin_market"),
        row("NYJ @ MIA", "moneyline", "MIA", ev_fair=0.10, screens="longshot;disagreement"),
        row("BUF @ KC", "moneyline", "KC", ev_fair=0.40, source="market"),
    ]


def test_the_card_lists_plays_by_execution_edge_and_keeps_the_vetoes(
    entries: list[LedgerEntry],
) -> None:
    card = build_card(entries, season=SEASON, week=WEEK)
    kc = next(game for game in card.games if game.matchup == "BUF @ KC")
    assert [play.market for play in kc.plays] == ["spread", "moneyline"]
    assert kc.vetoes == {"thin_market": 1}
    mia = next(game for game in card.games if game.matchup == "NYJ @ MIA")
    assert mia.plays == []
    assert mia.vetoes == {"longshot": 1, "disagreement": 1}


def test_a_benchmark_row_is_never_presented_as_one_of_our_plays(
    entries: list[LedgerEntry],
) -> None:
    card = build_card(entries, season=SEASON, week=WEEK)
    assert all(play.ev_fair != 0.40 for play in card.plays())
    assert card.selections == 4


def test_a_game_with_no_surviving_price_still_appears_on_the_card(
    entries: list[LedgerEntry],
) -> None:
    card = build_card(entries, season=SEASON, week=WEEK)
    text = render_markdown(card)
    assert "NYJ @ MIA" in text
    assert "every price on this game was vetoed" in text
    assert "longshot x1" in text
    assert "paper" in text.lower()
    page = render_html(card)
    assert "NYJ @ MIA" in page and "Vetoed:" in page


def test_another_week_is_not_pulled_into_this_one(entries: list[LedgerEntry]) -> None:
    other = row("CHI @ GB", "moneyline", "GB", ev_fair=0.09)
    other.week = 2
    card = build_card([*entries, other], season=SEASON, week=WEEK)
    assert all(game.matchup != "CHI @ GB" for game in card.games)


def test_the_record_is_absent_until_something_has_been_graded(
    entries: list[LedgerEntry],
) -> None:
    assert build_card(entries, season=SEASON, week=WEEK).record == []
    graded = row("CHI @ GB", "moneyline", "GB", ev_fair=0.09, result="win", clv=0.02)
    graded.pnl = 0.909
    card = build_card([*entries, graded], season=SEASON, week=WEEK)
    assert any(metric.label == "ALL" for metric in card.record)


def test_the_workbook_keeps_every_selection_and_sorts_clv_worst_first(
    entries: list[LedgerEntry],
) -> None:
    card = build_card(entries, season=SEASON, week=WEEK)
    book = load_workbook(BytesIO(build_workbook(card, entries)))
    assert book.sheetnames == ["Plays", "Selections", "Record", "CLV"]
    # Every priced row, vetoes and benchmark included: a rejected winner is only
    # diagnosable if it was written down.
    assert book["Selections"].max_row == 1 + len(entries)
    assert book["Plays"].cell(row=2, column=3).value == "KC -2.5"
    clv = book["CLV"]
    assert [clv.cell(row=r, column=9).value for r in (2, 3)] == [-0.01, 0.03]


def _configure(monkeypatch: pytest.MonkeyPatch, tmp_path, entries: list[LedgerEntry]):
    root = tmp_path / "state"
    root.mkdir()
    out = tmp_path / "output"
    path = cli.ledger_path(root)
    save_ledger(path, entries)
    monkeypatch.setattr(cli, "ledger_path", lambda: path)
    monkeypatch.setattr(cli, "output_dir", lambda: out)
    return out


def args(**overrides: object) -> argparse.Namespace:
    base = {"season": SEASON, "week": WEEK, "email": False, "to": None}
    base.update(overrides)
    return argparse.Namespace(**base)


def test_a_box_without_weasyprint_still_gets_the_workbook(
    monkeypatch: pytest.MonkeyPatch, tmp_path, entries: list[LedgerEntry], capsys
) -> None:
    out = _configure(monkeypatch, tmp_path, entries)

    def no_pdf(_html: str) -> bytes:
        raise OSError("cannot load library 'libgobject-2.0-0'")

    monkeypatch.setattr(cli, "render_pdf", no_pdf)
    assert cli.cmd_card(args()) == 0
    written = {path.suffix for path in out.iterdir()}
    assert written == {".md", ".html", ".xlsx"}
    assert "PDF not rendered" in capsys.readouterr().out


def test_missing_credentials_cost_the_email_not_the_package(
    monkeypatch: pytest.MonkeyPatch, tmp_path, entries: list[LedgerEntry], capsys
) -> None:
    out = _configure(monkeypatch, tmp_path, entries)
    monkeypatch.setattr(cli, "render_pdf", lambda _html: b"%PDF-1.7")

    def unconfigured(*_a: object, **_k: object) -> str:
        raise EmailNotConfigured("GMAIL_APP_PASSWORD is not set")

    monkeypatch.setattr(cli, "send_package", unconfigured)
    assert cli.cmd_card(args(email=True)) == 0
    assert (out / f"NFL_{SEASON}_Week{WEEK:02d}.xlsx").exists()
    assert "email not sent" in capsys.readouterr().out


def test_the_email_carries_the_card_the_workbook_and_the_pdf(
    monkeypatch: pytest.MonkeyPatch, tmp_path, entries: list[LedgerEntry]
) -> None:
    _configure(monkeypatch, tmp_path, entries)
    monkeypatch.setattr(cli, "render_pdf", lambda _html: b"%PDF-1.7")
    sent: dict[str, object] = {}

    def capture_send(_cfg: object, **kwargs: object) -> str:
        sent.update(kwargs)
        return "someone@example.com"

    monkeypatch.setattr(cli, "send_package", capture_send)
    assert cli.cmd_card(args(email=True, to="someone@example.com")) == 0
    names = [name for name, _data in sent["attachments"]]
    assert names == [
        f"NFL_{SEASON}_Week{WEEK:02d}.md",
        f"NFL_{SEASON}_Week{WEEK:02d}.xlsx",
        f"NFL_{SEASON}_Week{WEEK:02d}.pdf",
    ]
    assert "paper" in str(sent["subject"]).lower()


def test_a_week_that_was_never_priced_writes_nothing(
    monkeypatch: pytest.MonkeyPatch, tmp_path, entries: list[LedgerEntry], capsys
) -> None:
    out = _configure(monkeypatch, tmp_path, entries)
    assert cli.cmd_card(args(week=9)) == 0
    assert not out.exists()
    assert "no priced rows" in capsys.readouterr().out
