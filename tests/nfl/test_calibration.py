"""What turns a calibration map on, and what has to stay untouched when it is off.

The layer ships applying nothing, so most of what matters here is the refusals: a
map is fitted out of time, judged per market, and ignored the moment its basis or
its evidence stops holding.
"""

from __future__ import annotations

import json
from dataclasses import replace
from datetime import date as Date
from io import BytesIO

import pytest
from openpyxl import load_workbook

from nfl_engine import calibration
from nfl_engine.calibration import (
    BASIS,
    MIN_FIT_ROWS,
    MIN_HOLDOUT_ROWS,
    Calibrator,
    IsotonicMap,
    MarketFit,
    fit,
    read_maps,
    write_maps,
)
from nfl_engine.market.board import GameOdds, MarketQuote
from nfl_engine.market.ev import MONEYLINE, SPREAD, TOTAL
from nfl_engine.models.drives import DriveSim
from nfl_engine.output.card import build_card, render_html, render_markdown
from nfl_engine.output.excel import build_workbook
from nfl_engine.pipeline import price_slate
from nfl_engine.schemas import Game, TeamGameInfo

HOME, AWAY = "KC", "BUF"
MATCHUP = f"{AWAY} @ {HOME}"


def game() -> Game:
    return Game(
        game_id="1",
        season=2025,
        week=3,
        game_date=Date(2025, 9, 21),
        home=TeamGameInfo(name="Kansas City Chiefs", abbrev=HOME, is_home=True),
        away=TeamGameInfo(name="Buffalo Bills", abbrev=AWAY, is_home=False),
    )


def board() -> dict[str, GameOdds]:
    odds = GameOdds(matchup=MATCHUP)
    for index in range(3):
        book = f"book{index}"
        odds.add_ml(HOME, MarketQuote(book=book, american=-150, opposite_american=130))
        odds.add_ml(AWAY, MarketQuote(book=book, american=130, opposite_american=-150))
        odds.add_spread(-3.0, HOME, MarketQuote(book=book, american=-110, opposite_american=-110))
        odds.add_spread(-3.0, AWAY, MarketQuote(book=book, american=-110, opposite_american=-110))
        odds.add_total(47.5, True, MarketQuote(book=book, american=-110, opposite_american=-110))
        odds.add_total(47.5, False, MarketQuote(book=book, american=-110, opposite_american=-110))
    return {MATCHUP: odds}


LEVELS, REPS = 90, 40
PER_SEASON = LEVELS * REPS


def biased_rows(
    market: str, *, seasons: range, bias: float
) -> list[tuple[int, str, float, int]]:
    """A market that says ``p`` and wins ``p - bias``: exactly what a map can fix.

    Every probability level gets the same number of rows and the exact realised
    rate its bias implies, so acceptance is a statement about the code rather than
    about a seed. ``bias=0`` is therefore a genuinely calibrated market.
    """
    rows: list[tuple[int, str, float, int]] = []
    for season in seasons:
        for level in range(LEVELS):
            prob = 0.05 + 0.9 * (level / (LEVELS - 1))
            truth = min(max(prob - bias, 0.01), 0.99)
            wins = round(truth * REPS)
            rows.extend(
                (season, market, prob, 1 if rep < wins else 0) for rep in range(REPS)
            )
    return rows


def fitted_map() -> IsotonicMap:
    return IsotonicMap([0.2, 0.8], [0.1, 0.7])


def accepted_fit(market: str = SPREAD) -> MarketFit:
    return MarketFit(
        market=market,
        curve=fitted_map(),
        accepted=True,
        n_fit=MIN_FIT_ROWS,
        n_holdout=MIN_HOLDOUT_ROWS,
        brier_identity=0.25,
        brier_fitted=0.24,
        gain=0.01,
        gain_se=0.001,
        fit_seasons="2007-2019",
        holdout_seasons="2020-2025",
    )


# -- fitting ---------------------------------------------------------------
def test_the_split_is_by_season_so_a_curve_is_never_graded_on_a_season_it_saw() -> None:
    rows = biased_rows(SPREAD, seasons=range(2017, 2022), bias=0.08)
    got = fit(rows, cutoff=2019)[SPREAD]
    assert got.fit_seasons == "2017-2019"
    assert got.holdout_seasons == "2020-2021"
    assert got.n_fit == 3 * PER_SEASON
    assert got.n_holdout == 2 * PER_SEASON


def test_a_market_that_lies_by_eight_points_gets_a_map() -> None:
    rows = biased_rows(MONEYLINE, seasons=range(2017, 2022), bias=0.08)
    got = fit(rows, cutoff=2019)[MONEYLINE]
    assert got.accepted, got
    assert got.verdict() == "applied"
    assert got.brier_fitted < got.brier_identity
    assert got.curve.apply(0.60) < 0.60


def test_a_market_that_is_already_calibrated_is_left_alone() -> None:
    rows = biased_rows(TOTAL, seasons=range(2017, 2022), bias=0.0)
    got = fit(rows, cutoff=2019)[TOTAL]
    assert not got.accepted
    assert got.verdict() == "no correction (identity wins)"


def test_one_weak_market_does_not_retire_the_market_beside_it() -> None:
    rows = biased_rows(MONEYLINE, seasons=range(2017, 2022), bias=0.08)
    rows += biased_rows(TOTAL, seasons=range(2017, 2022), bias=0.0)
    fits = fit(rows, cutoff=2019)
    assert fits[MONEYLINE].accepted
    assert not fits[TOTAL].accepted
    maps = Calibrator.from_fits(fits)
    assert maps.applied_markets() == (MONEYLINE,)
    # The uncorrected market takes no correction at all -- not the other market's.
    assert maps.apply(TOTAL, 0.60) == 0.60


def test_a_thin_market_is_measured_but_never_fitted() -> None:
    rows = biased_rows(SPREAD, seasons=range(2019, 2020), bias=0.08)[:600]
    rows += biased_rows(SPREAD, seasons=range(2020, 2021), bias=0.08)[:300]
    got = fit(rows, cutoff=2019)[SPREAD]
    assert not got.accepted
    assert got.verdict() == "not fitted (thin)"
    assert got.curve.x == []
    assert got.n_fit and got.n_holdout


# -- persistence and refusal ----------------------------------------------
def test_every_measured_market_is_stored_applied_or_not(tmp_path) -> None:
    rejected = replace(accepted_fit(TOTAL), accepted=False, gain=0.0001)
    fits = {SPREAD: accepted_fit(), TOTAL: rejected}
    path = tmp_path / "calibration.json"
    write_maps(path, fits)
    stored = json.loads(path.read_text())
    assert sorted(stored["markets"]) == [SPREAD, TOTAL]
    assert stored["markets"][TOTAL]["n_holdout"] == MIN_HOLDOUT_ROWS
    loaded = read_maps(path)
    assert loaded.applied_markets() == (SPREAD,)
    assert sorted(loaded.fits) == [SPREAD, TOTAL]
    assert loaded.fits[TOTAL].gain == pytest.approx(0.0001)


def test_a_map_fitted_on_another_basis_is_ignored_until_it_is_refitted(tmp_path) -> None:
    path = tmp_path / "calibration.json"
    write_maps(path, {SPREAD: replace(accepted_fit(), basis="older-basis-2019.01")})
    loaded = read_maps(path)
    assert loaded.applied_markets() == ()
    assert loaded.fits[SPREAD].verdict() == "retired (basis)"
    assert loaded.apply(SPREAD, 0.62) == 0.62


def test_a_stored_gain_that_no_longer_clears_the_bar_retires_without_a_refit(tmp_path) -> None:
    path = tmp_path / "calibration.json"
    write_maps(path, {SPREAD: replace(accepted_fit(), gain=0.0001)})
    assert read_maps(path).applied_markets() == ()


def test_an_unreadable_map_prices_uncalibrated_rather_than_raising(tmp_path, monkeypatch) -> None:
    path = tmp_path / "calibration.json"
    path.write_text("{not json")
    monkeypatch.setattr(calibration, "shipped_path", lambda: path)
    assert calibration.load().applied_markets() == ()


def test_a_missing_map_file_prices_uncalibrated(tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(calibration, "shipped_path", lambda: tmp_path / "absent.json")
    assert calibration.load().applied_markets() == ()


def test_the_shipped_file_is_on_this_basis_and_corrects_nothing() -> None:
    """The measurement ships; the correction does not. Both on purpose."""
    shipped = calibration.load()
    assert shipped.applied_markets() == ()
    assert sorted(shipped.fits) == [MONEYLINE, SPREAD, TOTAL]
    assert all(f.basis == BASIS for f in shipped.fits.values())
    assert all(f.n_holdout >= MIN_HOLDOUT_ROWS for f in shipped.fits.values())
    assert shipped.stamp() == f"calibration: none applied (basis {BASIS})"


# -- pricing ---------------------------------------------------------------
def test_pricing_without_a_map_is_the_pricing_that_shipped_before_this_layer() -> None:
    sim = DriveSim(n_sims=4000)
    plain = price_slate([game()], board(), sim=sim)[0].bets
    through = price_slate([game()], board(), sim=DriveSim(n_sims=4000), calibrator=Calibrator())[
        0
    ].bets
    assert [b.model_prob for b in plain] == [b.model_prob for b in through]


def test_a_map_moves_its_own_market_and_leaves_the_execution_price_alone() -> None:
    maps = Calibrator(maps={SPREAD: IsotonicMap([0.0, 1.0], [0.0, 0.5])})
    plain = price_slate([game()], board(), sim=DriveSim(n_sims=4000))[0].bets
    priced = price_slate(
        [game()], board(), sim=DriveSim(n_sims=4000), calibrator=maps
    )[0].bets
    before = {(b.market, b.side, b.book): b for b in plain}
    for bet in priced:
        was = before[(bet.market, bet.side, bet.book)]
        if bet.market == SPREAD:
            assert bet.model_prob == pytest.approx(was.model_prob / 2, abs=1e-6)
        else:
            assert bet.model_prob == pytest.approx(was.model_prob)
        # The de-vigged consensus is the market's price, not the model's opinion.
        assert bet.fair_prob == pytest.approx(was.fair_prob)
        assert bet.ev_fair == pytest.approx(was.ev_fair)
        assert bet.push_prob == pytest.approx(was.push_prob)


def test_the_card_says_what_pricing_did_without_doing_any_of_it() -> None:
    card = build_card([], season=2025, week=3, calibration="calibration: none applied (basis x)")
    assert card.calibration in render_markdown(card)
    assert "none applied" in render_html(card)
    assert build_card([], season=2025, week=3).calibration == ""

    book = load_workbook(BytesIO(build_workbook(card, [])))
    assert any(
        cell == card.calibration for row in book["Record"].iter_rows(values_only=True) for cell in row
    )


def test_a_halved_probability_is_a_disagreement_the_screens_can_see() -> None:
    """Calibration lands before the screens, so a corrected price can be vetoed."""
    maps = Calibrator(maps={SPREAD: IsotonicMap([0.0, 1.0], [0.0, 0.5])})
    priced = price_slate([game()], board(), sim=DriveSim(n_sims=4000), calibrator=maps)[0].bets
    spreads = [bet for bet in priced if bet.market == SPREAD]
    assert spreads
    assert all(not bet.is_bet() for bet in spreads)
