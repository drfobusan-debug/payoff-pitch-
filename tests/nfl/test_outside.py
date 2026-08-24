"""An outside forecast beside ours, and the wall between it and every price.

The contracts here are the MLB benchmark lessons, restated for FPI:

* a benchmark row is read from its own source, graded like anything else, and
  excluded from every engine measurement -- record, ROI, CLV, tiers;
* it carries no invented price, and therefore no units and no ROI of its own;
* the published number is stored exactly as published, including the pair that
  does not sum to 100;
* a missing projection costs that game's benchmark and nothing else;
* the card shows agreement and disagreement, and shows the same play either way.
"""

from __future__ import annotations

import argparse
import json
from dataclasses import asdict
from io import BytesIO

import pandas as pd
import pytest
from openpyxl import load_workbook

from nfl_engine import cli
from nfl_engine.audit.ledger import (
    ENGINE,
    PAPER,
    LedgerEntry,
    grade,
    load_ledger,
    market_metrics,
    save_ledger,
    tier_metrics,
)
from nfl_engine.audit.outside import (
    FPI,
    benchmark_metrics,
    entries_from_fpi,
    head_to_head,
)
from nfl_engine.data import espn
from nfl_engine.output.card import build_card, render_html, render_markdown
from nfl_engine.output.excel import build_workbook

SEASON, WEEK = 2024, 1


def fpi_game(
    home: str = "KC",
    away: str = "BAL",
    *,
    home_prob: float = 58.167,
    home_margin: float = 3.201,
) -> espn.FpiGame:
    return espn.FpiGame(
        season=SEASON,
        week=WEEK,
        date="2024-09-05",
        matchup=f"{away} @ {home}",
        home=home,
        away=away,
        home_prob=home_prob,
        home_margin=home_margin,
    )


def engine_row(
    matchup: str = "BAL @ KC",
    side: str = "KC",
    *,
    market: str = "moneyline",
    screens: str = "",
    result: str = "",
    tier: str = "Strong buy",
) -> LedgerEntry:
    return LedgerEntry(
        season=SEASON,
        week=WEEK,
        date="2024-09-05",
        matchup=matchup,
        market=market,
        side=side,
        line=None,
        book="dk",
        odds=-140.0,
        opposite_odds=120.0,
        tier=tier,
        model_prob=0.62,
        fair_prob=0.57,
        ev_model=0.05,
        ev_fair=0.02,
        paired_books=3,
        screens=screens,
        result=result,
        pnl=0.714 if result == "win" else (-1.0 if result == "loss" else 0.0),
        clv=0.01,
    )


# -- the published number -------------------------------------------------
def test_published_probabilities_are_stored_as_published() -> None:
    """ESPN's two sides need not sum to 100, and are not made to."""
    home = fpi_game(home_prob=58.167)
    away = fpi_game(home_prob=41.5)
    assert home.home_prob + (100.0 - away.home_prob) != pytest.approx(100.0)
    rows = entries_from_fpi([home])
    assert rows[0].model_prob == pytest.approx(0.58167)


def test_pick_is_the_side_the_projection_favours() -> None:
    assert fpi_game(home_prob=58.167).pick == "KC"
    assert fpi_game(home_prob=41.5).pick == "BAL"
    assert fpi_game(home_prob=41.5).pick_prob == pytest.approx(0.585)


def test_an_exactly_even_projection_makes_no_call() -> None:
    """No side, so no row: a coin flip is not a forecast to grade."""
    assert fpi_game(home_prob=50.0).pick == ""
    assert entries_from_fpi([fpi_game(home_prob=50.0)]) == []


# -- what a benchmark row is ---------------------------------------------
def test_benchmark_rows_carry_their_source_and_no_price() -> None:
    row = entries_from_fpi([fpi_game()], captured_at="2024-09-04T12:00:00Z")[0]
    assert (row.source, row.mode) == (FPI, PAPER)
    assert row.odds is None and row.opposite_odds is None
    assert row.fair_prob is None and row.ev_model is None and row.ev_fair is None
    assert row.captured_at == "2024-09-04T12:00:00Z"
    assert row.market == "moneyline" and row.side == "KC"
    assert row.tier == "FPI +3.2"


def test_a_won_benchmark_call_stakes_nothing() -> None:
    """Right about the game, paid nothing: FPI quotes no price to be paid at."""
    row = entries_from_fpi([fpi_game()], {"BAL @ KC": (27, 20)})[0]
    assert row.result == "win"
    assert row.pnl == 0.0
    graded = grade(entries_from_fpi([fpi_game()])[0], 27, 20, home="KC")
    assert (graded.result, graded.pnl) == ("win", 0.0)


def test_a_missing_final_leaves_the_call_ungraded() -> None:
    row = entries_from_fpi([fpi_game()], {})[0]
    assert row.result == "" and row.pnl == 0.0


# -- isolation from every engine measurement -----------------------------
def test_benchmark_rows_are_excluded_from_engine_metrics() -> None:
    ours = engine_row(result="loss")
    theirs = entries_from_fpi([fpi_game()], {"BAL @ KC": (27, 20)})
    both = [ours, *theirs]
    for split, alone in zip(tier_metrics(both), tier_metrics([ours]), strict=True):
        assert (split.label, split.n, split.units, split.roi) == (
            alone.label,
            alone.n,
            alone.units,
            alone.roi,
        )
    for split, alone in zip(market_metrics(both), market_metrics([ours]), strict=True):
        assert (split.label, split.n, split.units) == (alone.label, alone.n, alone.units)


def test_benchmark_metrics_are_the_benchmarks_own_row() -> None:
    rows = [
        *entries_from_fpi([fpi_game()], {"BAL @ KC": (27, 20)}),
        *entries_from_fpi([fpi_game(home="SF", away="NYJ")], {"NYJ @ SF": (10, 24)}),
    ]
    bench = benchmark_metrics(rows)
    assert bench is not None
    assert (bench.n, bench.label) == (2, "FPI (benchmark)")
    assert bench.win_pct == pytest.approx(0.5)
    # No price, so nothing was staked and no return can be claimed.
    assert bench.units == 0.0 and bench.roi == 0.0


def test_benchmark_metrics_need_a_graded_call() -> None:
    assert benchmark_metrics(entries_from_fpi([fpi_game()])) is None
    assert benchmark_metrics([engine_row(result="win")]) is None


def test_the_card_never_presents_a_benchmark_row_as_a_play() -> None:
    entries = [engine_row(), *entries_from_fpi([fpi_game(home="SF", away="NYJ")])]
    card = build_card(entries, season=SEASON, week=WEEK)
    assert [p.side for p in card.plays()] == ["KC"]
    assert [g.matchup for g in card.games] == ["BAL @ KC"]


def test_the_clv_sheet_holds_no_benchmark_row() -> None:
    entries = [engine_row(result="win"), *entries_from_fpi([fpi_game()], {"BAL @ KC": (27, 20)})]
    card = build_card(entries, season=SEASON, week=WEEK)
    book = load_workbook(BytesIO(build_workbook(card, entries)))
    clv_rows = [r for r in book["CLV"].iter_rows(min_row=2, values_only=True) if r[0]]
    assert all(FPI not in str(row) for row in clv_rows)
    assert len(clv_rows) == 1


# -- agreement, disagreement, and passing --------------------------------
def test_agreement_and_fade_are_both_reported() -> None:
    entries = [
        engine_row(),
        *entries_from_fpi([fpi_game()]),
        engine_row(matchup="NYJ @ SF", side="NYJ"),
        *entries_from_fpi([fpi_game(home="SF", away="NYJ")]),
    ]
    pairs = {h.matchup: h for h in head_to_head(entries, season=SEASON, week=WEEK)}
    assert pairs["BAL @ KC"].agree and not pairs["BAL @ KC"].contested
    assert pairs["BAL @ KC"].mark() == "agrees"
    assert pairs["NYJ @ SF"].contested and not pairs["NYJ @ SF"].agree
    assert pairs["NYJ @ SF"].mark() == "fade: SF"


def test_a_game_we_passed_still_shows_their_call() -> None:
    entries = [engine_row(screens="edge"), *entries_from_fpi([fpi_game()])]
    pair = head_to_head(entries, season=SEASON, week=WEEK)[0]
    assert pair.ours == "" and pair.theirs == "KC"
    assert pair.mark() == "KC (we passed)"
    assert not pair.agree and not pair.contested


def test_a_game_they_have_no_number_on_shows_ours_alone() -> None:
    pair = head_to_head([engine_row()], season=SEASON, week=WEEK)[0]
    assert pair.ours == "KC" and pair.theirs == ""
    assert pair.mark() == ""


def test_only_moneyline_plays_are_compared() -> None:
    """FPI forecasts a winner, so a spread or total play is not the same claim."""
    entries = [
        engine_row(market="spread", side="KC"),
        *entries_from_fpi([fpi_game()]),
    ]
    pair = head_to_head(entries, season=SEASON, week=WEEK)[0]
    assert pair.ours == "" and pair.theirs == "KC"


# -- what the reader sees ------------------------------------------------
def test_the_card_prints_their_number_and_the_disagreement() -> None:
    entries = [engine_row(side="BAL"), *entries_from_fpi([fpi_game()])]
    card = build_card(entries, season=SEASON, week=WEEK)
    assert card.contested == 1
    markdown = render_markdown(card)
    assert "FPI: KC 58.2% (FPI +3.2) -- fade: KC" in markdown
    assert "backs the other side on 1" in markdown
    assert "FPI" in render_html(card)


def test_the_plays_sheet_carries_their_call_in_its_own_columns() -> None:
    entries = [engine_row(), *entries_from_fpi([fpi_game()])]
    card = build_card(entries, season=SEASON, week=WEEK)
    sheet = load_workbook(BytesIO(build_workbook(card, entries)))["Plays"]
    header = [cell.value for cell in sheet[1]]
    assert header[-4:] == ["FPI", "FPI %", "FPI margin", "Agrees"]
    played = [c.value for c in sheet[2]]
    assert played[-4:] == ["KC", pytest.approx(0.58167), "FPI +3.2", "agrees"]


def test_the_engine_card_is_identical_with_and_without_the_benchmark() -> None:
    """The wall, stated as a test: their number moves no play of ours."""
    ours = [engine_row(), engine_row(market="total", side="over")]
    plain = build_card(ours, season=SEASON, week=WEEK)
    with_bench = build_card([*ours, *entries_from_fpi([fpi_game()])], season=SEASON, week=WEEK)
    assert [(p.side, p.tier, p.ev_fair) for p in plain.plays()] == [
        (p.side, p.tier, p.ev_fair) for p in with_bench.plays()
    ]


# -- reading ESPN --------------------------------------------------------
def _powerindex(projection: float, margin: float) -> dict[str, object]:
    return {
        "stats": [
            {"name": "gameprojection", "value": projection},
            {"name": "teampredptdiff", "value": margin},
        ]
    }


def test_a_projection_is_read_off_the_published_stats(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(espn, "_get", lambda url: _powerindex(58.167, 3.201))
    assert espn.game_projection("401671789", "KC") == (58.167, 3.201)


def test_an_unknown_team_is_not_guessed(monkeypatch: pytest.MonkeyPatch) -> None:
    called = []
    monkeypatch.setattr(espn, "_get", lambda url: called.append(url) or None)
    assert espn.game_projection("401671789", "XXX") is None
    assert called == []


def test_a_failed_or_empty_response_is_no_projection(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(espn, "_get", lambda url: None)
    assert espn.game_projection("401671789", "KC") is None
    monkeypatch.setattr(espn, "_get", lambda url: {"stats": [{"name": "other", "value": 1.0}]})
    assert espn.game_projection("401671789", "KC") is None


def test_one_missing_projection_drops_that_game_only(
    monkeypatch: pytest.MonkeyPatch, tmp_path
) -> None:
    frame = pd.DataFrame(
        [
            {
                "season": SEASON,
                "week": WEEK,
                "gameday": "2024-09-05",
                "home_team": "KC",
                "away_team": "BAL",
                "espn": "401671789",
                "home_score": 27.0,
            },
            {
                "season": SEASON,
                "week": WEEK,
                "gameday": "2024-09-08",
                "home_team": "SF",
                "away_team": "NYJ",
                "espn": "401671800",
                "home_score": 24.0,
            },
        ]
    )
    monkeypatch.setattr(espn.nflverse, "games", lambda: frame)
    monkeypatch.setattr(espn, "cache_dir", lambda: tmp_path)
    monkeypatch.setattr(
        espn,
        "game_projection",
        lambda event_id, home: (58.167, 3.201) if event_id == "401671789" else None,
    )
    games = espn.projections(SEASON, WEEK)
    assert [g.matchup for g in games] == ["BAL @ KC"]


def test_a_game_without_an_espn_id_is_skipped(monkeypatch: pytest.MonkeyPatch, tmp_path) -> None:
    frame = pd.DataFrame(
        [
            {
                "season": SEASON,
                "week": WEEK,
                "gameday": "2024-09-05",
                "home_team": "KC",
                "away_team": "BAL",
                "espn": None,
                "home_score": 27.0,
            }
        ]
    )
    monkeypatch.setattr(espn.nflverse, "games", lambda: frame)
    monkeypatch.setattr(espn, "cache_dir", lambda: tmp_path)
    monkeypatch.setattr(
        espn, "game_projection", lambda event_id, home: pytest.fail("asked ESPN for a missing id")
    )
    assert espn.projections(SEASON, WEEK) == []


def test_a_played_week_is_read_once_and_then_cached(
    monkeypatch: pytest.MonkeyPatch, tmp_path
) -> None:
    frame = pd.DataFrame(
        [
            {
                "season": SEASON,
                "week": WEEK,
                "gameday": "2024-09-05",
                "home_team": "KC",
                "away_team": "BAL",
                "espn": "401671789",
                "home_score": 27.0,
            }
        ]
    )
    monkeypatch.setattr(espn.nflverse, "games", lambda: frame)
    monkeypatch.setattr(espn, "cache_dir", lambda: tmp_path)
    calls: list[str] = []

    def once(event_id: str, home: str) -> tuple[float, float]:
        calls.append(event_id)
        return (58.167, 3.201)

    monkeypatch.setattr(espn, "game_projection", once)
    first = espn.projections(SEASON, WEEK)
    second = espn.projections(SEASON, WEEK)
    assert calls == ["401671789"]
    assert first == second
    cached = json.loads((tmp_path / "espn" / f"fpi_{SEASON}_wk{WEEK}.json").read_text())
    assert cached[0]["home_prob"] == 58.167


# -- the command ---------------------------------------------------------
def _bench_args(**overrides: object) -> argparse.Namespace:
    base: dict[str, object] = {"season": SEASON, "week": WEEK, "write": True}
    base.update(overrides)
    return argparse.Namespace(**base)


def test_the_command_writes_benchmark_rows_and_touches_no_engine_row(
    monkeypatch: pytest.MonkeyPatch, tmp_path, capsys
) -> None:
    path = tmp_path / "nfl_ledger.csv"
    ours = engine_row()
    save_ledger(path, [ours])
    before = asdict(ours)
    monkeypatch.setattr(cli, "ledger_path", lambda: path)
    monkeypatch.setattr(cli.espn, "projections", lambda season, week: [fpi_game()])
    monkeypatch.setattr(cli, "_final_scores", lambda season: {("BAL @ KC", "2024-09-05"): (27, 20)})
    assert cli.cmd_benchmark(_bench_args()) == 0

    held = load_ledger(path)
    kept = [e for e in held if e.source == ENGINE]
    theirs = [e for e in held if e.source == FPI]
    assert [asdict(e) for e in kept] == [before]
    assert len(theirs) == 1 and theirs[0].result == "win" and theirs[0].pnl == 0.0
    assert "display only, never an input" in capsys.readouterr().out


def test_a_second_capture_keeps_the_first_read(
    monkeypatch: pytest.MonkeyPatch, tmp_path, capsys
) -> None:
    """The read of record is the one taken first, as with our own prices."""
    path = tmp_path / "nfl_ledger.csv"
    monkeypatch.setattr(cli, "ledger_path", lambda: path)
    monkeypatch.setattr(cli, "_final_scores", lambda season: {})
    monkeypatch.setattr(cli.espn, "projections", lambda season, week: [fpi_game(home_prob=58.167)])
    cli.cmd_benchmark(_bench_args())
    monkeypatch.setattr(cli.espn, "projections", lambda season, week: [fpi_game(home_prob=71.0)])
    cli.cmd_benchmark(_bench_args())
    rows = load_ledger(path)
    assert [r.model_prob for r in rows] == [pytest.approx(0.58167)]
    assert "0 new benchmark rows" in capsys.readouterr().out


def test_no_projection_writes_nothing(monkeypatch: pytest.MonkeyPatch, tmp_path, capsys) -> None:
    path = tmp_path / "nfl_ledger.csv"
    monkeypatch.setattr(cli, "ledger_path", lambda: path)
    monkeypatch.setattr(cli.espn, "projections", lambda season, week: [])
    monkeypatch.setattr(
        cli, "_final_scores", lambda season: pytest.fail("read finals with nothing to grade")
    )
    assert cli.cmd_benchmark(_bench_args()) == 0
    assert not path.exists()
    assert "no FPI projections" in capsys.readouterr().out


def test_espn_being_down_costs_the_benchmark_not_the_week(
    monkeypatch: pytest.MonkeyPatch, capsys
) -> None:
    def refused(season: int, week: int) -> list[espn.FpiGame]:
        raise OSError("connection reset by peer")

    monkeypatch.setattr(cli.espn, "projections", refused)
    assert cli._benchmark_step(_bench_args()) == 0
    assert "FPI unavailable" in capsys.readouterr().out
