"""The price band inside the pipeline, and one ordering across every output."""

from __future__ import annotations

from datetime import date
from pathlib import Path

import numpy as np
import pytest
from openpyxl import load_workbook

from cfb_engine.config import Config
from cfb_engine.data.cfbd import CFBDClient
from cfb_engine.features.adjustments import Adjustment
from cfb_engine.market.board import GameOdds
from cfb_engine.market.confidence import MatchupSignal
from cfb_engine.market.ev import MarketQuote
from cfb_engine.market.ordering import order_buys
from cfb_engine.market.priceband import LONG_GATE, SHORT_GATE
from cfb_engine.market.tiers import Tier
from cfb_engine.models.montecarlo import GameSimResult
from cfb_engine.output.card import _best_bets
from cfb_engine.output.excel import write_workbook
from cfb_engine.pipeline import Pipeline, _GameCtx
from cfb_engine.recommendations import Recommendation
from cfb_engine.schemas import Game, TeamGameInfo

DAY = date(2026, 9, 5)


def _ctx(home_win_prob: float) -> _GameCtx:
    """A sim that wins exactly ``home_win_prob`` of its margins."""
    wins = int(round(home_win_prob * 1000))
    margins = np.array([7.0] * wins + [-7.0] * (1000 - wins))
    sim = GameSimResult(margins=margins, totals=np.full(1000, 52.0))
    game = Game(
        game_id="1",
        game_date=DAY,
        home=TeamGameInfo(name="Alabama", abbrev="ALA", is_home=True),
        away=TeamGameInfo(name="Georgia", abbrev="UGA", is_home=False),
    )
    return _GameCtx(game, sim, Adjustment(), MatchupSignal())


def _ml_board(home_american: float, away_american: float) -> GameOdds:
    odds = GameOdds(matchup="UGA @ ALA")
    odds.add_ml("ALA", MarketQuote("b", home_american, opposite_american=away_american))
    odds.add_ml("UGA", MarketQuote("b", away_american, opposite_american=home_american))
    return odds


def _pipeline(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Pipeline:
    monkeypatch.setenv("CFBE_DATA_DIR", str(tmp_path))
    monkeypatch.setenv("CFBE_MARKET_ANCHOR", "0")  # price the model, not the market
    monkeypatch.setenv("CFBE_SHRINK_TAILS", "0")  # isolate the band from tail shrink
    return Pipeline(Config(), cfbd=CFBDClient(None))


def _short_favourite(pipe: Pipeline) -> Recommendation:
    """Alabama at -400 with the model 6 points above the price -- a buy the band
    should be able to refuse for its price alone."""
    recs = pipe._price_ml(_ctx(0.82), _ml_board(-400.0, 300.0))
    return next(r for r in recs if r.selection.startswith("ALA"))


def _arm_short_tail(monkeypatch: pytest.MonkeyPatch) -> None:
    """Put the moneyline's short tail back in play.

    The shipped moneyline band is long-side only, so a test about short prices
    has to say so rather than inherit it.
    """
    monkeypatch.setenv("CFBE_PRICE_MIN_GAME_ML", "-250")


def test_a_short_price_is_only_annotated_while_the_band_is_off(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _arm_short_tail(monkeypatch)
    monkeypatch.setenv("CFBE_PRICE_BAND_GAME_ML", "0")
    rec = _short_favourite(_pipeline(tmp_path, monkeypatch))
    assert rec.tier != Tier.PASS
    assert rec.pass_gate is None
    assert any("band off, measuring" in r for r in rec.reasons)


def test_an_armed_band_passes_the_row_and_names_the_gate(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setenv("CFBE_PRICE_BAND", "1")
    _arm_short_tail(monkeypatch)
    rec = _short_favourite(_pipeline(tmp_path, monkeypatch))
    assert rec.tier == Tier.PASS
    assert rec.pass_gate == SHORT_GATE


def test_a_band_armed_for_the_moneyline_leaves_the_spread_alone(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setenv("CFBE_PRICE_BAND_GAME_ML", "1")
    _arm_short_tail(monkeypatch)
    pipe = _pipeline(tmp_path, monkeypatch)
    assert _short_favourite(pipe).pass_gate == SHORT_GATE
    assert pipe.price_band.for_market("game_ats").enabled is False


def test_the_long_end_bites_the_dog_not_the_favourite(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setenv("CFBE_PRICE_BAND", "1")
    _arm_short_tail(monkeypatch)
    pipe = _pipeline(tmp_path, monkeypatch)
    recs = pipe._price_ml(_ctx(0.71), _ml_board(-160.0, 400.0))
    dog = next(r for r in recs if r.selection.startswith("UGA"))
    fav = next(r for r in recs if r.selection.startswith("ALA"))
    assert dog.pass_gate == LONG_GATE
    assert fav.pass_gate is None


def test_a_long_moneyline_dog_is_refused_with_nothing_configured(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """The shipped default, and the row that prompted it: a +260 dog whose only
    support is model-versus-price -- the input MLB measured at AUC 0.33."""
    pipe = _pipeline(tmp_path, monkeypatch)
    recs = pipe._price_ml(_ctx(0.68), _ml_board(-300.0, 260.0))
    dog = next(r for r in recs if r.selection.startswith("UGA"))
    fav = next(r for r in recs if r.selection.startswith("ALA"))
    assert (dog.tier, dog.pass_gate) == (Tier.PASS, LONG_GATE)
    # The short tail ships disarmed, so the favourite is neither refused nor
    # annotated -- the ledger's own odds column is what grades that half.
    assert fav.pass_gate is None
    assert not any("shorter than" in r for r in fav.reasons)


def test_the_shipped_moneyline_band_is_switched_off_per_market(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setenv("CFBE_PRICE_BAND_GAME_ML", "0")
    pipe = _pipeline(tmp_path, monkeypatch)
    dog = next(
        r
        for r in pipe._price_ml(_ctx(0.68), _ml_board(-300.0, 260.0))
        if r.selection.startswith("UGA")
    )
    assert dog.pass_gate is None
    assert any("band off, measuring" in r for r in dog.reasons)


def test_a_row_the_tiers_already_passed_is_not_reattributed_to_the_band(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Attribution has to stay honest: the band only owns rows it alone refused,
    or probation grades it on bets it never touched."""
    monkeypatch.setenv("CFBE_PRICE_BAND", "1")
    _arm_short_tail(monkeypatch)
    pipe = _pipeline(tmp_path, monkeypatch)
    # Model agrees with a -400 price, so the EV screen passes it on its own.
    rec = next(
        r
        for r in pipe._price_ml(_ctx(0.77), _ml_board(-400.0, 300.0))
        if r.selection.startswith("ALA")
    )
    assert rec.tier == Tier.PASS
    assert rec.pass_gate is None


# -- one ordering everywhere ---------------------------------------------


def _priced(selection: str, american: float, ev: float, edge: float) -> Recommendation:
    return Recommendation(
        game_date=DAY,
        game_id="g1",
        matchup="UGA @ ALA",
        market="game_ml",
        selection=selection,
        model_prob=0.5,
        market_american=american,
        ev=ev,
        edge=edge,
        fair_prob=0.45,
        tier=Tier.STRONG,
    )


def test_pipeline_card_and_workbook_all_run_the_same_order(tmp_path: Path) -> None:
    """The old code sorted in three places on two different keys, so the Excel
    tab and the article disagreed about which bet was strongest."""
    dog = _priced("UGA ML", 400.0, 0.30, 0.06)
    fav = _priced("ALA -160", -160.0, 0.12, 0.08)
    recs = [dog, fav]

    expected = [r.selection for r in order_buys(recs)]
    assert expected == [r.selection for r in _best_bets(recs)]

    path = write_workbook(recs, tmp_path / "card.xlsx", DAY)
    ws = load_workbook(path)["Strong Buys"]
    header = [c.value for c in ws[1]]
    assert "Kelly" in header
    col = header.index("Selection") + 1
    assert [ws.cell(row=r, column=col).value for r in (2, 3)] == expected
