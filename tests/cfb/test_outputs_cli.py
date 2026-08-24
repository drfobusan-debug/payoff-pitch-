"""Workbook generation and CLI wiring."""

from __future__ import annotations

from datetime import date

import pytest
from openpyxl import load_workbook

from cfb_engine.audit.ledger import daily_rollup, entries_from_graded, overall_metrics
from cfb_engine.audit.priced import engine_priced_stat
from cfb_engine.audit.probation import market_probation
from cfb_engine.cli import _build_parser, main
from cfb_engine.market.tiers import Tier
from cfb_engine.output.audit_report import build_audit_article
from cfb_engine.output.excel import write_ledger_workbook, write_workbook
from cfb_engine.recommendations import Recommendation, load_json, save_json

DAY = date(2025, 11, 1)


def _rec(tier: Tier, market: str = "game_ml") -> Recommendation:
    return Recommendation(
        game_date=DAY,
        game_id="g1",
        matchup="Alabama vs Georgia",
        market=market,
        selection="Georgia ML",
        model_prob=0.6,
        market_american=-120,
        ev=0.05,
        edge=0.04,
        fair_prob=0.55,
        tier=tier,
        home_abbrev="Georgia",
        away_abbrev="Alabama",
        team_side="home",
        side="win",
    )


def test_write_workbook_has_expected_tabs(tmp_path):
    recs = [_rec(Tier.STRONG), _rec(Tier.PASS, "game_total")]
    out = write_workbook(recs, tmp_path / "cfb.xlsx", DAY)
    wb = load_workbook(out)
    for tab in ("Strong Buys", "Moderate Buys", "Fades", "All"):
        assert tab in wb.sheetnames


def test_write_ledger_workbook(tmp_path):
    graded = [(_rec(Tier.STRONG), "win"), (_rec(Tier.MODERATE), "loss")]
    entries = entries_from_graded(graded, DAY)
    out = write_ledger_workbook(
        entries, overall_metrics(entries), daily_rollup(entries), tmp_path / "ledger.xlsx"
    )
    wb = load_workbook(out)
    assert "Overall" in wb.sheetnames
    assert "Bets" in wb.sheetnames


def test_money_and_probation_reach_the_workbook(tmp_path):
    graded = [(_rec(Tier.STRONG), "win"), (_rec(Tier.MODERATE), "loss")]
    entries = entries_from_graded(graded, DAY)
    out = write_ledger_workbook(
        entries,
        overall_metrics(entries),
        daily_rollup(entries),
        tmp_path / "ledger.xlsx",
        money_rows=[engine_priced_stat(entries)],
        probation_rows=market_probation(entries),
    )
    wb = load_workbook(out)
    assert "Money (priced buys)" in wb.sheetnames
    assert "Probation" in wb.sheetnames
    money = wb["Money (priced buys)"]
    header = [c.value for c in money[1]]
    assert "Needs" in header and "ROI" in header
    assert money.cell(row=2, column=header.index("N") + 1).value == 2
    verdict = wb["Probation"]
    assert verdict.cell(row=2, column=1).value == "WATCHING"


def test_the_audit_article_says_what_the_prices_demanded():
    """A reader who only hears the win rate cannot tell a winner from a loser."""
    graded = [(_rec(Tier.STRONG), "win"), (_rec(Tier.MODERATE), "loss")]
    entries = entries_from_graded(graded, DAY)
    stat = engine_priced_stat(entries)
    html, narration = build_audit_article(
        DAY,
        overall_metrics(entries),
        [],
        2,
        money_rows=[stat],
        probation=["game_ml: shut game_ml until the refit"],
    )
    assert "What the prices did" in html
    assert "shut game_ml until the refit" in html
    assert "where they needed" in narration


def test_the_audit_article_says_so_when_nothing_crossed_the_bar():
    html, _ = build_audit_article(DAY, [], [], 0)
    assert "no market or screen is on probation" in html


def test_predictions_json_roundtrip(tmp_path):
    recs = [_rec(Tier.STRONG), _rec(Tier.MODERATE, "game_ats")]
    path = tmp_path / "preds.json"
    save_json(recs, path)
    back = load_json(path)
    assert len(back) == 2
    assert back[0].tier == Tier.STRONG
    assert back[0].game_date == DAY


def test_cli_help_exits_zero():
    with pytest.raises(SystemExit) as exc:
        main(["--help"])
    assert exc.value.code == 0


def test_cli_parser_has_all_commands():
    parser = _build_parser()
    ns = parser.parse_args(["run"])
    assert ns.command == "run"
    for cmd in ("card", "close", "audit", "report", "calibrate", "probation"):
        assert parser.parse_args([cmd]).command == cmd
    assert parser.parse_args(["probation", "--since", "2025-09-01"]).since == "2025-09-01"
